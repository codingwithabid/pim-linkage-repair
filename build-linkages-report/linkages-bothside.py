# """
# Walk every VERSION of every variant in both live + draft files, detect link
# problems across version history, then aggregate the findings into a
# product-side view.

# Six variant-side issue patterns are detected:

#   1. STALE_LATEST_EMPTY        — parent_id appeared in an earlier version but
#                                  the latest version's parent_id is empty.
#                                  (e.g. versions: P1, -, P1, -  → latest is empty)
#   2. PARENT_CHANGED            — variant was reparented across versions AND the
#                                  OLD parent still exists AND still lists the
#                                  variant in its variant_id (genuine ownership
#                                  conflict — two products both claim V).
#                                  Clean reparenting (old parent properly updated)
#                                  is normal lifecycle and not flagged.
#   3. ORPHAN_BUT_REFERENCED     — variant has no parent_id in any version, but
#                                  some product references it.
#   4. PARENT_NOT_FOUND          — A parent the variant has ever pointed at
#                                  (latest OR any historical version) doesn't
#                                  exist as a PRODUCT anywhere in the catalog.
#   5. PARENT_DOES_NOT_LINK_BACK — latest parent exists but its variant_id list
#                                  doesn't include this variant.
#   6. NO_LINK_ANYWHERE          — no parent_id in any version AND no product
#                                  references this variant anywhere.

# Output workbook (5 sheets):

#   Product-side summary   one row per affected product:
#                          product_id | missing_variants | attached_variants
#                                     | total_variants  | issue_summary
#   Variant details         one row per variant with at least one issue:
#                           variant_id | latest_parent | all_versions_parent
#                                      | issues | claimed_by_products
#   Definitions             explanation of each issue type
#   Inventory               counts of products/variants per source
#   Summary                 overall counts by issue category

# Usage:
#     python variant_version_scan.py <live.csv> [--draft <draft.csv>] [-o out.xlsx]
# """

# import argparse
# import ast
# import csv
# import re
# import sys
# from collections import defaultdict

# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.utils import get_column_letter


# # ---------------- helpers ----------------

# def parse_id_list(value) -> list:
#     if value is None:
#         return []
#     s = str(value).strip()
#     if s == "" or s in ("[]", "['']", '[""]', "null", "None", "—"):
#         return []
#     if s.startswith("[") and s.endswith("]"):
#         try:
#             parsed = ast.literal_eval(s)
#             if isinstance(parsed, (list, tuple)):
#                 return [str(x).strip() for x in parsed if str(x).strip()]
#         except (ValueError, SyntaxError):
#             pass
#         inner = s[1:-1]
#         parts = re.split(r"\s*,\s*", inner)
#         return [p.strip().strip("'\"") for p in parts if p.strip().strip("'\"")]
#     parts = re.split(r"\s*,\s*", s)
#     return [p.strip().strip("'\"") for p in parts if p.strip().strip("'\"")]


# import json as _json


# def _normalize_whitespace_escapes_outside_strings(s: str) -> str:
#     """Walk through a JSON-ish string. When OUTSIDE a JSON string value, convert
#     literal `\\n` / `\\t` / `\\r` escape sequences to real whitespace (they were
#     meant as inter-token whitespace by an over-escaping writer). When INSIDE a
#     string value, leave escape sequences alone so json.loads handles them.

#     This is a safer replacement for blanket s.replace('\\n', '\\n') which mangles
#     real JSON escape sequences inside string values."""
#     out = []
#     i = 0
#     n = len(s)
#     in_string = False
#     while i < n:
#         c = s[i]
#         if not in_string:
#             if c == '\\' and i + 1 < n:
#                 nxt = s[i + 1]
#                 if nxt == 'n':
#                     out.append('\n'); i += 2; continue
#                 if nxt == 't':
#                     out.append('\t'); i += 2; continue
#                 if nxt == 'r':
#                     out.append('\r'); i += 2; continue
#                 if nxt == '\\':
#                     # Two literal backslashes between tokens — keep one
#                     out.append('\\'); i += 2; continue
#             if c == '"':
#                 in_string = True
#             out.append(c)
#             i += 1
#         else:
#             # Inside string: preserve all escape sequences as-is
#             if c == '\\' and i + 1 < n:
#                 out.append(c); out.append(s[i + 1]); i += 2; continue
#             if c == '"':
#                 in_string = False
#             out.append(c)
#             i += 1
#     return ''.join(out)


# def _repair_unescaped_quotes(s: str) -> str:
#     """Walk through a JSON-ish string and escape any quote that's inside a
#     string value but isn't followed by JSON-structural characters. Handles the
#     common case where the source writer escaped quotes for CSV but forgot to
#     escape them for JSON, producing lone unescaped quotes inside string values.

#     Examples this fixes:
#       - "2.5" Gel Visco"  →  "2.5\\" Gel Visco"
#       - "2'3\\" x 8'"     →  "2'3\\\\\\" x 8'" (already-broken double-backslash + quote)
#     """
#     out = []
#     i = 0
#     n = len(s)
#     in_string = False
#     while i < n:
#         c = s[i]
#         if c == '\\' and i + 1 < n:
#             # Pass through any escape sequence untouched
#             out.append(c)
#             out.append(s[i + 1])
#             i += 2
#             continue
#         if c == '"':
#             if not in_string:
#                 in_string = True
#                 out.append(c)
#                 i += 1
#                 continue
#             # Inside a string. Decide if this is the real closing quote by
#             # peeking ahead past whitespace for a JSON-structural character.
#             j = i + 1
#             while j < n and s[j] in ' \t\n\r':
#                 j += 1
#             if j == n or s[j] in ',:}]':
#                 # Real closing quote
#                 in_string = False
#                 out.append(c)
#             else:
#                 # Content quote — escape it
#                 out.append('\\')
#                 out.append(c)
#             i += 1
#             continue
#         out.append(c)
#         i += 1
#     return ''.join(out)


# def _safe_json_loads(s: str):
#     """Try to parse s as JSON, applying several recovery strategies on failure.
#     Uses strict=False so control characters inside string values don't break parsing."""
#     try:
#         return _json.loads(s, strict=False)
#     except (ValueError, TypeError):
#         pass
#     # Strategy 1: stateful normalization of inter-token whitespace escapes
#     cleaned = _normalize_whitespace_escapes_outside_strings(s)
#     if cleaned != s:
#         try:
#             return _json.loads(cleaned, strict=False)
#         except (ValueError, TypeError):
#             pass
#     # Strategy 2: repair lone unescaped quotes inside string values
#     repaired = _repair_unescaped_quotes(s)
#     if repaired != s:
#         try:
#             return _json.loads(repaired, strict=False)
#         except (ValueError, TypeError):
#             pass
#     # Strategy 3: combine — repair quotes on the whitespace-normalized form
#     repaired_clean = _repair_unescaped_quotes(cleaned)
#     if repaired_clean != cleaned:
#         try:
#             return _json.loads(repaired_clean, strict=False)
#         except (ValueError, TypeError):
#             pass
#     return None


# def parse_template_id(raw: str) -> str:
#     """Extract id from template_id column, which looks like:
#        {"id":"3956b34b-...","version":12}"""
#     if not raw:
#         return ""
#     s = raw.strip()
#     if not s or s in ("null", "None"):
#         return ""
#     obj = _safe_json_loads(s)
#     if isinstance(obj, dict) and "id" in obj:
#         return str(obj["id"]).strip()
#     return ""


# def _extract_en_us_displayname_via_regex(s: str) -> str:
#     """Last-resort: find "displayName" : { ... "en_us" : "VALUE" ... } using a
#     regex, when full JSON parsing fails. Uses a permissive value capture that
#     only treats a quote as the closing quote if it's followed by JSON
#     structural characters (`,` `}` `]` with optional whitespace, where
#     whitespace can also be `\\n`, `\\t`, `\\r` escape sequences)."""
#     m = re.search(
#         r'"displayName"[\s\S]{0,500}?"localizations"[\s\S]{0,300}?"en_us"\s*:\s*"',
#         s,
#     )
#     if not m:
#         return ""
#     start = m.end()
#     i = start
#     n = len(s)
#     while i < n:
#         c = s[i]
#         if c == '\\' and i + 1 < n:
#             i += 2
#             continue
#         if c == '"':
#             # Peek ahead past whitespace (real or escape sequences) for a JSON delimiter
#             j = i + 1
#             while j < n:
#                 if s[j] in ' \t\n\r':
#                     j += 1
#                     continue
#                 if s[j] == '\\' and j + 1 < n and s[j + 1] in 'ntr':
#                     j += 2
#                     continue
#                 break
#             if j == n or s[j] in ',}]':
#                 val = s[start:i]
#                 # Cleanup escapes — collapse `\\` pairs FIRST, then handle remaining `\"`
#                 val = (val
#                        .replace('\\\\', '\\')      # 2 backslashes → 1
#                        .replace('\\"', '"')        # backslash-quote → quote
#                        .replace("\\'", "'")
#                        .replace('\\n', ' ')
#                        .replace('\\t', ' ')
#                        .replace('\\r', ' '))
#                 val = " ".join(val.split())
#                 return val
#             i += 1
#             continue
#         i += 1
#     return ""


# def parse_product_name(raw: str) -> str:
#     """Extract displayName.localizations.en_us from the request_schema JSON column."""
#     if not raw:
#         return ""
#     s = raw.strip()
#     if not s or s in ("null", "None"):
#         return ""
#     obj = _safe_json_loads(s)
#     if isinstance(obj, dict):
#         display = obj.get("displayName") or {}
#         if isinstance(display, dict):
#             locs = display.get("localizations") or {}
#             if isinstance(locs, dict):
#                 val = locs.get("en_us")
#                 if val:
#                     cleaned = str(val)
#                     # Order matters: collapse `\\` pairs FIRST, then handle remaining `\"`
#                     cleaned = (cleaned
#                                .replace('\\\\', '\\')
#                                .replace('\\"', '"')
#                                .replace("\\'", "'")
#                                .replace("\\r\\n", " ")
#                                .replace("\\n", " ")
#                                .replace("\\t", " ")
#                                .replace("\\r", " "))
#                     cleaned = " ".join(cleaned.split())
#                     return cleaned
#     # JSON parsing failed or no displayName found — try regex extraction
#     val = _extract_en_us_displayname_via_regex(s)
#     return val


# def _looks_like_backslash_escaped(path: str) -> bool:
#     """Peek at the first ~10KB of body. If we see `\\"` inside a quoted field,
#     assume backslash-escape format."""
#     try:
#         with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
#             f.readline()  # skip header
#             sample = f.read(10000)
#         return '\\"' in sample
#     except OSError:
#         return False


# def _normalize_backslash_escaped_csv(path: str) -> "io.StringIO":
#     """
#     Read a CSV that uses backslash-escaped quotes inside quoted fields and
#     return a StringIO buffer with standard CSV escaping (doubled quotes).

#     Rules:
#       - `\\"` inside a quoted field  -> `""` (CSV doubled-quote)
#       - `\\\\` (literal backslash)   -> single `\\`
#       - everything else passes through unchanged (so embedded \\n stays as \\n
#         and JSON parsers see whitespace, not the letter 'n')

#     We track whether we're inside a quoted field by toggling on `"` chars,
#     handling the doubled-quote escape afterward.
#     """
#     import io
#     out = []
#     with open(path, "r", encoding="utf-8-sig", newline="") as f:
#         content = f.read()

#     in_quotes = False
#     i = 0
#     n = len(content)
#     while i < n:
#         c = content[i]
#         if c == '"':
#             # Check whether the previous char was a backslash inside quotes
#             # (we should never see this case because we normalize backslash-escapes
#             # below — so a literal " always opens/closes a field).
#             out.append('"')
#             in_quotes = not in_quotes
#             i += 1
#             continue
#         if c == '\\' and in_quotes and i + 1 < n:
#             nxt = content[i + 1]
#             if nxt == '"':
#                 # \"  ->  ""  (CSV doubled-quote escape)
#                 out.append('""')
#                 i += 2
#                 continue
#             if nxt == '\\':
#                 # \\  ->  \  (literal backslash)
#                 out.append('\\')
#                 i += 2
#                 continue
#             # Other escapes (\n, \t, \uXXXX, ...) inside JSON-bearing fields
#             # should be PASSED THROUGH untouched so json.loads sees them.
#             out.append(c)
#             i += 1
#             continue
#         out.append(c)
#         i += 1

#     return io.StringIO("".join(out))


# def load_file(path: str, source_label: str):
#     """
#     Reads ALL versions, not just the latest. Returns:
#         product_versions: dict[item_id] -> list of {version, parent_id_raw,
#                                                     variant_id_raw, state, ...}
#         variant_versions: dict[item_id] -> list of the same per-version dicts
#     """
#     product_versions = defaultdict(list)
#     variant_versions = defaultdict(list)

#     # Pre-process if needed
#     use_backslash = _looks_like_backslash_escaped(path)
#     if use_backslash:
#         f = _normalize_backslash_escaped_csv(path)
#         first_line = f.readline()
#         f.seek(0)
#         dialect = csv.excel  # standard now
#     else:
#         f = open(path, "r", encoding="utf-8-sig", newline="")
#         first_line = f.readline()
#         f.seek(0)
#         dialect = csv.excel_tab if "\t" in first_line else csv.excel

#     try:
#         reader = csv.DictReader(f, dialect=dialect)
#         if "item_id" not in (reader.fieldnames or []):
#             sys.exit(f"ERROR: {path} missing 'item_id' column (got {reader.fieldnames})")
#         has_version = "version" in (reader.fieldnames or [])
#         for row in reader:
#             item_id = (row.get("item_id") or "").strip()
#             scope = (row.get("item_scope") or "").strip().upper()
#             if not item_id or scope not in ("PRODUCT", "VARIANT"):
#                 continue
#             try:
#                 version = int(row["version"]) if has_version else 1
#             except (ValueError, TypeError):
#                 version = 1
#             entry = {
#                 "version": version,
#                 "source": source_label,
#                 "scope": scope,
#                 "state": row.get("state", ""),
#                 "site_ids": row.get("site_ids", ""),
#                 "sku": row.get("sku", ""),
#                 "date_modified": row.get("date_modified", ""),
#                 "parent_ids": parse_id_list(row.get("parent_id", "")),
#                 "variant_ids": parse_id_list(row.get("variant_id", "")),
#                 "template_id": parse_template_id(row.get("template_id", "")),
#                 "product_name": parse_product_name(row.get("request_schema", "")),
#             }
#             if scope == "PRODUCT":
#                 product_versions[item_id].append(entry)
#             else:
#                 variant_versions[item_id].append(entry)
#     finally:
#         try:
#             f.close()
#         except Exception:
#             pass
#     return product_versions, variant_versions


# def merge_version_dicts(*dicts):
#     """Combine the same item's version lists from multiple sources."""
#     merged = defaultdict(list)
#     for d in dicts:
#         for k, v_list in d.items():
#             merged[k].extend(v_list)
#     # Sort each item's versions: by source ('live' first), then by version
#     for k in merged:
#         merged[k].sort(key=lambda e: (0 if e["source"] == "live" else 1, e["version"]))
#     return merged


# def build_variant_meta(variant_versions):
#     """Return dict[variant_id] -> {"name": str, "template_id": str} for every
#     variant.

#     Name/template_id selection rules:
#       - Prefer LIVE: take the latest live version's value if non-empty.
#       - Fall back to DRAFT: if live is missing or has an empty value, use the
#         latest draft version's value.
#     """
#     meta = {}
#     for vid, versions in variant_versions.items():
#         if not versions:
#             continue
#         name = pick_live_first(versions, "product_name")
#         tid = pick_live_first(versions, "template_id")
#         meta[vid] = {"name": name, "template_id": tid}
#     return meta


# def pick_live_first(versions, field):
#     """Pick the value of `field` preferring live (latest live version with a
#     non-empty value), falling back to draft if live is empty/missing."""
#     if not versions:
#         return ""
#     live_rows = [v for v in versions if v.get("source") == "live"]
#     draft_rows = [v for v in versions if v.get("source") == "draft"]

#     def latest_nonempty(rows):
#         # Sort newest first and return the first non-empty value
#         try:
#             ordered = sorted(rows, key=lambda r: r.get("version", 0), reverse=True)
#         except TypeError:
#             ordered = list(reversed(rows))
#         for r in ordered:
#             val = r.get(field)
#             if val:
#                 return val
#         return ""

#     return latest_nonempty(live_rows) or latest_nonempty(draft_rows)


# def pick_latest_parent_per_variant(variant_versions, product_latest=None,
#                                    product_versions=None):
#     """
#     For each variant_id, decide its intended parent set using these rules:

#     1. If the absolute-latest version has a non-empty parent_id, use that.
#     2. Else if any earlier version had a non-empty parent_id, fall back to the
#        highest-version row that did.
#     3. Else (the variant has NEVER claimed a parent in any version), fall back
#        to whichever products CURRENTLY claim this variant in their variant_id
#        list (product_latest). This is the "orphan-but-referenced" rescue.
#     4. Else fall back to whichever products HISTORICALLY claimed this variant
#        in any version of their variant_id list (product_versions). This is the
#        "both sides forgot" rescue — the variant exists, an old product version
#        listed it, the link should still hold even though both sides went quiet.
#     5. Else return empty (truly orphan, no claimers anywhere).

#     Returns:
#         dict[variant_id] -> {"latest_parents": set, "latest_version": int,
#                              "latest_source": str, "from_fallback": bool,
#                              "from_product_claim": bool,
#                              "from_historical_product_claim": bool}
#     """
#     # Build a referrers index from product_latest: variant_id -> set of products
#     referrers_latest = defaultdict(set)
#     if product_latest:
#         for pid, info in product_latest.items():
#             for vid in info.get("variant_ids", set()):
#                 referrers_latest[vid].add(pid)
#     # Historical: any version of any product that listed this variant
#     referrers_historical = defaultdict(set)
#     if product_versions:
#         for pid, versions in product_versions.items():
#             for v in versions:
#                 for vid in v["variant_ids"]:
#                     referrers_historical[vid].add(pid)

#     out = {}
#     for vid, versions in variant_versions.items():
#         if not versions:
#             continue

#         try:
#             absolute_latest = max(versions, key=lambda v: v["version"])
#         except TypeError:
#             absolute_latest = versions[-1]

#         from_fallback = False
#         from_product_claim = False
#         from_historical_product_claim = False

#         if absolute_latest["parent_ids"]:
#             chosen_parents = set(absolute_latest["parent_ids"])
#             chosen_version = absolute_latest["version"]
#             chosen_source = absolute_latest["source"]
#         else:
#             # Try fallback: highest-version row with non-empty parent_ids
#             with_parent = [v for v in versions if v["parent_ids"]]
#             if with_parent:
#                 try:
#                     fb = max(with_parent, key=lambda v: v["version"])
#                 except TypeError:
#                     fb = with_parent[-1]
#                 chosen_parents = set(fb["parent_ids"])
#                 chosen_version = fb["version"]
#                 chosen_source = fb["source"]
#                 from_fallback = True
#             else:
#                 # No version ever had a parent — try product claims
#                 current_claimers = referrers_latest.get(vid, set())
#                 historical_claimers = referrers_historical.get(vid, set())
#                 if current_claimers:
#                     chosen_parents = set(current_claimers)
#                     chosen_version = absolute_latest["version"]
#                     chosen_source = absolute_latest["source"]
#                     from_product_claim = True
#                 elif historical_claimers:
#                     chosen_parents = set(historical_claimers)
#                     chosen_version = absolute_latest["version"]
#                     chosen_source = absolute_latest["source"]
#                     from_historical_product_claim = True
#                 else:
#                     chosen_parents = set()
#                     chosen_version = absolute_latest["version"]
#                     chosen_source = absolute_latest["source"]

#         out[vid] = {
#             "latest_parents": chosen_parents,
#             "latest_version": chosen_version,
#             "latest_source": chosen_source,
#             "from_fallback": from_fallback,
#             "from_product_claim": from_product_claim,
#             "from_historical_product_claim": from_historical_product_claim,
#         }
#     return out


# def latest_product_state(product_versions):
#     """For each product_id, pick the latest version per source and merge.
#     Returns dict[product_id] -> {variant_ids: set (union of live+draft latest),
#                                  latest_per_source: {source: version dict}}"""
#     out = {}
#     for pid, versions in product_versions.items():
#         by_source = defaultdict(list)
#         for v in versions:
#             by_source[v["source"]].append(v)
#         latest_per_source = {}
#         union_variants = set()
#         for src, vs in by_source.items():
#             try:
#                 latest = max(vs, key=lambda x: x["version"])
#             except TypeError:
#                 latest = vs[-1]
#             latest_per_source[src] = latest
#             union_variants.update(latest["variant_ids"])
#         out[pid] = {
#             "variant_ids": union_variants,
#             "latest_per_source": latest_per_source,
#         }
#     return out


# # ---------------- variant-side analysis ----------------

# def analyze_variants(variant_versions, product_latest, product_versions=None):
#     """
#     For each variant, walk all versions across both sources and classify any
#     issues. Returns:
#         variant_findings: dict[variant_id] -> {
#             "latest_parent": str (or "—"),
#             "all_versions_parent": str,
#             "issues": list[str],
#             "issue_codes": set[str],
#             "claimed_by_products": list[str],
#         }
#     """
#     # Build "products that claim this variant" indexes — both latest and historical
#     referrers_latest = defaultdict(set)
#     for pid, info in product_latest.items():
#         for vid in info["variant_ids"]:
#             referrers_latest[vid].add(pid)
#     # Historical referrers: any version of any product that ever listed the variant
#     referrers_ever = defaultdict(set)
#     if product_versions:
#         for pid, versions in product_versions.items():
#             for v in versions:
#                 for vid in v["variant_ids"]:
#                     referrers_ever[vid].add(pid)
#     else:
#         referrers_ever = referrers_latest

#     findings = {}
#     for vid, versions in variant_versions.items():
#         if not versions:
#             continue

#         # Group versions by source and pick the latest per source
#         by_source = defaultdict(list)
#         for v in versions:
#             by_source[v["source"]].append(v)
#         latest_per_source = {}
#         for src, vs in by_source.items():
#             try:
#                 latest_per_source[src] = max(vs, key=lambda x: x["version"])
#             except TypeError:
#                 latest_per_source[src] = vs[-1]

#         # All parent_ids EVER seen across every version (any source)
#         ever_parents = set()
#         for v in versions:
#             ever_parents.update(v["parent_ids"])

#         # Latest parent_ids = union of latest-per-source's parent_ids
#         latest_parents = set()
#         for src_latest in latest_per_source.values():
#             latest_parents.update(src_latest["parent_ids"])

#         claimed_by_latest = sorted(referrers_latest.get(vid, set()))
#         claimed_by_ever = sorted(referrers_ever.get(vid, set()))
#         # For display: prefer current claimers; fall back to historical
#         claimed_by = claimed_by_latest or claimed_by_ever

#         # ----- Classify issues -----
#         issues = []
#         codes = set()

#         # Per-version trace for the "all_versions_parent" column.
#         # Deduplicate by (source, version) to avoid showing identical rows
#         # multiple times. Use ' | ' between version entries and '+' between
#         # parent_ids within one entry so commas inside parent lists don't
#         # collide with the version separator.
#         seen_sv = set()
#         trace_entries = []
#         for v in sorted(versions, key=lambda x: (x["source"], x["version"])):
#             sv = (v["source"], v["version"])
#             if sv in seen_sv:
#                 continue
#             seen_sv.add(sv)
#             parents_str = "+".join(v["parent_ids"]) if v["parent_ids"] else "-"
#             trace_entries.append(f"{v['source'][0]}v{v['version']}={parents_str}")
#         trace = " | ".join(trace_entries)

#         if not ever_parents:
#             # Variant never claimed any parent in any version
#             if claimed_by_latest:
#                 # Some product STILL lists this variant in its current variant_id
#                 issues.append(f"orphan but referenced by {len(claimed_by_latest)} product(s)")
#                 codes.add("ORPHAN_BUT_REFERENCED")
#             elif claimed_by_ever:
#                 # No current product lists it, but an OLDER version of some product did
#                 issues.append(
#                     f"product(s) dropped this variant in latest version: "
#                     f"{', '.join(claimed_by_ever)}"
#                 )
#                 codes.add("PRODUCT_DROPPED_VARIANT")
#             else:
#                 issues.append("no link anywhere — no parent in any version, no product references it")
#                 codes.add("NO_LINK_ANYWHERE")
#         else:
#             # The variant DID have parents at some point
#             if not latest_parents:
#                 # Latest version is empty even though earlier versions had parents.
#                 # BUT: if any historical parent no longer exists as a product, that's
#                 # a stronger signal than just "link cleared" — fire PARENT_NOT_FOUND
#                 # for those, and STALE_LATEST_EMPTY only if no missing historical parents.
#                 missing_historical = {p for p in ever_parents if p not in product_latest}
#                 if missing_historical:
#                     for pid in sorted(missing_historical):
#                         issues.append(
#                             f"parent {pid} (claimed in earlier version) does not exist as a product"
#                         )
#                     codes.add("PARENT_NOT_FOUND")
#                     # If there are also surviving historical parents that still exist,
#                     # also note the stale-clear so the user sees both pieces of context.
#                     surviving = ever_parents - missing_historical
#                     if surviving:
#                         issues.append(
#                             f"latest version is empty; earlier versions also referenced existing parent(s): "
#                             f"{', '.join(sorted(surviving))}"
#                         )
#                         codes.add("STALE_LATEST_EMPTY")
#                 else:
#                     issues.append(
#                         f"latest version is empty but earlier had: {', '.join(sorted(ever_parents))}"
#                     )
#                     codes.add("STALE_LATEST_EMPTY")
#             else:
#                 # Latest has parents — check if any earlier parent was dropped
#                 # AND still claims this variant (genuine ownership conflict).
#                 # A clean reparenting (old parent updated to remove V) is normal
#                 # lifecycle, not an issue.
#                 dropped = ever_parents - latest_parents
#                 conflicting_dropped = set()
#                 for old_pid in dropped:
#                     if old_pid not in product_latest:
#                         # Old parent no longer exists — not a conflict
#                         continue
#                     if vid in product_latest[old_pid]["variant_ids"]:
#                         # Old parent still currently lists this variant
#                         conflicting_dropped.add(old_pid)
#                 if conflicting_dropped:
#                     issues.append(
#                         f"parent changed but old parent(s) still claim this variant: "
#                         f"{', '.join(sorted(conflicting_dropped))}"
#                     )
#                     codes.add("PARENT_CHANGED")

#                 # For each latest parent: does it exist? does it link back?
#                 for pid in sorted(latest_parents):
#                     if pid not in product_latest:
#                         issues.append(f"parent {pid} does not exist as a product")
#                         codes.add("PARENT_NOT_FOUND")
#                     else:
#                         if vid not in product_latest[pid]["variant_ids"]:
#                             issues.append(f"parent {pid} does not link back")
#                             codes.add("PARENT_DOES_NOT_LINK_BACK")

#         if not issues:
#             continue  # clean variant — skip

#         findings[vid] = {
#             "latest_parent": ", ".join(sorted(latest_parents)) if latest_parents else "—",
#             "all_versions_parent": trace,
#             "issues": issues,
#             "issue_codes": codes,
#             "claimed_by_products": claimed_by,
#             "variant_name": pick_live_first(versions, "product_name") or "—",
#             "variant_template_id": pick_live_first(versions, "template_id") or "—",
#         }
#     return findings


# # ---------------- product-side aggregation ----------------

# def aggregate_into_products(variant_findings, variant_versions, product_latest,
#                             variant_latest_parent, product_versions=None):
#     """
#     Group variant-side findings by product. A variant counts under product P if:
#       - any version of the variant claimed P as parent, OR
#       - ANY version of P's variant_id list (live or draft) includes this variant.

#     Walking product version history (not just latest) catches the case where P
#     used to claim V in an older version, then removed V, and V never claimed
#     P back from its own side.

#     Variant ids that don't exist as a VARIANT row anywhere (in live or draft)
#     are treated as ghosts and silently dropped from the report — they're
#     references to non-existent variants and not actionable.

#     Also computes:
#       - should_be_attached:  for each product P, the variants whose LATEST
#                              parent_id (across both files, highest version)
#                              includes P. This is the intended attachment set.
#       - stale_leftover_variants: variants that P's variant_id still lists but
#                                  whose latest parent is NOT P — these are
#                                  leftover claims from before the variant moved.

#     Returns dict[product_id] -> {
#         "missing_variants": list[str],
#         "attached_variants": list[str],
#         "should_be_attached": list[str],
#         "stale_leftover_variants": list[str],
#         "issue_summary": dict[str, int],
#         "variants_by_issue": dict[str, list[str]],
#     }
#     """
#     # Set of variant ids that actually exist as VARIANT rows somewhere.
#     # Anything outside this set is a "ghost" reference and gets dropped.
#     known_variant_ids = set(variant_versions.keys())

#     # Build variant -> set of products it has EVER been associated with
#     variant_to_associated_products = defaultdict(set)
#     # Direction 1: any version of the variant claimed P as parent
#     for vid, versions in variant_versions.items():
#         for v in versions:
#             for pid in v["parent_ids"]:
#                 variant_to_associated_products[vid].add(pid)
#     # Direction 2: ANY version of any product listed this variant —
#     # but only if the variant actually exists as a row somewhere
#     if product_versions:
#         for pid, versions in product_versions.items():
#             for v in versions:
#                 for vid in v["variant_ids"]:
#                     if vid in known_variant_ids:
#                         variant_to_associated_products[vid].add(pid)
#     else:
#         for pid, info in product_latest.items():
#             for vid in info["variant_ids"]:
#                 if vid in known_variant_ids:
#                     variant_to_associated_products[vid].add(pid)

#     by_product = defaultdict(lambda: {
#         "missing_variants": [],
#         "attached_variants": [],
#         "issue_summary": defaultdict(int),
#         "variants_by_issue": defaultdict(set),  # code -> set of variant_ids
#     })

#     # Walk all products we know about
#     for pid, info in product_latest.items():
#         latest_variant_ids = info["variant_ids"]
#         # Variants associated with this product (ever claimed P, or P lists them)
#         associated_vids = {vid for vid, ps in variant_to_associated_products.items() if pid in ps}

#         for vid in associated_vids:
#             in_finding = vid in variant_findings
#             in_latest_list = vid in latest_variant_ids
#             if in_finding:
#                 by_product[pid]["missing_variants"].append(vid)
#                 for code in variant_findings[vid]["issue_codes"]:
#                     by_product[pid]["issue_summary"][code] += 1
#                     by_product[pid]["variants_by_issue"][code].add(vid)
#             elif in_latest_list:
#                 by_product[pid]["attached_variants"].append(vid)
#             else:
#                 # V isn't in P's CURRENT variant_id, and V has no flagged issue.
#                 # Two sub-cases — figure out which:
#                 #   (a) V's parent_id (any version) referenced P → one-sided link,
#                 #       P forgot to list V back. Label: PARENT_DOES_NOT_LINK_BACK
#                 #   (b) V never referenced P at all → P listed V in some OLD
#                 #       version, then dropped it, and V never knew about the
#                 #       relationship. Label: PRODUCT_DROPPED_VARIANT
#                 variant_ever_claimed_p = any(
#                     pid in v["parent_ids"]
#                     for v in variant_versions.get(vid, [])
#                 )
#                 code = "PARENT_DOES_NOT_LINK_BACK" if variant_ever_claimed_p \
#                        else "PRODUCT_DROPPED_VARIANT"
#                 by_product[pid]["missing_variants"].append(vid)
#                 by_product[pid]["issue_summary"][code] += 1
#                 by_product[pid]["variants_by_issue"][code].add(vid)

#     # Sort lists, drop empties
#     out = {}
#     for pid, data in by_product.items():
#         missing = sorted(set(data["missing_variants"]))
#         attached = sorted(set(data["attached_variants"]))
#         if not missing and not attached:
#             continue

#         # should_be_attached: variants whose LATEST parent (highest version
#         # across live + draft) includes this product
#         should = sorted(
#             vid for vid, info in variant_latest_parent.items()
#             if pid in info["latest_parents"]
#         )

#         # stale_leftover: variants in P's latest variant_id list whose latest
#         # parent is NOT P (the variant has moved on but P still claims it)
#         product_latest_variants = product_latest.get(pid, {}).get("variant_ids", set())
#         stale_leftover = []
#         for vid in product_latest_variants:
#             info = variant_latest_parent.get(vid)
#             if info and info["latest_parents"] and pid not in info["latest_parents"]:
#                 stale_leftover.append(vid)
#         stale_leftover = sorted(stale_leftover)

#         variants_by_issue = {
#             code: sorted(vids) for code, vids in data["variants_by_issue"].items()
#         }
#         out[pid] = {
#             "missing_variants": missing,
#             "attached_variants": attached,
#             "should_be_attached": should,
#             "stale_leftover_variants": stale_leftover,
#             "issue_summary": dict(data["issue_summary"]),
#             "variants_by_issue": variants_by_issue,
#         }
#     return out


# # ---------------- writer ----------------

# def write_excel(product_view, variant_findings, product_latest, variant_versions,
#                 variant_meta,
#                 live_counts, draft_counts, output_path):
#     wb = Workbook()

#     header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
#     body_font = Font(name="Arial", size=10)
#     center = Alignment(horizontal="center", vertical="center")
#     left = Alignment(horizontal="left", vertical="center", wrap_text=True)

#     blue_header = PatternFill("solid", start_color="305496")
#     red_header = PatternFill("solid", start_color="C00000")
#     issue_fill = PatternFill("solid", start_color="FCE4D6")
#     clean_fill = PatternFill("solid", start_color="E2EFDA")
#     section_fill = PatternFill("solid", start_color="595959")
#     section_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)

#     # ----- Sheet 1: Product-side summary -----
#     ws1 = wb.active
#     ws1.title = "Product-side summary"
#     headers = ["product_id", "product_name", "product_template_id",
#                "live_state", "draft_state", "issue_label",
#                "missing_variants", "attached_variants",
#                "attached_count", "should_be_attached",
#                "should_be_attached_variant_names",
#                "should_be_attached_variant_template_ids",
#                "conflicting_template_ids",
#                "should_not_attached_count",
#                "stale_leftover_variants",
#                "total_variants", "issue_summary", "variants_by_issue"]
#     ws1.append(headers)
#     for col_idx in range(1, len(headers) + 1):
#         c = ws1.cell(row=1, column=col_idx)
#         c.font = header_font
#         c.fill = red_header
#         c.alignment = center

#     stale_fill = PatternFill("solid", start_color="F4B084")
#     complete_fill = PatternFill("solid", start_color="C6EFCE")

#     for pid in sorted(product_view, key=lambda x: -len(product_view[x]["missing_variants"])):
#         data = product_view[pid]
#         missing = data["missing_variants"]
#         attached = data["attached_variants"]
#         should = data["should_be_attached"]
#         stale = data["stale_leftover_variants"]
#         total = len(missing) + len(attached)
#         attached_count = len(attached)
#         should_not_attached_count = len(set(should) - set(attached))
#         is_complete = total > 0 and attached_count == total and not missing

#         per_source = product_latest.get(pid, {}).get("latest_per_source", {})
#         live_state = per_source.get("live", {}).get("state") or "—"
#         draft_state = per_source.get("draft", {}).get("state") or "—"

#         # Prefer LIVE for name/template_id; fall back to DRAFT if live is empty
#         live_row = per_source.get("live") or {}
#         draft_row = per_source.get("draft") or {}
#         product_name = (
#             live_row.get("product_name")
#             or draft_row.get("product_name")
#             or "—"
#         )
#         product_template_id = (
#             live_row.get("template_id")
#             or draft_row.get("template_id")
#             or "—"
#         )

#         # ----- Build the three new variant-breakdown columns -----
#         # should_be_attached_variant_names: "variant_id (name)" entries, in
#         # the same order as should_be_attached
#         names_list = []
#         tid_to_vids = {}  # template_id -> list of variant_ids (preserves order)
#         for vid in should:
#             meta = variant_meta.get(vid, {})
#             name = meta.get("name") or "—"
#             names_list.append(f"{vid} ({name})")
#             tid = meta.get("template_id") or ""
#             if tid:
#                 tid_to_vids.setdefault(tid, []).append(vid)
#         variant_names_str = ", ".join(names_list) if names_list else "—"

#         if len(tid_to_vids) == 0:
#             variant_template_ids_str = "—"
#             conflicting_template_ids_str = ""
#         elif len(tid_to_vids) == 1:
#             # All variants share the same template_id
#             sole_tid = next(iter(tid_to_vids.keys()))
#             variant_template_ids_str = sole_tid
#             conflicting_template_ids_str = ""
#         else:
#             # Multiple distinct template_ids — leave variant_template_ids empty
#             variant_template_ids_str = ""
#             parts = [f"{tid} ({', '.join(vids)})" for tid, vids in tid_to_vids.items()]
#             conflicting_template_ids_str = "; ".join(parts)

#         codes = sorted(data["issue_summary"].keys())
#         issue_label = " | ".join(codes) if codes else "OK"

#         summary = "; ".join(f"{k}: {v}" for k, v in sorted(data["issue_summary"].items())) or "—"
#         breakdown_lines = []
#         for code in sorted(data["variants_by_issue"]):
#             vids = data["variants_by_issue"][code]
#             breakdown_lines.append(f"{code}: {', '.join(vids)}")
#         breakdown = "\n".join(breakdown_lines) if breakdown_lines else "—"

#         ws1.append([
#             pid,
#             product_name,
#             product_template_id,
#             live_state,
#             draft_state,
#             issue_label,
#             ", ".join(missing) if missing else "—",
#             ", ".join(attached) if attached else "—",
#             attached_count,
#             ", ".join(should) if should else "—",
#             variant_names_str,
#             variant_template_ids_str,
#             conflicting_template_ids_str,
#             should_not_attached_count,
#             ", ".join(stale) if stale else "—",
#             total,
#             summary,
#             breakdown,
#         ])
#         row_idx = ws1.max_row
#         for col_idx in range(1, len(headers) + 1):
#             cell = ws1.cell(row=row_idx, column=col_idx)
#             cell.font = body_font
#             cell.alignment = left
#             if is_complete:
#                 cell.fill = complete_fill
#             elif missing:
#                 cell.fill = issue_fill
#         # stale_leftover_variants is now column 15
#         if stale and not is_complete:
#             ws1.cell(row=row_idx, column=15).fill = stale_fill

#     widths = [42, 50, 42, 14, 14, 40, 60, 60, 16, 60, 60, 42, 60, 22, 60, 14, 40, 80]
#     for i, w in enumerate(widths, start=1):
#         ws1.column_dimensions[get_column_letter(i)].width = w
#     ws1.freeze_panes = "A2"
#     ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

#     # ----- Sheet 2: Variant details -----
#     ws2 = wb.create_sheet("Variant details")
#     headers2 = ["variant_id", "variant_name", "variant_template_id",
#                 "latest_parent", "all_versions_parent",
#                 "issues", "claimed_by_products"]
#     ws2.append(headers2)
#     for col_idx in range(1, len(headers2) + 1):
#         c = ws2.cell(row=1, column=col_idx)
#         c.font = header_font
#         c.fill = blue_header
#         c.alignment = center

#     for vid in sorted(variant_findings):
#         f = variant_findings[vid]
#         ws2.append([
#             vid,
#             f.get("variant_name", "—"),
#             f.get("variant_template_id", "—"),
#             f["latest_parent"],
#             f["all_versions_parent"],
#             "; ".join(f["issues"]),
#             ", ".join(f["claimed_by_products"]) if f["claimed_by_products"] else "—",
#         ])
#         row_idx = ws2.max_row
#         for col_idx in range(1, len(headers2) + 1):
#             cell = ws2.cell(row=row_idx, column=col_idx)
#             cell.font = body_font
#             cell.alignment = left
#             cell.fill = issue_fill

#     widths2 = [42, 50, 42, 40, 80, 80, 60]
#     for i, w in enumerate(widths2, start=1):
#         ws2.column_dimensions[get_column_letter(i)].width = w
#     ws2.freeze_panes = "A2"
#     ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

#     # ----- Sheet 3: Definitions -----
#     ws3 = wb.create_sheet("Definitions")
#     ws3.append(["issue code", "what it means"])
#     for col_idx in (1, 2):
#         c = ws3.cell(row=1, column=col_idx)
#         c.font = header_font
#         c.fill = blue_header
#         c.alignment = center
#     defs = [
#         ("STALE_LATEST_EMPTY",
#          "Some earlier version of this variant had a parent_id (e.g. P1), but the LATEST version "
#          "has an empty parent_id. Either the link was deliberately removed without back-cleanup, or "
#          "the data went stale."),
#         ("PARENT_CHANGED",
#          "The variant was reparented across versions and an older parent still both EXISTS as a "
#          "product AND still lists this variant in its variant_id. Two products both think they own "
#          "this variant — genuine conflict. Clean reparenting (where the old parent's variant_id was "
#          "also updated to remove the variant) is normal lifecycle and is NOT flagged."),
#         ("ORPHAN_BUT_REFERENCED",
#          "The variant has NEVER had a parent_id in any version, but some product's variant_id list "
#          "references it. The variant doesn't acknowledge the relationship from its side."),
#         ("PARENT_NOT_FOUND",
#          "A product id that this variant has ever pointed at (in latest OR any historical version) "
#          "doesn't exist as a PRODUCT row in either live or draft data. If the missing parent was only "
#          "claimed historically and the latest parent_id is empty, this fires INSTEAD of "
#          "STALE_LATEST_EMPTY because the deleted parent is the more important problem."),
#         ("PARENT_DOES_NOT_LINK_BACK",
#          "The variant currently claims a parent product (in its latest parent_id), and that "
#          "product exists, but the product's latest variant_id list does NOT include this variant. "
#          "Live one-sided link — variant says yes, product says no."),
#         ("NO_LINK_ANYWHERE",
#          "The variant has no parent_id in any version AND no product references it. Fully disconnected."),
#         ("PRODUCT_DROPPED_VARIANT",
#          "An older version of some product listed the variant in its variant_id, but the latest "
#          "version no longer does — and the variant itself never claimed this product as a parent "
#          "in any version. The link existed historically only from the product's side, then was "
#          "removed. Indicates a one-sided historical association that's now fully severed."),
#     ]
#     for code, meaning in defs:
#         ws3.append([code, meaning])
#         row_idx = ws3.max_row
#         for col_idx in (1, 2):
#             cell = ws3.cell(row=row_idx, column=col_idx)
#             cell.font = body_font
#             cell.alignment = left
#             cell.fill = issue_fill
#         ws3.row_dimensions[row_idx].height = 45
#     ws3.column_dimensions["A"].width = 32
#     ws3.column_dimensions["B"].width = 100

#     # ----- Sheet 4: Inventory -----
#     ws4 = wb.create_sheet("Inventory")
#     ws4["A1"] = "Source counts"
#     ws4["A1"].font = Font(name="Arial", bold=True)
#     ws4["A2"] = "live"
#     ws4["B2"] = f"{live_counts['products']} PRODUCT, {live_counts['variants']} VARIANT"
#     ws4["A3"] = "draft"
#     ws4["B3"] = f"{draft_counts['products']} PRODUCT, {draft_counts['variants']} VARIANT"
#     ws4.column_dimensions["A"].width = 20
#     ws4.column_dimensions["B"].width = 50

#     # ----- Sheet 5: Summary -----
#     ws5 = wb.create_sheet("Summary")
#     ws5["A1"] = "Total variants with issues"
#     ws5["B1"] = len(variant_findings)
#     ws5["A2"] = "Total products affected"
#     ws5["B2"] = len(product_view)

#     issue_totals = defaultdict(int)
#     for f in variant_findings.values():
#         for code in f["issue_codes"]:
#             issue_totals[code] += 1
#     row = 4
#     ws5.cell(row=row, column=1, value="By issue code (counts variants once per code)").font = Font(name="Arial", bold=True)
#     for code in sorted(issue_totals):
#         row += 1
#         ws5.cell(row=row, column=1, value=f"  {code}")
#         ws5.cell(row=row, column=2, value=issue_totals[code])

#     ws5.column_dimensions["A"].width = 50
#     ws5.column_dimensions["B"].width = 14

#     wb.save(output_path)


# # ---------------- main ----------------

# def main():
#     parser = argparse.ArgumentParser(
#         description=__doc__,
#         formatter_class=argparse.RawDescriptionHelpFormatter,
#     )
#     parser.add_argument("live", help="Path to LIVE CSV/TSV file")
#     parser.add_argument("-d", "--draft", default=None,
#                         help="Optional path to DRAFT CSV/TSV file")
#     parser.add_argument("-o", "--output", default="variant_version_scan.xlsx",
#                         help="Output Excel file path")
#     args = parser.parse_args()

#     print(f"Loading live: {args.live}")
#     live_products_v, live_variants_v = load_file(args.live, "live")
#     print(f"  {len(live_products_v)} products (versioned), {len(live_variants_v)} variants (versioned)")

#     draft_products_v, draft_variants_v = ({}, {})
#     if args.draft:
#         print(f"Loading draft: {args.draft}")
#         draft_products_v, draft_variants_v = load_file(args.draft, "draft")
#         print(f"  {len(draft_products_v)} products, {len(draft_variants_v)} variants")

#     # Merge per-item version lists from both sources
#     product_versions = merge_version_dicts(live_products_v, draft_products_v)
#     variant_versions = merge_version_dicts(live_variants_v, draft_variants_v)

#     # Compute latest state of every product (union of latest from each source)
#     product_latest = latest_product_state(product_versions)
#     # Compute the single latest parent per variant (highest version across files,
#     # falling back to historical parent_id, then to current product claimers,
#     # then to historical product claimers if needed)
#     variant_latest_parent = pick_latest_parent_per_variant(
#         variant_versions, product_latest, product_versions=product_versions
#     )
#     # name + template_id per variant (for the product-side variant breakdown columns)
#     variant_meta = build_variant_meta(variant_versions)

#     print("Analyzing variants across all versions...")
#     variant_findings = analyze_variants(variant_versions, product_latest,
#                                         product_versions=product_versions)
#     print(f"  {len(variant_findings)} variants have issues")

#     print("Aggregating into product-side view...")
#     product_view = aggregate_into_products(variant_findings, variant_versions,
#                                            product_latest, variant_latest_parent,
#                                            product_versions=product_versions)
#     print(f"  {len(product_view)} products have affected variants")

#     print(f"Writing {args.output}...")
#     write_excel(
#         product_view, variant_findings, product_latest, variant_versions,
#         variant_meta=variant_meta,
#         live_counts={"products": len(live_products_v), "variants": len(live_variants_v)},
#         draft_counts={"products": len(draft_products_v), "variants": len(draft_variants_v)},
#         output_path=args.output,
#     )
#     print("Done.")


# if __name__ == "__main__":
#     main()
"""
Walk every VERSION of every variant in both live + draft files, detect link
problems across version history, then aggregate the findings into a
product-side view.

Six variant-side issue patterns are detected:

  1. STALE_LATEST_EMPTY        — parent_id appeared in an earlier version but
                                 the latest version's parent_id is empty.
                                 (e.g. versions: P1, -, P1, -  → latest is empty)
  2. PARENT_CHANGED            — variant was reparented across versions AND the
                                 OLD parent still exists AND still lists the
                                 variant in its variant_id (genuine ownership
                                 conflict — two products both claim V).
                                 Clean reparenting (old parent properly updated)
                                 is normal lifecycle and not flagged.
  3. ORPHAN_BUT_REFERENCED     — variant has no parent_id in any version, but
                                 some product references it.
  4. PARENT_NOT_FOUND          — A parent the variant has ever pointed at
                                 (latest OR any historical version) doesn't
                                 exist as a PRODUCT anywhere in the catalog.
  5. PARENT_DOES_NOT_LINK_BACK — latest parent exists but its variant_id list
                                 doesn't include this variant.
  6. NO_LINK_ANYWHERE          — no parent_id in any version AND no product
                                 references this variant anywhere.

Output workbook (5 sheets):

  Product-side summary   one row per affected product:
                         product_id | missing_variants | attached_variants
                                    | total_variants  | issue_summary
  Variant details         one row per variant with at least one issue:
                          variant_id | latest_parent | all_versions_parent
                                     | issues | claimed_by_products
  Definitions             explanation of each issue type
  Inventory               counts of products/variants per source
  Summary                 overall counts by issue category

Usage:
    python variant_version_scan.py <live.csv> [--draft <draft.csv>] [-o out.xlsx]
"""

import argparse
import ast
import csv
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Raise csv field size limit — some rows (e.g. large request_schema JSON blobs)
# exceed the default 131,072 char limit. Use the largest value csv accepts on
# this platform.
_maxInt = sys.maxsize
while True:
    try:
        csv.field_size_limit(_maxInt)
        break
    except OverflowError:
        _maxInt = int(_maxInt / 10)


# ---------------- helpers ----------------

def parse_id_list(value) -> list:
    if value is None:
        return []
    s = str(value).strip()
    if s == "" or s in ("[]", "['']", '[""]', "null", "None", "—"):
        return []
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple)):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except (ValueError, SyntaxError):
            pass
        inner = s[1:-1]
        parts = re.split(r"\s*,\s*", inner)
        return [p.strip().strip("'\"") for p in parts if p.strip().strip("'\"")]
    parts = re.split(r"\s*,\s*", s)
    return [p.strip().strip("'\"") for p in parts if p.strip().strip("'\"")]


import json as _json


def _normalize_whitespace_escapes_outside_strings(s: str) -> str:
    """Walk through a JSON-ish string. When OUTSIDE a JSON string value, convert
    literal `\\n` / `\\t` / `\\r` escape sequences to real whitespace (they were
    meant as inter-token whitespace by an over-escaping writer). When INSIDE a
    string value, leave escape sequences alone so json.loads handles them.

    This is a safer replacement for blanket s.replace('\\n', '\\n') which mangles
    real JSON escape sequences inside string values."""
    out = []
    i = 0
    n = len(s)
    in_string = False
    while i < n:
        c = s[i]
        if not in_string:
            if c == '\\' and i + 1 < n:
                nxt = s[i + 1]
                if nxt == 'n':
                    out.append('\n'); i += 2; continue
                if nxt == 't':
                    out.append('\t'); i += 2; continue
                if nxt == 'r':
                    out.append('\r'); i += 2; continue
                if nxt == '\\':
                    # Two literal backslashes between tokens — keep one
                    out.append('\\'); i += 2; continue
            if c == '"':
                in_string = True
            out.append(c)
            i += 1
        else:
            # Inside string: preserve all escape sequences as-is
            if c == '\\' and i + 1 < n:
                out.append(c); out.append(s[i + 1]); i += 2; continue
            if c == '"':
                in_string = False
            out.append(c)
            i += 1
    return ''.join(out)


def _repair_unescaped_quotes(s: str) -> str:
    """Walk through a JSON-ish string and escape any quote that's inside a
    string value but isn't followed by JSON-structural characters. Handles the
    common case where the source writer escaped quotes for CSV but forgot to
    escape them for JSON, producing lone unescaped quotes inside string values.

    Examples this fixes:
      - "2.5" Gel Visco"  →  "2.5\\" Gel Visco"
      - "2'3\\" x 8'"     →  "2'3\\\\\\" x 8'" (already-broken double-backslash + quote)
    """
    out = []
    i = 0
    n = len(s)
    in_string = False
    while i < n:
        c = s[i]
        if c == '\\' and i + 1 < n:
            # Pass through any escape sequence untouched
            out.append(c)
            out.append(s[i + 1])
            i += 2
            continue
        if c == '"':
            if not in_string:
                in_string = True
                out.append(c)
                i += 1
                continue
            # Inside a string. Decide if this is the real closing quote by
            # peeking ahead past whitespace for a JSON-structural character.
            j = i + 1
            while j < n and s[j] in ' \t\n\r':
                j += 1
            if j == n or s[j] in ',:}]':
                # Real closing quote
                in_string = False
                out.append(c)
            else:
                # Content quote — escape it
                out.append('\\')
                out.append(c)
            i += 1
            continue
        out.append(c)
        i += 1
    return ''.join(out)


def _safe_json_loads(s: str):
    """Try to parse s as JSON, applying several recovery strategies on failure.
    Uses strict=False so control characters inside string values don't break parsing."""
    try:
        return _json.loads(s, strict=False)
    except (ValueError, TypeError):
        pass
    # Strategy 1: stateful normalization of inter-token whitespace escapes
    cleaned = _normalize_whitespace_escapes_outside_strings(s)
    if cleaned != s:
        try:
            return _json.loads(cleaned, strict=False)
        except (ValueError, TypeError):
            pass
    # Strategy 2: repair lone unescaped quotes inside string values
    repaired = _repair_unescaped_quotes(s)
    if repaired != s:
        try:
            return _json.loads(repaired, strict=False)
        except (ValueError, TypeError):
            pass
    # Strategy 3: combine — repair quotes on the whitespace-normalized form
    repaired_clean = _repair_unescaped_quotes(cleaned)
    if repaired_clean != cleaned:
        try:
            return _json.loads(repaired_clean, strict=False)
        except (ValueError, TypeError):
            pass
    return None


def parse_template_id(raw: str) -> str:
    """Extract id from template_id column, which looks like:
       {"id":"3956b34b-...","version":12}"""
    if not raw:
        return ""
    s = raw.strip()
    if not s or s in ("null", "None"):
        return ""
    obj = _safe_json_loads(s)
    if isinstance(obj, dict) and "id" in obj:
        return str(obj["id"]).strip()
    return ""


def _extract_en_us_displayname_via_regex(s: str) -> str:
    """Last-resort: find "displayName" : { ... "en_us" : "VALUE" ... } using a
    regex, when full JSON parsing fails. Uses a permissive value capture that
    only treats a quote as the closing quote if it's followed by JSON
    structural characters (`,` `}` `]` with optional whitespace, where
    whitespace can also be `\\n`, `\\t`, `\\r` escape sequences)."""
    m = re.search(
        r'"displayName"[\s\S]{0,500}?"localizations"[\s\S]{0,300}?"en_us"\s*:\s*"',
        s,
    )
    if not m:
        return ""
    start = m.end()
    i = start
    n = len(s)
    while i < n:
        c = s[i]
        if c == '\\' and i + 1 < n:
            i += 2
            continue
        if c == '"':
            # Peek ahead past whitespace (real or escape sequences) for a JSON delimiter
            j = i + 1
            while j < n:
                if s[j] in ' \t\n\r':
                    j += 1
                    continue
                if s[j] == '\\' and j + 1 < n and s[j + 1] in 'ntr':
                    j += 2
                    continue
                break
            if j == n or s[j] in ',}]':
                val = s[start:i]
                # Cleanup escapes — collapse `\\` pairs FIRST, then handle remaining `\"`
                val = (val
                       .replace('\\\\', '\\')      # 2 backslashes → 1
                       .replace('\\"', '"')        # backslash-quote → quote
                       .replace("\\'", "'")
                       .replace('\\n', ' ')
                       .replace('\\t', ' ')
                       .replace('\\r', ' '))
                val = " ".join(val.split())
                return val
            i += 1
            continue
        i += 1
    return ""


def parse_product_name(raw: str) -> str:
    """Extract displayName.localizations.en_us from the request_schema JSON column."""
    if not raw:
        return ""
    s = raw.strip()
    if not s or s in ("null", "None"):
        return ""
    obj = _safe_json_loads(s)
    if isinstance(obj, dict):
        display = obj.get("displayName") or {}
        if isinstance(display, dict):
            locs = display.get("localizations") or {}
            if isinstance(locs, dict):
                val = locs.get("en_us")
                if val:
                    cleaned = str(val)
                    # Order matters: collapse `\\` pairs FIRST, then handle remaining `\"`
                    cleaned = (cleaned
                               .replace('\\\\', '\\')
                               .replace('\\"', '"')
                               .replace("\\'", "'")
                               .replace("\\r\\n", " ")
                               .replace("\\n", " ")
                               .replace("\\t", " ")
                               .replace("\\r", " "))
                    cleaned = " ".join(cleaned.split())
                    return cleaned
    # JSON parsing failed or no displayName found — try regex extraction
    val = _extract_en_us_displayname_via_regex(s)
    return val


def _looks_like_backslash_escaped(path: str) -> bool:
    """Peek at the first ~10KB of body. If we see `\\"` inside a quoted field,
    assume backslash-escape format."""
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            f.readline()  # skip header
            sample = f.read(10000)
        return '\\"' in sample
    except OSError:
        return False


def _normalize_backslash_escaped_csv(path: str) -> "io.StringIO":
    """
    Read a CSV that uses backslash-escaped quotes inside quoted fields and
    return a StringIO buffer with standard CSV escaping (doubled quotes).

    Rules:
      - `\\"` inside a quoted field  -> `""` (CSV doubled-quote)
      - `\\\\` (literal backslash)   -> single `\\`
      - everything else passes through unchanged (so embedded \\n stays as \\n
        and JSON parsers see whitespace, not the letter 'n')

    We track whether we're inside a quoted field by toggling on `"` chars,
    handling the doubled-quote escape afterward.
    """
    import io
    out = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        content = f.read()

    in_quotes = False
    i = 0
    n = len(content)
    while i < n:
        c = content[i]
        if c == '"':
            # Check whether the previous char was a backslash inside quotes
            # (we should never see this case because we normalize backslash-escapes
            # below — so a literal " always opens/closes a field).
            out.append('"')
            in_quotes = not in_quotes
            i += 1
            continue
        if c == '\\' and in_quotes and i + 1 < n:
            nxt = content[i + 1]
            if nxt == '"':
                # \"  ->  ""  (CSV doubled-quote escape)
                out.append('""')
                i += 2
                continue
            if nxt == '\\':
                # \\  ->  \  (literal backslash)
                out.append('\\')
                i += 2
                continue
            # Other escapes (\n, \t, \uXXXX, ...) inside JSON-bearing fields
            # should be PASSED THROUGH untouched so json.loads sees them.
            out.append(c)
            i += 1
            continue
        out.append(c)
        i += 1

    return io.StringIO("".join(out))


def load_file(path: str, source_label: str):
    """
    Reads ALL versions, not just the latest. Returns:
        product_versions: dict[item_id] -> list of {version, parent_id_raw,
                                                    variant_id_raw, state, ...}
        variant_versions: dict[item_id] -> list of the same per-version dicts
    """
    product_versions = defaultdict(list)
    variant_versions = defaultdict(list)

    # Pre-process if needed
    use_backslash = _looks_like_backslash_escaped(path)
    if use_backslash:
        f = _normalize_backslash_escaped_csv(path)
        first_line = f.readline()
        f.seek(0)
        dialect = csv.excel  # standard now
    else:
        f = open(path, "r", encoding="utf-8-sig", newline="")
        first_line = f.readline()
        f.seek(0)
        dialect = csv.excel_tab if "\t" in first_line else csv.excel

    try:
        reader = csv.DictReader(f, dialect=dialect)
        if "item_id" not in (reader.fieldnames or []):
            sys.exit(f"ERROR: {path} missing 'item_id' column (got {reader.fieldnames})")
        has_version = "version" in (reader.fieldnames or [])
        for row in reader:
            item_id = (row.get("item_id") or "").strip()
            scope = (row.get("item_scope") or "").strip().upper()
            if not item_id or scope not in ("PRODUCT", "VARIANT"):
                continue
            try:
                version = int(row["version"]) if has_version else 1
            except (ValueError, TypeError):
                version = 1
            entry = {
                "version": version,
                "source": source_label,
                "scope": scope,
                "state": row.get("state", ""),
                "site_ids": row.get("site_ids", ""),
                "sku": row.get("sku", ""),
                "date_modified": row.get("date_modified", ""),
                "parent_ids": parse_id_list(row.get("parent_id", "")),
                "variant_ids": parse_id_list(row.get("variant_id", "")),
                "template_id": parse_template_id(row.get("template_id", "")),
                "product_name": parse_product_name(row.get("request_schema", "")),
            }
            if scope == "PRODUCT":
                product_versions[item_id].append(entry)
            else:
                variant_versions[item_id].append(entry)
    finally:
        try:
            f.close()
        except Exception:
            pass
    return product_versions, variant_versions


def merge_version_dicts(*dicts):
    """Combine the same item's version lists from multiple sources."""
    merged = defaultdict(list)
    for d in dicts:
        for k, v_list in d.items():
            merged[k].extend(v_list)
    # Sort each item's versions: by source ('live' first), then by version
    for k in merged:
        merged[k].sort(key=lambda e: (0 if e["source"] == "live" else 1, e["version"]))
    return merged


def build_variant_meta(variant_versions):
    """Return dict[variant_id] -> {"name": str, "template_id": str} for every
    variant.

    Name/template_id selection rules:
      - Prefer LIVE: take the latest live version's value if non-empty.
      - Fall back to DRAFT: if live is missing or has an empty value, use the
        latest draft version's value.
    """
    meta = {}
    for vid, versions in variant_versions.items():
        if not versions:
            continue
        name = pick_live_first(versions, "product_name")
        tid = pick_live_first(versions, "template_id")
        meta[vid] = {"name": name, "template_id": tid}
    return meta


def pick_live_first(versions, field):
    """Pick the value of `field` preferring live (latest live version with a
    non-empty value), falling back to draft if live is empty/missing."""
    if not versions:
        return ""
    live_rows = [v for v in versions if v.get("source") == "live"]
    draft_rows = [v for v in versions if v.get("source") == "draft"]

    def latest_nonempty(rows):
        # Sort newest first and return the first non-empty value
        try:
            ordered = sorted(rows, key=lambda r: r.get("version", 0), reverse=True)
        except TypeError:
            ordered = list(reversed(rows))
        for r in ordered:
            val = r.get(field)
            if val:
                return val
        return ""

    return latest_nonempty(live_rows) or latest_nonempty(draft_rows)


def pick_latest_parent_per_variant(variant_versions, product_latest=None,
                                   product_versions=None):
    """
    For each variant_id, decide its intended parent set using these rules:

    1. If the absolute-latest version has a non-empty parent_id, use that.
    2. Else if any earlier version had a non-empty parent_id, fall back to the
       highest-version row that did.
    3. Else (the variant has NEVER claimed a parent in any version), fall back
       to whichever products CURRENTLY claim this variant in their variant_id
       list (product_latest). This is the "orphan-but-referenced" rescue.
    4. Else fall back to whichever products HISTORICALLY claimed this variant
       in any version of their variant_id list (product_versions). This is the
       "both sides forgot" rescue — the variant exists, an old product version
       listed it, the link should still hold even though both sides went quiet.
    5. Else return empty (truly orphan, no claimers anywhere).

    Returns:
        dict[variant_id] -> {"latest_parents": set, "latest_version": int,
                             "latest_source": str, "from_fallback": bool,
                             "from_product_claim": bool,
                             "from_historical_product_claim": bool}
    """
    # Build a referrers index from product_latest: variant_id -> set of products
    referrers_latest = defaultdict(set)
    if product_latest:
        for pid, info in product_latest.items():
            for vid in info.get("variant_ids", set()):
                referrers_latest[vid].add(pid)
    # Historical: any version of any product that listed this variant
    referrers_historical = defaultdict(set)
    if product_versions:
        for pid, versions in product_versions.items():
            for v in versions:
                for vid in v["variant_ids"]:
                    referrers_historical[vid].add(pid)

    out = {}
    for vid, versions in variant_versions.items():
        if not versions:
            continue

        try:
            absolute_latest = max(versions, key=lambda v: v["version"])
        except TypeError:
            absolute_latest = versions[-1]

        from_fallback = False
        from_product_claim = False
        from_historical_product_claim = False

        if absolute_latest["parent_ids"]:
            chosen_parents = set(absolute_latest["parent_ids"])
            chosen_version = absolute_latest["version"]
            chosen_source = absolute_latest["source"]
        else:
            # Try fallback: highest-version row with non-empty parent_ids
            with_parent = [v for v in versions if v["parent_ids"]]
            if with_parent:
                try:
                    fb = max(with_parent, key=lambda v: v["version"])
                except TypeError:
                    fb = with_parent[-1]
                chosen_parents = set(fb["parent_ids"])
                chosen_version = fb["version"]
                chosen_source = fb["source"]
                from_fallback = True
            else:
                # No version ever had a parent — try product claims
                current_claimers = referrers_latest.get(vid, set())
                historical_claimers = referrers_historical.get(vid, set())
                if current_claimers:
                    chosen_parents = set(current_claimers)
                    chosen_version = absolute_latest["version"]
                    chosen_source = absolute_latest["source"]
                    from_product_claim = True
                elif historical_claimers:
                    chosen_parents = set(historical_claimers)
                    chosen_version = absolute_latest["version"]
                    chosen_source = absolute_latest["source"]
                    from_historical_product_claim = True
                else:
                    chosen_parents = set()
                    chosen_version = absolute_latest["version"]
                    chosen_source = absolute_latest["source"]

        out[vid] = {
            "latest_parents": chosen_parents,
            "latest_version": chosen_version,
            "latest_source": chosen_source,
            "from_fallback": from_fallback,
            "from_product_claim": from_product_claim,
            "from_historical_product_claim": from_historical_product_claim,
        }
    return out


def latest_product_state(product_versions):
    """For each product_id, pick the latest version per source and merge.
    Returns dict[product_id] -> {variant_ids: set (union of live+draft latest),
                                 latest_per_source: {source: version dict}}"""
    out = {}
    for pid, versions in product_versions.items():
        by_source = defaultdict(list)
        for v in versions:
            by_source[v["source"]].append(v)
        latest_per_source = {}
        union_variants = set()
        for src, vs in by_source.items():
            try:
                latest = max(vs, key=lambda x: x["version"])
            except TypeError:
                latest = vs[-1]
            latest_per_source[src] = latest
            union_variants.update(latest["variant_ids"])
        out[pid] = {
            "variant_ids": union_variants,
            "latest_per_source": latest_per_source,
        }
    return out


# ---------------- variant-side analysis ----------------

def analyze_variants(variant_versions, product_latest, product_versions=None):
    """
    For each variant, walk all versions across both sources and classify any
    issues. Returns:
        variant_findings: dict[variant_id] -> {
            "latest_parent": str (or "—"),
            "all_versions_parent": str,
            "issues": list[str],
            "issue_codes": set[str],
            "claimed_by_products": list[str],
        }
    """
    # Build "products that claim this variant" indexes — both latest and historical
    referrers_latest = defaultdict(set)
    for pid, info in product_latest.items():
        for vid in info["variant_ids"]:
            referrers_latest[vid].add(pid)
    # Historical referrers: any version of any product that ever listed the variant
    referrers_ever = defaultdict(set)
    if product_versions:
        for pid, versions in product_versions.items():
            for v in versions:
                for vid in v["variant_ids"]:
                    referrers_ever[vid].add(pid)
    else:
        referrers_ever = referrers_latest

    findings = {}
    for vid, versions in variant_versions.items():
        if not versions:
            continue

        # Group versions by source and pick the latest per source
        by_source = defaultdict(list)
        for v in versions:
            by_source[v["source"]].append(v)
        latest_per_source = {}
        for src, vs in by_source.items():
            try:
                latest_per_source[src] = max(vs, key=lambda x: x["version"])
            except TypeError:
                latest_per_source[src] = vs[-1]

        # All parent_ids EVER seen across every version (any source)
        ever_parents = set()
        for v in versions:
            ever_parents.update(v["parent_ids"])

        # Latest parent_ids = union of latest-per-source's parent_ids
        latest_parents = set()
        for src_latest in latest_per_source.values():
            latest_parents.update(src_latest["parent_ids"])

        claimed_by_latest = sorted(referrers_latest.get(vid, set()))
        claimed_by_ever = sorted(referrers_ever.get(vid, set()))
        # For display: prefer current claimers; fall back to historical
        claimed_by = claimed_by_latest or claimed_by_ever

        # ----- Classify issues -----
        issues = []
        codes = set()

        # Per-version trace for the "all_versions_parent" column.
        # Deduplicate by (source, version) to avoid showing identical rows
        # multiple times. Use ' | ' between version entries and '+' between
        # parent_ids within one entry so commas inside parent lists don't
        # collide with the version separator.
        seen_sv = set()
        trace_entries = []
        for v in sorted(versions, key=lambda x: (x["source"], x["version"])):
            sv = (v["source"], v["version"])
            if sv in seen_sv:
                continue
            seen_sv.add(sv)
            parents_str = "+".join(v["parent_ids"]) if v["parent_ids"] else "-"
            trace_entries.append(f"{v['source'][0]}v{v['version']}={parents_str}")
        trace = " | ".join(trace_entries)

        if not ever_parents:
            # Variant never claimed any parent in any version
            if claimed_by_latest:
                # Some product STILL lists this variant in its current variant_id
                issues.append(f"orphan but referenced by {len(claimed_by_latest)} product(s)")
                codes.add("ORPHAN_BUT_REFERENCED")
            elif claimed_by_ever:
                # No current product lists it, but an OLDER version of some product did
                issues.append(
                    f"product(s) dropped this variant in latest version: "
                    f"{', '.join(claimed_by_ever)}"
                )
                codes.add("PRODUCT_DROPPED_VARIANT")
            else:
                issues.append("no link anywhere — no parent in any version, no product references it")
                codes.add("NO_LINK_ANYWHERE")
        else:
            # The variant DID have parents at some point
            if not latest_parents:
                # Latest version is empty even though earlier versions had parents.
                # BUT: if any historical parent no longer exists as a product, that's
                # a stronger signal than just "link cleared" — fire PARENT_NOT_FOUND
                # for those, and STALE_LATEST_EMPTY only if no missing historical parents.
                missing_historical = {p for p in ever_parents if p not in product_latest}
                if missing_historical:
                    for pid in sorted(missing_historical):
                        issues.append(
                            f"parent {pid} (claimed in earlier version) does not exist as a product"
                        )
                    codes.add("PARENT_NOT_FOUND")
                    # If there are also surviving historical parents that still exist,
                    # also note the stale-clear so the user sees both pieces of context.
                    surviving = ever_parents - missing_historical
                    if surviving:
                        issues.append(
                            f"latest version is empty; earlier versions also referenced existing parent(s): "
                            f"{', '.join(sorted(surviving))}"
                        )
                        codes.add("STALE_LATEST_EMPTY")
                else:
                    issues.append(
                        f"latest version is empty but earlier had: {', '.join(sorted(ever_parents))}"
                    )
                    codes.add("STALE_LATEST_EMPTY")
            else:
                # Latest has parents — check if any earlier parent was dropped
                # AND still claims this variant (genuine ownership conflict).
                # A clean reparenting (old parent updated to remove V) is normal
                # lifecycle, not an issue.
                dropped = ever_parents - latest_parents
                conflicting_dropped = set()
                for old_pid in dropped:
                    if old_pid not in product_latest:
                        # Old parent no longer exists — not a conflict
                        continue
                    if vid in product_latest[old_pid]["variant_ids"]:
                        # Old parent still currently lists this variant
                        conflicting_dropped.add(old_pid)
                if conflicting_dropped:
                    issues.append(
                        f"parent changed but old parent(s) still claim this variant: "
                        f"{', '.join(sorted(conflicting_dropped))}"
                    )
                    codes.add("PARENT_CHANGED")

                # For each latest parent: does it exist? does it link back?
                for pid in sorted(latest_parents):
                    if pid not in product_latest:
                        issues.append(f"parent {pid} does not exist as a product")
                        codes.add("PARENT_NOT_FOUND")
                    else:
                        if vid not in product_latest[pid]["variant_ids"]:
                            issues.append(f"parent {pid} does not link back")
                            codes.add("PARENT_DOES_NOT_LINK_BACK")

        if not issues:
            continue  # clean variant — skip

        findings[vid] = {
            "latest_parent": ", ".join(sorted(latest_parents)) if latest_parents else "—",
            "all_versions_parent": trace,
            "issues": issues,
            "issue_codes": codes,
            "claimed_by_products": claimed_by,
            "variant_name": pick_live_first(versions, "product_name") or "—",
            "variant_template_id": pick_live_first(versions, "template_id") or "—",
        }
    return findings


# ---------------- product-side aggregation ----------------

def aggregate_into_products(variant_findings, variant_versions, product_latest,
                            variant_latest_parent, product_versions=None):
    """
    Group variant-side findings by product. A variant counts under product P if:
      - any version of the variant claimed P as parent, OR
      - ANY version of P's variant_id list (live or draft) includes this variant.

    Walking product version history (not just latest) catches the case where P
    used to claim V in an older version, then removed V, and V never claimed
    P back from its own side.

    Variant ids that don't exist as a VARIANT row anywhere (in live or draft)
    are treated as ghosts and silently dropped from the report — they're
    references to non-existent variants and not actionable.

    Also computes:
      - should_be_attached:  for each product P, the variants whose LATEST
                             parent_id (across both files, highest version)
                             includes P. This is the intended attachment set.
      - stale_leftover_variants: variants that P's variant_id still lists but
                                 whose latest parent is NOT P — these are
                                 leftover claims from before the variant moved.

    Returns dict[product_id] -> {
        "missing_variants": list[str],
        "attached_variants": list[str],
        "should_be_attached": list[str],
        "stale_leftover_variants": list[str],
        "issue_summary": dict[str, int],
        "variants_by_issue": dict[str, list[str]],
    }
    """
    # Set of variant ids that actually exist as VARIANT rows somewhere.
    # Anything outside this set is a "ghost" reference and gets dropped.
    known_variant_ids = set(variant_versions.keys())

    # Build variant -> set of products it has EVER been associated with
    variant_to_associated_products = defaultdict(set)
    # Direction 1: any version of the variant claimed P as parent
    for vid, versions in variant_versions.items():
        for v in versions:
            for pid in v["parent_ids"]:
                variant_to_associated_products[vid].add(pid)
    # Direction 2: ANY version of any product listed this variant —
    # but only if the variant actually exists as a row somewhere
    if product_versions:
        for pid, versions in product_versions.items():
            for v in versions:
                for vid in v["variant_ids"]:
                    if vid in known_variant_ids:
                        variant_to_associated_products[vid].add(pid)
    else:
        for pid, info in product_latest.items():
            for vid in info["variant_ids"]:
                if vid in known_variant_ids:
                    variant_to_associated_products[vid].add(pid)

    by_product = defaultdict(lambda: {
        "missing_variants": [],
        "attached_variants": [],
        "issue_summary": defaultdict(int),
        "variants_by_issue": defaultdict(set),  # code -> set of variant_ids
    })

    # Walk all products we know about
    for pid, info in product_latest.items():
        latest_variant_ids = info["variant_ids"]
        # Variants associated with this product (ever claimed P, or P lists them)
        associated_vids = {vid for vid, ps in variant_to_associated_products.items() if pid in ps}

        for vid in associated_vids:
            in_finding = vid in variant_findings
            in_latest_list = vid in latest_variant_ids
            if in_finding:
                by_product[pid]["missing_variants"].append(vid)
                for code in variant_findings[vid]["issue_codes"]:
                    by_product[pid]["issue_summary"][code] += 1
                    by_product[pid]["variants_by_issue"][code].add(vid)
            elif in_latest_list:
                by_product[pid]["attached_variants"].append(vid)
            else:
                # V isn't in P's CURRENT variant_id, and V has no flagged issue.
                # Two sub-cases — figure out which:
                #   (a) V's parent_id (any version) referenced P → one-sided link,
                #       P forgot to list V back. Label: PARENT_DOES_NOT_LINK_BACK
                #   (b) V never referenced P at all → P listed V in some OLD
                #       version, then dropped it, and V never knew about the
                #       relationship. Label: PRODUCT_DROPPED_VARIANT
                variant_ever_claimed_p = any(
                    pid in v["parent_ids"]
                    for v in variant_versions.get(vid, [])
                )
                code = "PARENT_DOES_NOT_LINK_BACK" if variant_ever_claimed_p \
                       else "PRODUCT_DROPPED_VARIANT"
                by_product[pid]["missing_variants"].append(vid)
                by_product[pid]["issue_summary"][code] += 1
                by_product[pid]["variants_by_issue"][code].add(vid)

    # Sort lists, drop empties
    out = {}
    for pid, data in by_product.items():
        missing = sorted(set(data["missing_variants"]))
        attached = sorted(set(data["attached_variants"]))
        if not missing and not attached:
            continue

        # should_be_attached: variants whose LATEST parent (highest version
        # across live + draft) includes this product
        should = sorted(
            vid for vid, info in variant_latest_parent.items()
            if pid in info["latest_parents"]
        )

        # stale_leftover: variants in P's latest variant_id list whose latest
        # parent is NOT P (the variant has moved on but P still claims it)
        product_latest_variants = product_latest.get(pid, {}).get("variant_ids", set())
        stale_leftover = []
        for vid in product_latest_variants:
            info = variant_latest_parent.get(vid)
            if info and info["latest_parents"] and pid not in info["latest_parents"]:
                stale_leftover.append(vid)
        stale_leftover = sorted(stale_leftover)

        variants_by_issue = {
            code: sorted(vids) for code, vids in data["variants_by_issue"].items()
        }
        out[pid] = {
            "missing_variants": missing,
            "attached_variants": attached,
            "should_be_attached": should,
            "stale_leftover_variants": stale_leftover,
            "issue_summary": dict(data["issue_summary"]),
            "variants_by_issue": variants_by_issue,
        }
    return out


# ---------------- writer ----------------

def write_excel(product_view, variant_findings, product_latest, variant_versions,
                variant_meta,
                live_counts, draft_counts, output_path):
    wb = Workbook()

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    blue_header = PatternFill("solid", start_color="305496")
    red_header = PatternFill("solid", start_color="C00000")
    issue_fill = PatternFill("solid", start_color="FCE4D6")
    clean_fill = PatternFill("solid", start_color="E2EFDA")
    section_fill = PatternFill("solid", start_color="595959")
    section_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)

    # ----- Sheet 1: Product-side summary -----
    ws1 = wb.active
    ws1.title = "Product-side summary"
    headers = ["product_id", "product_name", "product_template_id",
               "live_state", "draft_state", "issue_label",
               "missing_variants", "attached_variants",
               "attached_count", "should_be_attached",
               "should_be_attached_variant_names",
               "should_be_attached_variant_template_ids",
               "conflicting_template_ids",
               "should_not_attached_count",
               "stale_leftover_variants",
               "total_variants", "issue_summary", "variants_by_issue"]
    ws1.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws1.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = red_header
        c.alignment = center

    stale_fill = PatternFill("solid", start_color="F4B084")
    complete_fill = PatternFill("solid", start_color="C6EFCE")

    for pid in sorted(product_view, key=lambda x: -len(product_view[x]["missing_variants"])):
        data = product_view[pid]
        missing = data["missing_variants"]
        attached = data["attached_variants"]
        should = data["should_be_attached"]
        stale = data["stale_leftover_variants"]
        total = len(missing) + len(attached)
        attached_count = len(attached)
        should_not_attached_count = len(set(should) - set(attached))
        is_complete = total > 0 and attached_count == total and not missing

        per_source = product_latest.get(pid, {}).get("latest_per_source", {})
        live_state = per_source.get("live", {}).get("state") or "—"
        draft_state = per_source.get("draft", {}).get("state") or "—"

        # Prefer LIVE for name/template_id; fall back to DRAFT if live is empty
        live_row = per_source.get("live") or {}
        draft_row = per_source.get("draft") or {}
        product_name = (
            live_row.get("product_name")
            or draft_row.get("product_name")
            or "—"
        )
        product_template_id = (
            live_row.get("template_id")
            or draft_row.get("template_id")
            or "—"
        )

        # ----- Build the three new variant-breakdown columns -----
        # should_be_attached_variant_names: "variant_id (name)" entries, in
        # the same order as should_be_attached
        names_list = []
        tid_to_vids = {}  # template_id -> list of variant_ids (preserves order)
        for vid in should:
            meta = variant_meta.get(vid, {})
            name = meta.get("name") or "—"
            names_list.append(f"{vid} ({name})")
            tid = meta.get("template_id") or ""
            if tid:
                tid_to_vids.setdefault(tid, []).append(vid)
        variant_names_str = ", ".join(names_list) if names_list else "—"

        if len(tid_to_vids) == 0:
            variant_template_ids_str = "—"
            conflicting_template_ids_str = ""
        elif len(tid_to_vids) == 1:
            # All variants share the same template_id
            sole_tid = next(iter(tid_to_vids.keys()))
            variant_template_ids_str = sole_tid
            conflicting_template_ids_str = ""
        else:
            # Multiple distinct template_ids — leave variant_template_ids empty
            variant_template_ids_str = ""
            parts = [f"{tid} ({', '.join(vids)})" for tid, vids in tid_to_vids.items()]
            conflicting_template_ids_str = "; ".join(parts)

        codes = sorted(data["issue_summary"].keys())
        issue_label = " | ".join(codes) if codes else "OK"

        summary = "; ".join(f"{k}: {v}" for k, v in sorted(data["issue_summary"].items())) or "—"
        breakdown_lines = []
        for code in sorted(data["variants_by_issue"]):
            vids = data["variants_by_issue"][code]
            breakdown_lines.append(f"{code}: {', '.join(vids)}")
        breakdown = "\n".join(breakdown_lines) if breakdown_lines else "—"

        ws1.append([
            pid,
            product_name,
            product_template_id,
            live_state,
            draft_state,
            issue_label,
            ", ".join(missing) if missing else "—",
            ", ".join(attached) if attached else "—",
            attached_count,
            ", ".join(should) if should else "—",
            variant_names_str,
            variant_template_ids_str,
            conflicting_template_ids_str,
            should_not_attached_count,
            ", ".join(stale) if stale else "—",
            total,
            summary,
            breakdown,
        ])
        row_idx = ws1.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = left
            if is_complete:
                cell.fill = complete_fill
            elif missing:
                cell.fill = issue_fill
        # stale_leftover_variants is now column 15
        if stale and not is_complete:
            ws1.cell(row=row_idx, column=15).fill = stale_fill

    widths = [42, 50, 42, 14, 14, 40, 60, 60, 16, 60, 60, 42, 60, 22, 60, 14, 40, 80]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ----- Sheet 2: Variant details -----
    ws2 = wb.create_sheet("Variant details")
    headers2 = ["variant_id", "variant_name", "variant_template_id",
                "latest_parent", "all_versions_parent",
                "issues", "claimed_by_products"]
    ws2.append(headers2)
    for col_idx in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = blue_header
        c.alignment = center

    for vid in sorted(variant_findings):
        f = variant_findings[vid]
        ws2.append([
            vid,
            f.get("variant_name", "—"),
            f.get("variant_template_id", "—"),
            f["latest_parent"],
            f["all_versions_parent"],
            "; ".join(f["issues"]),
            ", ".join(f["claimed_by_products"]) if f["claimed_by_products"] else "—",
        ])
        row_idx = ws2.max_row
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = left
            cell.fill = issue_fill

    widths2 = [42, 50, 42, 40, 80, 80, 60]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}1"

    # ----- Sheet 3: Definitions -----
    ws3 = wb.create_sheet("Definitions")
    ws3.append(["issue code", "what it means"])
    for col_idx in (1, 2):
        c = ws3.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = blue_header
        c.alignment = center
    defs = [
        ("STALE_LATEST_EMPTY",
         "Some earlier version of this variant had a parent_id (e.g. P1), but the LATEST version "
         "has an empty parent_id. Either the link was deliberately removed without back-cleanup, or "
         "the data went stale."),
        ("PARENT_CHANGED",
         "The variant was reparented across versions and an older parent still both EXISTS as a "
         "product AND still lists this variant in its variant_id. Two products both think they own "
         "this variant — genuine conflict. Clean reparenting (where the old parent's variant_id was "
         "also updated to remove the variant) is normal lifecycle and is NOT flagged."),
        ("ORPHAN_BUT_REFERENCED",
         "The variant has NEVER had a parent_id in any version, but some product's variant_id list "
         "references it. The variant doesn't acknowledge the relationship from its side."),
        ("PARENT_NOT_FOUND",
         "A product id that this variant has ever pointed at (in latest OR any historical version) "
         "doesn't exist as a PRODUCT row in either live or draft data. If the missing parent was only "
         "claimed historically and the latest parent_id is empty, this fires INSTEAD of "
         "STALE_LATEST_EMPTY because the deleted parent is the more important problem."),
        ("PARENT_DOES_NOT_LINK_BACK",
         "The variant currently claims a parent product (in its latest parent_id), and that "
         "product exists, but the product's latest variant_id list does NOT include this variant. "
         "Live one-sided link — variant says yes, product says no."),
        ("NO_LINK_ANYWHERE",
         "The variant has no parent_id in any version AND no product references it. Fully disconnected."),
        ("PRODUCT_DROPPED_VARIANT",
         "An older version of some product listed the variant in its variant_id, but the latest "
         "version no longer does — and the variant itself never claimed this product as a parent "
         "in any version. The link existed historically only from the product's side, then was "
         "removed. Indicates a one-sided historical association that's now fully severed."),
    ]
    for code, meaning in defs:
        ws3.append([code, meaning])
        row_idx = ws3.max_row
        for col_idx in (1, 2):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = left
            cell.fill = issue_fill
        ws3.row_dimensions[row_idx].height = 45
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 100

    # ----- Sheet 4: Inventory -----
    ws4 = wb.create_sheet("Inventory")
    ws4["A1"] = "Source counts"
    ws4["A1"].font = Font(name="Arial", bold=True)
    ws4["A2"] = "live"
    ws4["B2"] = f"{live_counts['products']} PRODUCT, {live_counts['variants']} VARIANT"
    ws4["A3"] = "draft"
    ws4["B3"] = f"{draft_counts['products']} PRODUCT, {draft_counts['variants']} VARIANT"
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 50

    # ----- Sheet 5: Summary -----
    ws5 = wb.create_sheet("Summary")
    ws5["A1"] = "Total variants with issues"
    ws5["B1"] = len(variant_findings)
    ws5["A2"] = "Total products affected"
    ws5["B2"] = len(product_view)

    issue_totals = defaultdict(int)
    for f in variant_findings.values():
        for code in f["issue_codes"]:
            issue_totals[code] += 1
    row = 4
    ws5.cell(row=row, column=1, value="By issue code (counts variants once per code)").font = Font(name="Arial", bold=True)
    for code in sorted(issue_totals):
        row += 1
        ws5.cell(row=row, column=1, value=f"  {code}")
        ws5.cell(row=row, column=2, value=issue_totals[code])

    ws5.column_dimensions["A"].width = 50
    ws5.column_dimensions["B"].width = 14

    wb.save(output_path)


# ---------------- main ----------------

def main():
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("live", help="Path to LIVE CSV/TSV file")
    parser.add_argument("-d", "--draft", default=None,
                        help="Optional path to DRAFT CSV/TSV file")
    parser.add_argument("-o", "--output", default="variant_version_scan.xlsx",
                        help="Output Excel file path")
    args = parser.parse_args()

    print(f"Loading live: {args.live}")
    live_products_v, live_variants_v = load_file(args.live, "live")
    print(f"  {len(live_products_v)} products (versioned), {len(live_variants_v)} variants (versioned)")

    draft_products_v, draft_variants_v = ({}, {})
    if args.draft:
        print(f"Loading draft: {args.draft}")
        draft_products_v, draft_variants_v = load_file(args.draft, "draft")
        print(f"  {len(draft_products_v)} products, {len(draft_variants_v)} variants")

    # Merge per-item version lists from both sources
    product_versions = merge_version_dicts(live_products_v, draft_products_v)
    variant_versions = merge_version_dicts(live_variants_v, draft_variants_v)

    # Compute latest state of every product (union of latest from each source)
    product_latest = latest_product_state(product_versions)
    # Compute the single latest parent per variant (highest version across files,
    # falling back to historical parent_id, then to current product claimers,
    # then to historical product claimers if needed)
    variant_latest_parent = pick_latest_parent_per_variant(
        variant_versions, product_latest, product_versions=product_versions
    )
    # name + template_id per variant (for the product-side variant breakdown columns)
    variant_meta = build_variant_meta(variant_versions)

    print("Analyzing variants across all versions...")
    variant_findings = analyze_variants(variant_versions, product_latest,
                                        product_versions=product_versions)
    print(f"  {len(variant_findings)} variants have issues")

    print("Aggregating into product-side view...")
    product_view = aggregate_into_products(variant_findings, variant_versions,
                                           product_latest, variant_latest_parent,
                                           product_versions=product_versions)
    print(f"  {len(product_view)} products have affected variants")

    print(f"Writing {args.output}...")
    write_excel(
        product_view, variant_findings, product_latest, variant_versions,
        variant_meta=variant_meta,
        live_counts={"products": len(live_products_v), "variants": len(live_variants_v)},
        draft_counts={"products": len(draft_products_v), "variants": len(draft_variants_v)},
        output_path=args.output,
    )
    print("Done.")


if __name__ == "__main__":
    main()