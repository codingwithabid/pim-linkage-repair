"""
rebuild_variant_names.py

Rebuilds the `should_be_attached_variant_names` and `variants_by_issue` columns
in a report-linkages xlsx by looking up EVERY variant directly in the live and
draft CSVs. The CSVs are the source of truth — whatever name they hold is what
gets written, overwriting any name already in the report.

Name selection rule (identical to variant_version_scan.py):
  - Prefer the latest LIVE version with a non-empty displayName.en_us
  - Fall back to the latest DRAFT version
  - If neither has a name, the variant_id is written bare (no parens)

Affected columns per row:
  - `should_be_attached_variant_names`  — rewritten from `should_be_attached`
                                          (one entry per variant in that list)
  - `variants_by_issue`                 — each SKU gets " (Name)" appended
                                          where the CSV has a name

Usage:
    python rebuild_variant_names.py <report.xlsx> <live.csv> [--draft <draft.csv>] [-o <out.xlsx>]
"""

import argparse
import ast
import csv
import io
import json
import re
import sys
from collections import defaultdict

from openpyxl import load_workbook


# Raise csv field size limit — some rows (e.g. large request_schema JSON blobs)
# exceed the default 131,072 char limit. Use the largest value csv accepts on
# this platform.
_maxInt = sys.maxsize
while True:
    try:
        csv.field_size_limit(_maxInt)
        break
    except OverflowError:
        _maxInt = int(_maxInt / 10)


# ============ NAME-EXTRACTION LOGIC (copied verbatim from variant_version_scan.py) ============

def parse_id_list(value) -> list:
    if value is None:
        return []
    s = str(value).strip()
    if s == "" or s in ("[]", "['']", '[""]', "null", "None", "—"):
        return []
    if s.startswith("[") and s.endswith("]"):
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple)):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except (ValueError, SyntaxError):
            pass
        inner = s[1:-1]
        parts = re.split(r"\s*,\s*", inner)
        return [p.strip().strip("'\"") for p in parts if p.strip().strip("'\"")]
    parts = re.split(r"\s*,\s*", s)
    return [p.strip().strip("'\"") for p in parts if p.strip().strip("'\"")]


def _normalize_whitespace_escapes_outside_strings(s: str) -> str:
    out = []
    i, n, in_string = 0, len(s), False
    while i < n:
        c = s[i]
        if not in_string:
            if c == '\\' and i + 1 < n:
                nxt = s[i + 1]
                if nxt == 'n': out.append('\n'); i += 2; continue
                if nxt == 't': out.append('\t'); i += 2; continue
                if nxt == 'r': out.append('\r'); i += 2; continue
                if nxt == '\\': out.append('\\'); i += 2; continue
            if c == '"': in_string = True
            out.append(c); i += 1
        else:
            if c == '\\' and i + 1 < n:
                out.append(c); out.append(s[i + 1]); i += 2; continue
            if c == '"': in_string = False
            out.append(c); i += 1
    return ''.join(out)


def _repair_unescaped_quotes(s: str) -> str:
    out = []
    i, n, in_string = 0, len(s), False
    while i < n:
        c = s[i]
        if c == '\\' and i + 1 < n:
            out.append(c); out.append(s[i + 1]); i += 2; continue
        if c == '"':
            if not in_string:
                in_string = True; out.append(c); i += 1; continue
            j = i + 1
            while j < n and s[j] in ' \t\n\r': j += 1
            if j == n or s[j] in ',:}]':
                in_string = False; out.append(c)
            else:
                out.append('\\'); out.append(c)
            i += 1; continue
        out.append(c); i += 1
    return ''.join(out)


def _safe_json_loads(s: str):
    try: return json.loads(s, strict=False)
    except (ValueError, TypeError): pass
    cleaned = _normalize_whitespace_escapes_outside_strings(s)
    if cleaned != s:
        try: return json.loads(cleaned, strict=False)
        except (ValueError, TypeError): pass
    repaired = _repair_unescaped_quotes(s)
    if repaired != s:
        try: return json.loads(repaired, strict=False)
        except (ValueError, TypeError): pass
    repaired_clean = _repair_unescaped_quotes(cleaned)
    if repaired_clean != cleaned:
        try: return json.loads(repaired_clean, strict=False)
        except (ValueError, TypeError): pass
    return None


def _extract_en_us_displayname_via_regex(s: str) -> str:
    m = re.search(
        r'"displayName"[\s\S]{0,500}?"localizations"[\s\S]{0,300}?"en_us"\s*:\s*"', s
    )
    if not m: return ""
    start = m.end()
    i, n = start, len(s)
    while i < n:
        c = s[i]
        if c == '\\' and i + 1 < n:
            i += 2; continue
        if c == '"':
            j = i + 1
            while j < n:
                if s[j] in ' \t\n\r': j += 1; continue
                if s[j] == '\\' and j + 1 < n and s[j + 1] in 'ntr': j += 2; continue
                break
            if j == n or s[j] in ',}]':
                val = s[start:i]
                val = (val.replace('\\\\', '\\').replace('\\"', '"').replace("\\'", "'")
                          .replace('\\n', ' ').replace('\\t', ' ').replace('\\r', ' '))
                return " ".join(val.split())
            i += 1; continue
        i += 1
    return ""


def parse_product_name(raw: str) -> str:
    if not raw: return ""
    s = raw.strip()
    if not s or s in ("null", "None"): return ""
    obj = _safe_json_loads(s)
    if isinstance(obj, dict):
        display = obj.get("displayName") or {}
        if isinstance(display, dict):
            locs = display.get("localizations") or {}
            if isinstance(locs, dict):
                val = locs.get("en_us")
                if val:
                    cleaned = str(val)
                    cleaned = (cleaned.replace('\\\\', '\\').replace('\\"', '"').replace("\\'", "'")
                                      .replace("\\r\\n", " ").replace("\\n", " ")
                                      .replace("\\t", " ").replace("\\r", " "))
                    return " ".join(cleaned.split())
    return _extract_en_us_displayname_via_regex(s)


def _looks_like_backslash_escaped(path: str) -> bool:
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            f.readline()
            sample = f.read(10000)
        return '\\"' in sample
    except OSError:
        return False


def _normalize_backslash_escaped_csv(path: str):
    out = []
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        content = f.read()
    in_quotes = False
    i, n = 0, len(content)
    while i < n:
        c = content[i]
        if c == '"':
            out.append('"'); in_quotes = not in_quotes; i += 1; continue
        if c == '\\' and in_quotes and i + 1 < n:
            nxt = content[i + 1]
            if nxt == '"': out.append('""'); i += 2; continue
            if nxt == '\\': out.append('\\'); i += 2; continue
            out.append(c); i += 1; continue
        out.append(c); i += 1
    return io.StringIO("".join(out))


# ============ CSV LOADING ============

def load_variant_names(path: str, source_label: str) -> dict:
    """Return {variant_id: [(version, source, name), ...]} for VARIANT rows."""
    if not path:
        return {}
    use_backslash = _looks_like_backslash_escaped(path)
    if use_backslash:
        f = _normalize_backslash_escaped_csv(path)
        first_line = f.readline(); f.seek(0)
        dialect = csv.excel
    else:
        f = open(path, "r", encoding="utf-8-sig", newline="")
        first_line = f.readline(); f.seek(0)
        dialect = csv.excel_tab if "\t" in first_line else csv.excel

    variants = defaultdict(list)
    try:
        reader = csv.DictReader(f, dialect=dialect)
        if "item_id" not in (reader.fieldnames or []):
            sys.exit(f"ERROR: {path} missing 'item_id' column (got {reader.fieldnames})")
        has_version = "version" in (reader.fieldnames or [])
        for row in reader:
            scope = (row.get("item_scope") or "").strip().upper()
            if scope != "VARIANT":
                continue
            vid = (row.get("item_id") or "").strip()
            if not vid:
                continue
            try:
                version = int(row["version"]) if has_version else 1
            except (ValueError, TypeError):
                version = 1
            name = parse_product_name(row.get("request_schema", ""))
            variants[vid].append((version, source_label, name))
    finally:
        try: f.close()
        except Exception: pass
    return variants


def build_name_map(live_variants: dict, draft_variants: dict) -> dict:
    """Apply live-first / draft-fallback rule and return {vid: name}."""
    all_vids = set(live_variants) | set(draft_variants)
    out = {}
    for vid in all_vids:
        live_rows = list(live_variants.get(vid, []))
        draft_rows = list(draft_variants.get(vid, []))

        def latest_nonempty(rows):
            for v, s, n in sorted(rows, key=lambda r: r[0], reverse=True):
                if n: return n
            return ""

        name = latest_nonempty(live_rows) or latest_nonempty(draft_rows)
        if name:
            out[vid] = name
    return out


# ============ XLSX REBUILD ============

def format_name_entry(vid: str, name: str) -> str:
    return f"{vid} ({name})" if name else vid


def rebuild_issues_cell(issues_cell: str, name_map: dict) -> str:
    """Rewrite each 'CODE: sku1, sku2, ...' line so every SKU gets ' (Name)'
    appended when the CSV has a name."""
    if not issues_cell:
        return issues_cell
    out_lines = []
    for line in str(issues_cell).splitlines():
        if ":" not in line:
            out_lines.append(line); continue
        label, _, skus_part = line.partition(":")
        skus = [s.strip() for s in skus_part.split(",") if s.strip()]
        # Strip any pre-existing " (...)" suffix
        clean = []
        for s in skus:
            m = re.match(r"^([^()]+?)\s*\(.*\)$", s)
            clean.append(m.group(1).strip() if m else s)
        annotated = [format_name_entry(sku, name_map.get(sku, "")) for sku in clean]
        out_lines.append(f"{label}: {', '.join(annotated)}")
    return "\n".join(out_lines)


def rebuild_workbook(xlsx_path: str, name_map: dict, out_path: str) -> dict:
    wb = load_workbook(xlsx_path)
    stats = {
        "rows_processed": 0,
        "names_written": 0,
        "names_found_in_csv": 0,
        "missing_from_csv": set(),
        "issues_rewritten": 0,
    }

    for sheet in wb.worksheets:
        header = {c.value: c.column for c in sheet[1] if c.value}
        sba_col = header.get("should_be_attached")
        names_col = header.get("should_be_attached_variant_names")
        issues_col = header.get("variants_by_issue")
        if not sba_col or not names_col:
            continue

        for row in range(2, sheet.max_row + 1):
            sba = sheet.cell(row=row, column=sba_col).value
            if sba and str(sba).strip() not in ("", "—", "-"):
                sba_ids = [x.strip() for x in str(sba).split(",") if x.strip()]
                stats["rows_processed"] += 1
                entries = []
                for vid in sba_ids:
                    name = name_map.get(vid, "")
                    entries.append(format_name_entry(vid, name))
                    if name:
                        stats["names_found_in_csv"] += 1
                    else:
                        stats["missing_from_csv"].add(vid)
                    stats["names_written"] += 1
                sheet.cell(row=row, column=names_col).value = ", ".join(entries)

            # Rewrite variants_by_issue using the GLOBAL name_map so SKUs that
            # only appear here (not in should_be_attached) also get their names.
            if issues_col:
                cur = sheet.cell(row=row, column=issues_col).value
                if cur:
                    new = rebuild_issues_cell(str(cur), name_map)
                    if new != cur:
                        sheet.cell(row=row, column=issues_col).value = new
                        stats["issues_rewritten"] += 1
                        # Track any SKU here that's not in the CSV
                        for line in str(cur).splitlines():
                            if ":" not in line:
                                continue
                            _, _, skus_part = line.partition(":")
                            for s in skus_part.split(","):
                                s = s.strip()
                                m = re.match(r"^([^()]+?)\s*\(.*\)$", s)
                                sku = m.group(1).strip() if m else s
                                if sku and sku not in name_map:
                                    stats["missing_from_csv"].add(sku)

    wb.save(out_path)
    return stats


# ============ MAIN ============

def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("report", help="Path to report-linkages xlsx")
    p.add_argument("live", help="Path to LIVE product_item CSV")
    p.add_argument("-d", "--draft", default=None, help="Optional path to DRAFT CSV")
    p.add_argument("-o", "--output", default=None,
                   help="Output xlsx path (default: <report>-rebuilt.xlsx)")
    args = p.parse_args()

    out_path = args.output or args.report.replace(".xlsx", "-rebuilt.xlsx")

    print(f"Loading live CSV: {args.live}")
    live_vars = load_variant_names(args.live, "live")
    print(f"  {len(live_vars)} variants in live")

    draft_vars = {}
    if args.draft:
        print(f"Loading draft CSV: {args.draft}")
        draft_vars = load_variant_names(args.draft, "draft")
        print(f"  {len(draft_vars)} variants in draft")

    name_map = build_name_map(live_vars, draft_vars)
    print(f"Combined name lookup: {len(name_map)} variants with names\n")

    print(f"Rebuilding: {args.report}")
    stats = rebuild_workbook(args.report, name_map, out_path)
    print(f"  Rows processed:           {stats['rows_processed']}")
    print(f"  Total name entries written: {stats['names_written']}")
    print(f"  Names found in CSV:       {stats['names_found_in_csv']}")
    print(f"  Variants missing from CSV: {len(stats['missing_from_csv'])}")
    print(f"  variants_by_issue cells rewritten: {stats['issues_rewritten']}")
    if stats["missing_from_csv"]:
        sample = sorted(stats["missing_from_csv"])[:10]
        print(f"  Sample missing: {sample}")
    print(f"\nSaved: {out_path}")


if __name__ == "__main__":
    main()