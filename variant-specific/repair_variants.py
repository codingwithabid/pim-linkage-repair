#!/usr/bin/env python3
"""
Variant payload repair script.

Pipeline:
  1. Read all source sheets from the linkages report:
        - 'Variant Specific(+ve)'  ('variants having issues' column)
        - 'Product Specific(+ve)'  ('variants links to be added' column)
        - 'Product to verify(-ve)' ('links to be fixed' column)
     Parse each variants cell. Cells look like:
        "CODE1: id (name), id (name)  CODE2: id (name) ..."
     Only variants tagged with STALE_LATEST_EMPTY or ORPHAN_BUT_REFERENCED
     are kept; other issue codes (PARENT_CHANGED, PARENT_DOES_NOT_LINK_BACK,
     PRODUCT_DROPPED_VARIANT) appear in the sheets too but represent
     different repair operations that this script doesn't handle, so their
     variants are dropped at parse time.

  2. Detect duplicates (variants claimed by 2+ different products across
     any of the sheets). Write duplicate_variants.json with each duplicate's
     per-product info (state/status derived from the sheet's draftState/
     liveState columns), plus null placeholders for variant-level
     state/status/siteId that the next step fills in.

  3. Enrich duplicates. For each duplicate variant, GET the API (same
     multi-site / multi-view retry as the main fetch) and update the entry
     with variant-level state, status, and siteId. No payload file written
     — duplicates need a human decision about which product owns the
     variant before they can be safely rebound. Rewrite duplicate_variants.json
     with the enriched data.

  4. Snapshot product states (pre-repair). Sheet-only — no API. For every
     productId referenced by any kept occurrence (across clean map AND
     duplicates), derive "DRAFT" or "PUBLISHED" from the sheet's
     draft_state / live_state columns using the same rule as the
     per-product status in duplicate_variants.json (draft beats live).
     Write product_states_pre_repair.json:
        { productId: { "state": "DRAFT" | "PUBLISHED" | null } }
     Purpose: after Step 6's variant PUTs land, some products may need a
     state change (e.g. republish) to make the variant edits visible. This
     snapshot is the only authoritative record of what state each product
     was in BEFORE the run started, so a follow-up step can decide which
     products need which transitions.

  5. Build the clean variant_map.json (excludes duplicates). Initially:
        { variantId: { "productId":         <from sheet>,
                       "productName":       <from sheet>,
                       "productTemplateId": <from sheet>,
                       "variantName":       <from sheet>,
                       "status":            null,   # populated by step 6
                       "state":             null,   # populated by step 6
                       "siteIds":           null,   # populated by step 6 (array)
                       "issues":            [issue codes] } }
     The map is written to disk now so it exists even if the run dies
     mid-fetch.

  6. For each variant in the clean map: GET the API (multi-site +
     multi-view retry), then "repair" the response:
        - productIds  → [<product_id from sheet>]
        - status      → "DRAFT" if served by LATEST_INCLUDE_DRAFT
                        "PUBLISHED" if served by LATEST
        - state       → from API response (kept in map; stripped from file)
        - strip server-managed fields (links, siteIds, hasDraft, hasReady,
          hasPublished, syncedProductType, status, state, dateCreated,
          dateModified)
     Write the repaired payload to:
        <output_root>/<productTemplateId>/<siteId>/<variantId>.json
     The siteId subfolder is the site that actually returned the data
     (winning site from the multi-site retry). This lets the eventual
     update phase iterate `<output>/*/<siteId>/*.json` and PUT each batch
     back to the same site it came from — no fallback logic needed during
     update, because the file's location already encodes the right target.
     Rewrite variant_map.json with the API state/status/siteId merged in.
     Variants the API didn't return keep their null placeholders — null =
     "fetch missed this one" (see errors.log for the reason).

Auth:
  Headers are set per the user's curl snippet. The X-Upstart-Session-Id is
  a placeholder — override via the API_SESSION_ID env var.

Usage:
  python repair_variants.py \
      --input  /mnt/user-data/uploads/report-linkages.xlsx \
      --output ./out \
      --api-host https://nochannel-test-1-api.nochannel-test.upstart.team
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
#
# Three loggers, three files:
#
#   repair        → run.log         everything (INFO+), rotating, durable record
#                                   also mirrored to stderr at INFO+
#   repair.ok     → success.log     one structured line per successful variant
#   repair.err    → errors.log      every WARNING/ERROR with full context
#
# Why three files instead of one with levels?
#   - Ops people want to grep "did variant X succeed today?" → success.log
#   - On-call wants "what broke last night?" → errors.log (no noise)
#   - Auditors / debuggers want the full timeline → run.log
#
# Format: timestamp + level + logger + message. Pipe-delimited key=value
# pairs for the per-variant lines so they're easy to grep and easy to parse
# with awk/cut without bringing in jsonlines tooling.

LOG_FMT = "%(asctime)s | %(levelname)-7s | %(name)-10s | %(message)s"
LOG_DATEFMT = "%Y-%m-%d %H:%M:%S"


def setup_logging(log_dir: Path, verbose: bool = False) -> tuple[logging.Logger, logging.Logger, logging.Logger]:
    """
    Configure the three loggers and return them as (main, success, error).

    Files:
      <log_dir>/run.log       — rotating (10MB × 5 backups), all messages
      <log_dir>/success.log   — rotating, INFO-only from repair.ok
      <log_dir>/errors.log    — rotating, WARNING+ from repair.err

    Console: INFO+ to stderr so the run is visible in real time.
    """
    log_dir.mkdir(parents=True, exist_ok=True)
    formatter = logging.Formatter(LOG_FMT, datefmt=LOG_DATEFMT)

    def _rotating(path: Path, level: int) -> RotatingFileHandler:
        # 10MB per file, keep 5 backups → up to 50MB history per stream.
        # Production-typical defaults; tune via env vars if you outgrow them.
        h = RotatingFileHandler(
            path, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8"
        )
        h.setLevel(level)
        h.setFormatter(formatter)
        return h

    # --- root logger for the project: catches everything ---
    main = logging.getLogger("repair")
    main.setLevel(logging.DEBUG if verbose else logging.INFO)
    main.propagate = False
    # Clear stale handlers if someone calls setup_logging twice in one process.
    main.handlers.clear()
    main.addHandler(_rotating(log_dir / "run.log", logging.DEBUG if verbose else logging.INFO))

    console = logging.StreamHandler(sys.stderr)
    console.setLevel(logging.INFO)
    console.setFormatter(formatter)
    main.addHandler(console)

    # --- success-only logger: never propagates, dedicated file ---
    success = logging.getLogger("repair.ok")
    success.setLevel(logging.INFO)
    success.propagate = False
    success.handlers.clear()
    success.addHandler(_rotating(log_dir / "success.log", logging.INFO))

    # --- error-only logger: dedicated file + bubbles up to main/console ---
    error = logging.getLogger("repair.err")
    error.setLevel(logging.WARNING)
    error.propagate = True  # so errors ALSO show up in run.log + console
    error.handlers.clear()
    error.addHandler(_rotating(log_dir / "errors.log", logging.WARNING))
    # ↑ propagate=True means parent ("repair") will also handle the record,
    # so it lands in run.log and on the console. That's intentional — errors
    # should be visible everywhere, not hidden in errors.log alone.

    main.info("logging initialised: dir=%s verbose=%s", log_dir, verbose)
    return main, success, error


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

# Site IDs from the user's spec. The primary goes on the request header;
# fallbacks are tried in order if the primary returns 404/empty.
PRIMARY_SITE_ID = "201cb789-4198-488b-a5eb-4e7df0fb4bee"
FALLBACK_SITE_IDS = [
    "8d3ea3bc-f65b-4227-9fa6-6fae40e4575a",
    "fbfcd92b-d271-4002-9163-4f84986b41be",
]

# View precedence: try LATEST_INCLUDE_DRAFT first, then LATEST.
VIEW_PRECEDENCE = ["LATEST_INCLUDE_DRAFT", "LATEST"]

# Known issue codes from the Definitions sheet. Used so the parser knows what
# a "code header" looks like vs. a name fragment that happens to contain a colon.
# All issue codes that appear in any source sheet. The parser needs to
# recognize each as a code-header so it knows where one block ends and the
# next begins (a cell like "PARENT_CHANGED: X (...)  STALE_LATEST_EMPTY: Y
# (...)" must split on both). Whether a given code's variants make it into
# the output is a separate question, controlled by KEEP_ISSUE_CODES below.
ISSUE_CODES = {
    "STALE_LATEST_EMPTY",
    "PARENT_CHANGED",
    "ORPHAN_BUT_REFERENCED",
    "PARENT_NOT_FOUND",
    "PARENT_DOES_NOT_LINK_BACK",
    "NO_LINK_ANYWHERE",
    "PRODUCT_DROPPED_VARIANT",
}

# Regex that splits the "variants having issues" cell into (code, body) blocks.
# Cells look like:
#   "STALE_LATEST_EMPTY: ID1 (name), ID2 (name)  PRODUCT_DROPPED_VARIANT: ID3 ..."
# We split on `<CODE>:` where CODE is one of the known issue codes.
_CODE_RE = re.compile(
    r"\b(" + "|".join(sorted(ISSUE_CODES, key=len, reverse=True)) + r")\s*:\s*"
)

# Inside a block body, IDs look like `<id> (<name>)`. A name can contain commas
# and parentheses-free text. We match an ID, then a parenthesised name.
# ID grammar: UUID, or letters/digits/dashes/underscores (e.g. BL-JLHPAQFL,
# prod1560004). We accept anything that isn't whitespace, comma, or paren.
_ID_NAME_RE = re.compile(r"([^\s,()]+)\s*\(([^)]*)\)")


# ---------------------------------------------------------------------------
# Step 1+2: parse the sheet
# ---------------------------------------------------------------------------

def parse_issues_cell(cell: str) -> list[tuple[str, str, str]]:
    """
    Parse one 'variants having issues' cell.

    Returns a list of (variant_id, variant_name, issue_code) tuples.
    Names are kept for traceability but aren't strictly required downstream.
    """
    if not cell or cell == "—":
        return []

    out: list[tuple[str, str, str]] = []

    # Find all code markers, then for each marker take the text up to the
    # next marker (or end-of-string) as that block's body.
    matches = list(_CODE_RE.finditer(cell))
    if not matches:
        return []

    for i, m in enumerate(matches):
        code = m.group(1)
        body_start = m.end()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(cell)
        body = cell[body_start:body_end]

        for id_match in _ID_NAME_RE.finditer(body):
            variant_id = id_match.group(1).strip().rstrip(",")
            variant_name = id_match.group(2).strip()
            if variant_id:
                out.append((variant_id, variant_name, code))

    return out


def pick_state(live_state: Any, draft_state: Any) -> str:
    """
    Pick a single 'state' for the variant. Prefer live_state when present,
    otherwise draft_state. '—' or None means missing.
    """
    def _clean(v: Any) -> str | None:
        if v is None:
            return None
        s = str(v).strip()
        if not s or s == "—":
            return None
        return s

    return _clean(live_state) or _clean(draft_state) or "UNKNOWN"


# Which sheets we read, and what their "variant list" column is called.
# Each sheet uses a different header for the same kind of data:
SHEET_VARIANTS_COLUMN = {
    "Variant Specific(+ve)":  "variants having issues",
    "Product Specific(+ve)":  "variants links to be added",
    "Product to verify(-ve)": "links to be fixed",
}

# Only variants tagged with one of these issue codes go into variant_map /
# duplicate_variants. Other codes (PARENT_CHANGED, PARENT_DOES_NOT_LINK_BACK,
# PRODUCT_DROPPED_VARIANT) appear in the source sheets too but represent
# different repair operations that this script doesn't handle. They still
# need to be recognized by the parser (so its regex knows where one
# code-block ends and the next begins), they just don't produce occurrences.
KEEP_ISSUE_CODES = {"STALE_LATEST_EMPTY", "ORPHAN_BUT_REFERENCED"}


def collect_all_occurrences(
    xlsx_path: Path,
) -> list[dict[str, Any]]:
    """
    Scan ALL THREE source sheets and return every (variant, product)
    occurrence whose issue code is in KEEP_ISSUE_CODES.

    Why all three sheets:
      - Variant Specific(+ve):  "variants having issues" column, codes:
                                STALE_LATEST_EMPTY, ORPHAN_BUT_REFERENCED
      - Product Specific(+ve):  "variants links to be added" column, codes:
                                PARENT_CHANGED, PARENT_DOES_NOT_LINK_BACK,
                                STALE_LATEST_EMPTY (intent: link variants TO
                                this product)
      - Product to verify(-ve): "links to be fixed" column, codes:
                                PRODUCT_DROPPED_VARIANT, STALE_LATEST_EMPTY,
                                etc. (intent: the listed product_id IS the
                                rebinding target — same semantics as the
                                others for our purposes per user direction)

    Why the filter:
      Other issue codes mean different repair operations. Including them
      would mix unrelated work into the same payload-rewrite output.

    Each item: {
        "variantId": ..., "variantName": ...,
        "productId": ..., "productName": ...,
        "productTemplateId": ...,
        "state": ..., "liveState": ..., "draftState": ...,
        "issue": <STALE_LATEST_EMPTY | ORPHAN_BUT_REFERENCED>,
        "sourceSheet": <which sheet this came from>,
    }
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    occurrences: list[dict[str, Any]] = []
    log = logging.getLogger("repair")

    for sheet_name, var_col in SHEET_VARIANTS_COLUMN.items():
        if sheet_name not in wb.sheetnames:
            # Don't hard-fail — older report formats may lack a sheet.
            log.warning("sheet %r not found in workbook; skipping", sheet_name)
            continue

        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            log.warning("sheet %r is empty; skipping", sheet_name)
            continue
        idx = {h: i for i, h in enumerate(headers)}

        # All three sheets share these columns; only the variant-list column
        # name differs. Validate each before reading.
        needed = [
            var_col, "product_id", "product_name",
            "product_template_id", "live_state", "draft_state",
        ]
        missing = [c for c in needed if c not in idx]
        if missing:
            log.warning(
                "sheet %r missing columns %s; skipping",
                sheet_name, missing,
            )
            continue

        kept = filtered_out = 0
        for row in rows:
            cell = row[idx[var_col]]
            product_id = row[idx["product_id"]]
            product_name = row[idx["product_name"]]
            product_template_id = row[idx["product_template_id"]]
            live_state = row[idx["live_state"]]
            draft_state = row[idx["draft_state"]]

            if not cell or not product_id:
                continue

            state = pick_state(live_state, draft_state)

            for variant_id, variant_name, issue_code in parse_issues_cell(str(cell)):
                if issue_code not in KEEP_ISSUE_CODES:
                    # Other codes represent different repair operations
                    # (parent rebinding, dropped-variant cleanup, etc.) and
                    # are out of scope for the payload rewrite this script
                    # produces. Count for visibility, then skip.
                    filtered_out += 1
                    continue
                kept += 1
                occurrences.append({
                    "variantId": variant_id,
                    "variantName": variant_name,
                    "productId": product_id,
                    "productName": product_name,
                    "productTemplateId": product_template_id,
                    "state": state,
                    "liveState": str(live_state) if live_state is not None else None,
                    "draftState": str(draft_state) if draft_state is not None else None,
                    "issue": issue_code,
                    "sourceSheet": sheet_name,
                })
        log.info(
            "sheet=%r kept=%d filtered_out=%d (codes not in %s)",
            sheet_name, kept, filtered_out, sorted(KEEP_ISSUE_CODES),
        )

    return occurrences


def _is_sheet_value_set(v: Any) -> bool:
    """The sheet uses '—' (em dash) for missing/empty state values."""
    if v is None:
        return False
    s = str(v).strip()
    return bool(s) and s != "—"


def derive_product_status(live_state: Any, draft_state: Any) -> str | None:
    """
    Derive a variant-status hint from a product's sheet state columns.

    Rule (per user spec): draft beats live.
      - draft_state is set       → status="DRAFT"
        (the product has an editable draft revision, so the variant payload
         we'd care about is the draft one)
      - else liveState is set    → status="PUBLISHED"
      - else                     → None
    """
    if _is_sheet_value_set(draft_state):
        return "DRAFT"
    if _is_sheet_value_set(live_state):
        return "PUBLISHED"
    return None


def detect_duplicates(
    occurrences: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """
    Find variant IDs that are claimed by more than one DISTINCT product.

    Returns:
        {
          variant_id: {
            "variantName": ...,
            "products": [
              {"productId": ..., "productName": ..., "productTemplateId": ...,
               "state":  <product's state from the sheet>,
               "status": "DRAFT" | "PUBLISHED" | None    (sheet-derived),
               "issues": [<distinct issue codes>]},
              ...
            ],
            # Populated post-fetch from the API (one extra GET per duplicate):
            "state":  null,   # → from the variant's API response root
            "status": null,   # → from which view served it
            "siteId": null,   # → which site returned the data
          },
          ...
        }

    A variant flagged twice by the SAME product (e.g. listed under two issue
    codes within one row) is NOT a duplicate — it's just a multi-issue
    variant. Those get their issue codes merged but stay in the clean map.
    """
    # variant_id -> product_id -> aggregated product info
    by_variant: dict[str, dict[str, dict[str, Any]]] = {}
    variant_names: dict[str, str] = {}

    for occ in occurrences:
        vid = occ["variantId"]
        pid = occ["productId"]
        variant_names.setdefault(vid, occ["variantName"])

        prods = by_variant.setdefault(vid, {})
        if pid not in prods:
            prods[pid] = {
                "productId": pid,
                "productName": occ["productName"],
                "productTemplateId": occ["productTemplateId"],
                # Sheet-sourced product state. `state` is the picked one
                # (liveState else draftState); `status` is the DRAFT/PUBLISHED
                # hint derived from which sheet column was populated.
                "state": occ["state"],
                "status": derive_product_status(occ["liveState"], occ["draftState"]),
                "issues": [occ["issue"]],
            }
        else:
            # Same variant, same product, another issue code → just record it.
            existing = prods[pid]["issues"]
            if occ["issue"] not in existing:
                existing.append(occ["issue"])

    duplicates: dict[str, dict[str, Any]] = {}
    for vid, prods in by_variant.items():
        if len(prods) > 1:
            duplicates[vid] = {
                "variantName": variant_names[vid],
                "products": list(prods.values()),
                # Variant-level (NOT per-product). Filled in by the post-fetch
                # enrichment pass that calls the API once per duplicate.
                "state": None,
                "status": None,
                "siteId": None,
            }
    return duplicates


def build_variant_map(
    occurrences: list[dict[str, Any]],
    duplicate_ids: set[str],
) -> dict[str, dict[str, Any]]:
    """
    Build the clean {variant_id: {...}} map, EXCLUDING duplicates.

    Duplicates are reported separately (see detect_duplicates) and skipped
    here so the API-fetch / repair pipeline doesn't pick a wrong product
    binding for them. Re-run those by hand after a human decides the owner.
    """
    variant_map: dict[str, dict[str, Any]] = {}

    for occ in occurrences:
        vid = occ["variantId"]
        if vid in duplicate_ids:
            continue
        if vid in variant_map:
            # Same variant, same product, multiple issue codes — merge codes.
            existing_issues = variant_map[vid]["issues"]
            if occ["issue"] not in existing_issues:
                existing_issues.append(occ["issue"])
            continue

        variant_map[vid] = {
            "productId": occ["productId"],
            "productName": occ["productName"],
            "productTemplateId": occ["productTemplateId"],
            "variantName": occ["variantName"],
            # status, state, siteIds, fetchedViaSite are placeholders here —
            # populated after the API fetch:
            #   status:         LATEST_INCLUDE_DRAFT → DRAFT, LATEST → PUBLISHED
            #   state:          `state` attribute at root of API response
            #   siteIds:        the full `siteIds` array from the variant's
            #                   payload (source of truth — a variant can be
            #                   on multiple sites)
            #   fetchedViaSite: which site we actually fetched from (for
            #                   audit; the primary may have served it, or
            #                   we fell through to a fallback)
            # If the fetch never succeeds on any site, all four stay null.
            "status": None,
            "state": None,
            "siteIds": None,
            "fetchedViaSite": None,
            "issues": [occ["issue"]],
        }

    return variant_map


def build_product_state_snapshot(
    occurrences: list[dict[str, Any]],
) -> dict[str, dict[str, str | None]]:
    """
    Build a one-shot snapshot of every product's state BEFORE any variant
    PUT touches the system.

    Why: after the main fetch (Step 5) PUTs variant payloads back, some
    products may need a state change (e.g. publish/unpublish) to make those
    variant edits visible. To know which products needed which transitions,
    you have to know what state they were in BEFORE the run started.
    Knowing the original state lets a follow-up step decide things like
    "this product was in DRAFT before the repair, keep it in DRAFT"
    or "this was PUBLISHED, republish after the variant change."

    Source: SHEET ONLY. No API calls. The sheet's draft_state / live_state
    columns are the same data we already use to derive per-product status
    in duplicate_variants.json, just rolled up to one row per productId.

    Sanity check (run on the user's report): every productId has exactly
    one (live_state, draft_state) pair across all sheets — zero
    inconsistencies. So "first occurrence wins" is safe; the choice
    would be the same regardless.

    Shape:
        { productId: { "state": "DRAFT" | "PUBLISHED" | None } }

    Includes products referenced by EITHER the clean map or the duplicates
    (i.e. every productId that appears in `occurrences`, which is built
    from both clean and duplicate variants before they're split).
    """
    snapshot: dict[str, dict[str, str | None]] = {}
    for occ in occurrences:
        pid = occ["productId"]
        if pid in snapshot:
            continue
        snapshot[pid] = {
            "state": derive_product_status(occ["liveState"], occ["draftState"]),
        }
    return snapshot


# ---------------------------------------------------------------------------
# Step 3: fetch from the API
# ---------------------------------------------------------------------------

def make_session(
    session_id: str,
    cookie_header: str | None = None,
) -> requests.Session:
    """
    Build the requests session with the headers the API expects.

    Note: X-UpStart-Site is NOT set on the session — it's passed per request
    so the same Session can be reused across multiple sites without races
    between worker threads. (Mutating session.headers from concurrent threads
    would be a bug; per-request kwargs are thread-safe.)

    `cookie_header`, if given, is the raw `Cookie:` header value copied from
    a working browser request (DevTools → Network → request → "Copy as cURL"
    and pull out the `-b '...'` argument). Most Upstart endpoints rely on
    session cookies for auth, not just X-Upstart-Session-Id — if cookies are
    missing, the API gateway returns the SPA's HTML shell on every request
    instead of variant JSON. (See fetch_variant: kind=non_json content_type=text/html
    is the canonical signature of this failure.)
    """
    s = requests.Session()
    headers = {
        "X-Upstart-Tenant": "denvermattress",
        "Referer": "https://denvermattress.nochannel-test-1.nochannel-test.upstart.team/",
        "X-Upstart-Session-Id": session_id,
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
    }
    if cookie_header:
        headers["Cookie"] = cookie_header
    s.headers.update(headers)
    return s


def fetch_variant(
    session: requests.Session,
    api_host: str,
    variant_id: str,
    site_ids: list[str],
    timeout: float = 30.0,
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    """
    Fetch one variant, trying each site in `site_ids` (in order), and for
    each site each view in VIEW_PRECEDENCE.

    Stops at the first (site, view) combination that returns usable JSON.

    Returns:
        (payload, view_used, site_used)   on success
        (None,    None,      None)        if no (site, view) returned data

    Logging policy (per your request):
      - per-site / per-view attempts log at DEBUG only (see --verbose).
        With ~15k variants and several sites/views each, WARNINGs on every
        failed attempt would drown errors.log.
      - the FINAL not-found verdict (with all tried sites) is emitted by
        process_variant, not here. fetch_variant stays purely about
        "did any attempt work?".
      - the one exception: kind=auth_html stays at ERROR. That's a
        configuration problem the operator needs to see immediately,
        regardless of retries — if cookies are wrong, every site will
        return the same SPA shell and silently quiet logging would hide it.
    """
    main_log = logging.getLogger("repair")
    err_log = logging.getLogger("repair.err")

    url = f"{api_host.rstrip('/')}/pim/variants/{variant_id}"

    for site_id in site_ids:
        for view in VIEW_PRECEDENCE:
            try:
                r = session.get(
                    url,
                    params={"view": view},
                    headers={"X-UpStart-Site": site_id},
                    timeout=timeout,
                )
            except requests.RequestException as e:
                main_log.debug(
                    "attempt_fail variantId=%s siteId=%s view=%s kind=network detail=%r",
                    variant_id, site_id, view, str(e),
                )
                continue

            if r.status_code == 200:
                try:
                    data = r.json()
                except ValueError:
                    content_type = r.headers.get("Content-Type", "")
                    if "text/html" in content_type.lower():
                        # Auth/config problem — must surface even with quiet
                        # logging because no amount of retries will fix it.
                        err_log.error(
                            "fetch_error variantId=%s siteId=%s view=%s kind=auth_html "
                            "status=200 content_type=%r body_len=%d "
                            "HINT='Server returned HTML (likely the SPA shell). "
                            "Auth is probably missing or expired. Pass --cookie or "
                            "set API_COOKIE env var with the Cookie header from a "
                            "working browser request.'",
                            variant_id, site_id, view, content_type, len(r.content),
                        )
                    else:
                        main_log.debug(
                            "attempt_fail variantId=%s siteId=%s view=%s kind=non_json "
                            "status=200 content_type=%r body_len=%d",
                            variant_id, site_id, view, content_type, len(r.content),
                        )
                    continue
                if data:
                    main_log.debug(
                        "attempt_ok variantId=%s siteId=%s view=%s",
                        variant_id, site_id, view,
                    )
                    return data, view, site_id
                main_log.debug(
                    "attempt_empty variantId=%s siteId=%s view=%s",
                    variant_id, site_id, view,
                )
                continue

            if r.status_code in (404, 204):
                main_log.debug(
                    "attempt_miss variantId=%s siteId=%s view=%s status=%s",
                    variant_id, site_id, view, r.status_code,
                )
                continue

            # Unexpected non-2xx — still debug-level per policy, but include
            # body so --verbose runs can diagnose without rerunning.
            main_log.debug(
                "attempt_fail variantId=%s siteId=%s view=%s kind=http_status "
                "status=%s body=%r",
                variant_id, site_id, view, r.status_code, r.text[:200],
            )

    return None, None, None


# ---------------------------------------------------------------------------
# Step 4+5: repair and write
# ---------------------------------------------------------------------------

_VIEW_TO_STATUS = {
    "LATEST_INCLUDE_DRAFT": "DRAFT",
    "LATEST": "PUBLISHED",
}


# Fields stripped from the API response before the payload is written to disk.
# These are all server-managed: either set by the platform on read (HATEOAS
# links, derived booleans, computed timestamps) or owned by separate APIs
# (state/status are mutated via change-state, not via the variant PUT).
# Including any of them in a write/update request would either be ignored or
# rejected with 400.
#
# Grouped by category for readability. Order within the tuple doesn't matter
# — repair_payload iterates and pops; missing keys are silently skipped.
_STRIP_FROM_PAYLOAD = (
    # --- HATEOAS / envelope ---
    "links",                  # link envelope, always regenerated on read

    # --- Identity (per user direction) ---
    # NOTE: stripping `id` is unusual — most PUT endpoints require it in
    # the body. Keep this in mind if the eventual update step starts
    # returning 4xx; the platform may need `id` re-added.
    "id",

    # --- Platform-derived booleans ---
    "hasDraft",               # derived from existence of a draft revision
    "hasReady",               # derived
    "hasPublished",           # derived

    # --- Platform-derived metadata ---
    # `siteIds` is the source of truth for which sites a variant lives on,
    # BUT it's encoded in the file's folder path now (e.g.
    # tpl-id/siteA,siteB/variantId.json). Stripping it from the payload
    # body keeps the saved JSON minimal and avoids two sources of truth
    # going out of sync. push_variants.py reads the folder name to know
    # which sites to PUT to; it doesn't need siteIds in the body.
    "siteIds",
    "syncedProductType",      # derived from productTemplate sync status
    "productTemplateVersion", # incremented by the platform on template change
    "productType",            # derived from productTemplate
    "sellable",               # derived state, not directly writable
    "status",                 # owned by change-state API, not by variant PUT
    "state",                  # owned by change-state API, not by variant PUT

    # --- Server-set timestamps ---
    # The platform sets/updates these on every read/write. Including stale
    # values in a PUT either gets ignored or causes optimistic-concurrency
    # 4xx depending on the endpoint. Drop them all. Variant names overlap
    # with product-style names (createdDate vs dateCreated etc.) so we
    # cover both spellings.
    "dateCreated", "createdDate",
    "dateModified", "modifiedDate", "lastModifiedDate",
    "datePublished", "publishDate",
    "dateArchived", "archiveDate",
)


def repair_payload(
    payload: dict[str, Any],
    product_id: str,
    view: str,
) -> dict[str, Any]:
    """
    Turn an API GET response into a clean payload safe to write to disk for
    replay/update use:
      - productIds → [product_id]   (the rebinding intent)
      - strip server-managed fields listed in _STRIP_FROM_PAYLOAD (see comment
        on that constant for why each one is removed)

    `view` is no longer used for status (status is stripped), but stays in
    the signature so callers don't have to change and so we can later derive
    other fields from it if needed.

    Operates on a shallow copy of `payload`; the original isn't mutated.
    """
    del view  # currently unused after status was removed from the output
    repaired = dict(payload)
    repaired["productIds"] = [product_id]
    for k in _STRIP_FROM_PAYLOAD:
        repaired.pop(k, None)
    return repaired


def write_payload(
    output_root: Path,
    repaired_payload: dict[str, Any],
    variant_id: str,
    fetched_via_site: str,
    fallback_template_id: str | None,
    *,
    site_ids_source: dict[str, Any] | None = None,
) -> tuple[Path, list[str]]:
    """
    Write `<output_root>/<productTemplateId>/<siteIds-joined>/<variantId>.json`
    and return (path, list_of_siteIds).

    Folder naming
    -------------
    The folder is named by the COMMA-JOINED list of `siteIds`, IN ORIGINAL
    PAYLOAD ORDER (not sorted). A variant whose response says
    `siteIds: ["siteB", "siteA"]` lands under one folder named "siteB,siteA".
    The order matters because downstream push uses the first site as the
    X-UpStart-Site header and the rest as the ?sites= query — keeping
    order stable means the header/query split is deterministic across runs.

    Why pass site_ids_source separately
    ------------------------------------
    repair_payload strips `siteIds` from the payload (it's server-managed
    and shouldn't be in a PUT body). But we still need to KNOW the siteIds
    to decide where the file goes. So we pull them from the original
    response (`site_ids_source`) and write the cleaned `repaired_payload`
    to disk. If `site_ids_source` isn't given, fall back to checking the
    repaired payload itself (so the function still works in unit tests
    where strip hasn't happened).

    Why not the fetch site
    ----------------------
    The X-UpStart-Site header used during fetch tells us where the variant
    was REACHABLE from — but the variant's actual `siteIds` field is the
    source of truth for which sites it BELONGS to. A variant can be
    reachable via the primary site but live on multiple sites; pre-binding
    the file to the primary alone would lose that info.

    `fetched_via_site` is recorded only for diagnostic purposes (logged by
    the caller); it's not used to pick the folder.

    productTemplateId is read from the API response if present, else falls
    back to the sheet's product_template_id, else 'unknown'.
    """
    tpl_id = (
        repaired_payload.get("productTemplateId")
        or repaired_payload.get("product_template_id")
        or fallback_template_id
        or "unknown"
    )

    # Pull siteIds from the original (un-stripped) payload, preserving its
    # order. The first site in this list becomes the push step's
    # X-UpStart-Site header; the rest go into the ?sites= query.
    source = site_ids_source if site_ids_source is not None else repaired_payload
    raw = source.get("siteIds") or []
    if not isinstance(raw, list):
        raw = []
    site_ids = [str(s) for s in raw if s]
    if not site_ids:
        # Defensive: API should always return siteIds. If somehow missing,
        # fall back to the site we successfully fetched from rather than
        # losing the file. Use a marker prefix so this is easy to grep for
        # in the output tree.
        site_ids = [f"fallback-{fetched_via_site}"] if fetched_via_site else ["unknown-site"]

    folder_name = ",".join(site_ids)
    folder = output_root / str(tpl_id) / folder_name
    folder.mkdir(parents=True, exist_ok=True)
    out = folder / f"{variant_id}.json"
    out.write_text(json.dumps(repaired_payload, indent=2, ensure_ascii=False))
    return out, site_ids


# ---------------------------------------------------------------------------
# Step 6: duplicate enrichment (read-only; no payload writes)
# ---------------------------------------------------------------------------

def enrich_duplicate(
    session: requests.Session,
    api_host: str,
    variant_id: str,
    site_ids: list[str],
) -> dict[str, Any]:
    """
    Fetch a single duplicate variant and return its API-derived state/status.

    Unlike `process_variant`, this does NOT write any payload to disk —
    duplicates need a human decision (which product owns the variant?)
    before they can be safely rebound. We only enrich duplicate_variants.json
    with API truth so the reviewer has the variant's actual state at a
    glance when deciding.

    Returns:
        {"variantId": ..., "state": ..., "status": ..., "siteId": ...}
    All three values are None if the API returned nothing on any site.
    """
    payload, view, winning_site = fetch_variant(
        session, api_host, variant_id, site_ids,
    )
    if payload is None:
        return {
            "variantId": variant_id,
            "state": None, "status": None, "siteId": None,
        }
    return {
        "variantId": variant_id,
        "state": payload.get("state"),
        "status": _VIEW_TO_STATUS.get(view),
        "siteId": winning_site,
    }


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def process_variant(
    session: requests.Session,
    api_host: str,
    variant_id: str,
    meta: dict[str, Any],
    output_root: Path,
    site_ids: list[str],
) -> dict[str, Any]:
    """
    Fetch (across sites + views), repair, and write one variant.

    `site_ids` is the ordered list of sites to try: primary first, then
    fallbacks. Stops at the first hit. The site that succeeded is recorded
    on the success log line and in the returned dict, so downstream code
    (map update, audit) knows where the data came from.

    Returns a result dict:
      {
        "variantId": ...,
        "verdict":   "ok" | "not_found" | "error",
        "info":      <path on ok, error string otherwise>,
        "status":    "DRAFT" | "PUBLISHED" | None,   # for map update
        "state":     <api response state> | None,    # for map update
        "siteId":    <site that succeeded> | None,
      }

    Logging contract:
      - success.log: one INFO line per "ok" outcome with the winning site.
      - errors.log:  one WARNING per variant on full not-found, listing
                     ALL sites tried (no per-site noise).
                     one ERROR for write/repair failure.
                     one ERROR (from fetch_variant) for auth_html, since
                     that's a config problem the operator needs to see.
    """
    ok_log = logging.getLogger("repair.ok")
    err_log = logging.getLogger("repair.err")

    t_start = time.monotonic()
    product_id = meta["productId"]

    payload, view, winning_site = fetch_variant(
        session, api_host, variant_id, site_ids,
    )
    if payload is None:
        # Single consolidated warning: variant wasn't found on ANY site.
        # Per-attempt detail lives at DEBUG in run.log (--verbose).
        err_log.warning(
            "verdict=not_found variantId=%s productId=%s "
            "triedSiteIds=%s reason=no_view_returned_data_on_any_site",
            variant_id, product_id, list(site_ids),
        )
        return {
            "variantId": variant_id,
            "verdict": "not_found",
            "info": "no view returned data on any site",
            "status": None,
            "state": None,
            "siteId": None,
        }

    try:
        repaired = repair_payload(payload, product_id, view)
        path, file_site_ids = write_payload(
            output_root,
            repaired,
            variant_id,
            fetched_via_site=winning_site,
            fallback_template_id=meta.get("productTemplateId"),
            site_ids_source=payload,  # un-stripped, has siteIds
        )
    except OSError as e:
        err_log.error(
            "verdict=error variantId=%s productId=%s fetchedVia=%s "
            "kind=write_failed exc=%s detail=%r",
            variant_id, product_id, winning_site, type(e).__name__, str(e),
        )
        return {
            "variantId": variant_id,
            "verdict": "error",
            "info": f"write failed: {e}",
            "status": None,
            "state": None,
            "siteIds": None,
            "fetchedViaSite": winning_site,
        }
    except Exception as e:
        err_log.exception(
            "verdict=error variantId=%s productId=%s fetchedVia=%s "
            "kind=unexpected exc=%s",
            variant_id, product_id, winning_site, type(e).__name__,
        )
        return {
            "variantId": variant_id,
            "verdict": "error",
            "info": f"unexpected: {e}",
            "status": None,
            "state": None,
            "siteIds": None,
            "fetchedViaSite": winning_site,
        }

    elapsed_ms = int((time.monotonic() - t_start) * 1000)
    tpl_id = (
        repaired.get("productTemplateId")
        or repaired.get("product_template_id")
        or meta.get("productTemplateId")
        or "unknown"
    )
    final_status = _VIEW_TO_STATUS.get(view)
    final_state = payload.get("state")
    ok_log.info(
        "verdict=ok variantId=%s productId=%s productTemplateId=%s "
        "siteIds=%s fetchedVia=%s view=%s state=%s status=%s "
        "elapsed_ms=%d path=%s",
        variant_id, product_id, tpl_id, file_site_ids, winning_site, view,
        final_state, final_status,
        elapsed_ms, path,
    )
    return {
        "variantId": variant_id,
        "verdict": "ok",
        "info": str(path),
        "status": final_status,
        "state": final_state,
        "siteIds": file_site_ids,
        "fetchedViaSite": winning_site,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", required=True, type=Path, help="Path to report-linkages.xlsx")
    ap.add_argument("--output", required=True, type=Path, help="Output folder root")
    ap.add_argument(
        "--api-host",
        default="https://nochannel-test-1-api.nochannel-test.upstart.team",
        help="API host (no trailing slash)",
    )
    ap.add_argument(
        "--site-id",
        default=PRIMARY_SITE_ID,
        help="Primary X-UpStart-Site to try first (default: PRIMARY_SITE_ID constant)",
    )
    ap.add_argument(
        "--fallback-site-ids",
        default=",".join(FALLBACK_SITE_IDS),
        help="Comma-separated list of sites to try if --site-id returns nothing. "
             "Default: FALLBACK_SITE_IDS constant. Pass empty string to disable fallback.",
    )
    ap.add_argument(
        "--session-id",
        default=os.environ.get("API_SESSION_ID", "BmVa7sjFeCNL_NlOBxiADQ.NJPHpP9NccNdJqSzZmi5lk7F_EYEfBaLZO2uR-u4i1wq0o0QO2fZF79ABWHUV991GQj2zICYqFcNDn83IlksPcBP4DFD8DDPQMJT4NdxUNTsuaUhTNtA_d4DDFyobyYX"),
        help="X-Upstart-Session-Id (or set API_SESSION_ID env var)",
    )
    ap.add_argument(
        "--cookie",
        default=os.environ.get("API_COOKIE", ""),
        help="Raw Cookie header value for auth. Most Upstart endpoints need "
             "session cookies, not just X-Upstart-Session-Id. To get yours: "
             "Chrome DevTools → Network → any working API request → right-click "
             "→ Copy as cURL → grab the -b/--cookie value. Or set API_COOKIE env var.",
    )
    ap.add_argument(
        "--map-only",
        action="store_true",
        help="Just build variant_map.json and duplicate_variants.json, then "
             "exit (no API calls of any kind).",
    )
    ap.add_argument(
        "--skip-duplicate-enrichment",
        action="store_true",
        help="Skip the optional pass that fetches state/status for duplicate "
             "variants. Speeds up the run; duplicate_variants.json will have "
             "null variant-level state/status. Per-product state/status is "
             "always populated from the sheet regardless of this flag.",
    )
    ap.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Concurrent fetch workers (default 1)",
    )
    ap.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Process only the first N variants (0 = all). Useful for dry-run.",
    )
    ap.add_argument(
        "--log-dir",
        type=Path,
        default=None,
        help="Where to put run.log / success.log / errors.log. "
             "Default: <output>/logs",
    )
    ap.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable DEBUG-level logging (per-view fetch attempts, etc).",
    )
    args = ap.parse_args()

    if not args.input.exists():
        # Logging isn't up yet — go to stderr directly so this doesn't get lost.
        print(f"Input file not found: {args.input}", file=sys.stderr)
        return 1

    args.output.mkdir(parents=True, exist_ok=True)
    log_dir = args.log_dir or (args.output / "logs")
    log, _ok_log, _err_log = setup_logging(log_dir, verbose=args.verbose)
    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log.info("run_start run_id=%s input=%s output=%s", run_id, args.input, args.output)

    log.info("[1/6] reading sheet from %s", args.input)
    occurrences = collect_all_occurrences(args.input)
    log.info("parsed %d (variant, product, issue) occurrences", len(occurrences))

    log.info("[2/6] detecting duplicate variants (claimed by multiple products)")
    duplicates = detect_duplicates(occurrences)
    dup_path = args.output / "duplicate_variants.json"
    dup_path.write_text(json.dumps(duplicates, indent=2, ensure_ascii=False))
    log.info("found %d duplicate variants — wrote %s", len(duplicates), dup_path)

    duplicate_ids = set(duplicates.keys())

    # ----------------------------------------------------------------------
    # Session + site list setup
    # ----------------------------------------------------------------------
    # Needed before either enrichment (step 3) or the main fetch (step 5).
    # Build it once, reuse across both passes. Skipping under --map-only
    # avoids the auth-cookie warning when the user just wants the JSON.
    if not args.map_only:
        raw_fallbacks = [s.strip() for s in (args.fallback_site_ids or "").split(",")]
        site_ids: list[str] = []
        for s in [args.site_id, *raw_fallbacks]:
            if s and s not in site_ids:
                site_ids.append(s)
        if not site_ids:
            log.error("no site IDs configured — set --site-id at minimum")
            return 1
        log.info("siteIds (in retry order): %s", site_ids)
        if not args.cookie:
            log.warning(
                "no Cookie header provided — if the API gateway requires session "
                "cookies you'll get HTML responses (kind=auth_html in errors.log). "
                "Pass --cookie or set API_COOKIE env var. See --help for how."
            )
        session = make_session(args.session_id, cookie_header=args.cookie or None)
    else:
        site_ids = []
        session = None  # type: ignore[assignment]

    # ----------------------------------------------------------------------
    # Step 3 — Enrich duplicates (BEFORE clean-map build and main fetch).
    # ----------------------------------------------------------------------
    # Run this first so the duplicate_variants.json is fully populated by
    # the time the long main fetch (step 5) starts — a human reviewer can
    # triage duplicates in parallel while the 15k clean-map fetches run.
    # No payload files written: duplicates need a human decision.
    if duplicates and not args.map_only and not args.skip_duplicate_enrichment:
        log.info(
            "[3/6] enriching %d duplicate variants with API state/status "
            "(no payload writes)",
            len(duplicates),
        )
        dup_lock = threading.Lock()
        dup_ok = dup_miss = 0
        with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
            futures = {
                pool.submit(enrich_duplicate, session, args.api_host, vid, site_ids): vid
                for vid in duplicates
            }
            for fut in as_completed(futures):
                result = fut.result()
                vid = result["variantId"]
                with dup_lock:
                    duplicates[vid]["state"] = result["state"]
                    duplicates[vid]["status"] = result["status"]
                    duplicates[vid]["siteId"] = result["siteId"]
                if result["state"] is not None:
                    dup_ok += 1
                else:
                    dup_miss += 1
                    # Same one-line consolidated warning as the main pass.
                    logging.getLogger("repair.err").warning(
                        "duplicate_enrich_not_found variantId=%s "
                        "triedSiteIds=%s reason=no_view_returned_data_on_any_site",
                        vid, site_ids,
                    )
        dup_path.write_text(json.dumps(duplicates, indent=2, ensure_ascii=False))
        log.info(
            "rewrote %s with API state/status (ok=%d miss=%d)",
            dup_path, dup_ok, dup_miss,
        )
    elif args.skip_duplicate_enrichment:
        log.info("[3/6] skipping duplicate enrichment (--skip-duplicate-enrichment)")

    # ----------------------------------------------------------------------
    # Step 4 — Snapshot product states BEFORE any variant PUT.
    # ----------------------------------------------------------------------
    # Sheet-only derivation. Includes every productId that appears in
    # occurrences (i.e. both clean-map and duplicate-variant products).
    # Persists to product_states_pre_repair.json so a follow-up step can
    # decide which products need a re-publish or state change after the
    # variant PUTs land.
    log.info("[4/6] snapshotting product states (sheet-derived, pre-repair)")
    product_states = build_product_state_snapshot(occurrences)
    pstate_path = args.output / "product_states_pre_repair.json"
    pstate_path.write_text(json.dumps(product_states, indent=2, ensure_ascii=False))
    # Log the breakdown so it's obvious at a glance what we're dealing with
    state_counts = Counter(v["state"] for v in product_states.values())
    log.info(
        "snapshotted %d unique products (DRAFT=%d PUBLISHED=%d null=%d) — wrote %s",
        len(product_states),
        state_counts.get("DRAFT", 0),
        state_counts.get("PUBLISHED", 0),
        state_counts.get(None, 0),
        pstate_path,
    )

    log.info("[5/6] building clean variant map (excluding duplicates)")
    variant_map = build_variant_map(occurrences, duplicate_ids)
    map_path = args.output / "variant_map.json"
    map_path.write_text(json.dumps(variant_map, indent=2, ensure_ascii=False))
    log.info("%d variants in clean map — wrote %s", len(variant_map), map_path)

    if args.map_only:
        log.info("--map-only set — skipping API calls. run_id=%s", run_id)
        return 0

    items = list(variant_map.items())
    if args.limit > 0:
        items = items[: args.limit]
        log.info("--limit %d applied — processing first %d variants", args.limit, len(items))

    log.info(
        "[6/6] fetching %d variants from %s",
        len(items), args.api_host,
    )

    ok = bad = nf = 0
    t0 = time.time()
    # Workers concurrently write status/state into variant_map under this
    # lock. CPython's GIL would probably make distinct-key dict assignment
    # safe in practice, but being explicit keeps the script robust against
    # future Python versions (e.g. free-threaded 3.13+) and against the
    # final json.dumps reading the dict mid-update.
    map_lock = threading.Lock()

    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        futures = {
            pool.submit(process_variant, session, args.api_host, vid, meta, args.output, site_ids): vid
            for vid, meta in items
        }
        for fut in as_completed(futures):
            result = fut.result()
            vid = result["variantId"]
            verdict = result["verdict"]
            # The per-variant log lines are already emitted inside
            # process_variant() (success.log on ok, errors.log on miss/err).
            # Here we update counters and merge API truth into the map.
            if verdict == "ok":
                ok += 1
                with map_lock:
                    # Update placeholders with what the API actually returned.
                    # siteIds (plural) comes from the payload itself (truth
                    # for multi-site variants); fetchedViaSite records which
                    # site served the GET (one of siteIds, for audit).
                    if result["status"] is not None:
                        variant_map[vid]["status"] = result["status"]
                    if result["state"] is not None:
                        variant_map[vid]["state"] = result["state"]
                    if result.get("siteIds") is not None:
                        variant_map[vid]["siteIds"] = result["siteIds"]
                    if result.get("fetchedViaSite") is not None:
                        variant_map[vid]["fetchedViaSite"] = result["fetchedViaSite"]
            elif verdict == "not_found":
                nf += 1
            else:
                bad += 1

            done = ok + bad + nf
            if done % 50 == 0:
                log.info(
                    "progress %d/%d (ok=%d miss=%d err=%d)",
                    done, len(items), ok, nf, bad,
                )

    # Persist the updated map. Variants the API didn't return keep their
    # null placeholders, which is the documented signal that the fetch
    # missed them (cross-reference errors.log for the reason).
    with map_lock:
        map_path.write_text(json.dumps(variant_map, indent=2, ensure_ascii=False))
    log.info("rewrote %s with API-sourced status/state for %d variants", map_path, ok)

    elapsed = time.time() - t0
    log.info(
        "run_end run_id=%s ok=%d not_found=%d errors=%d elapsed_s=%.1f out=%s",
        run_id, ok, nf, bad, elapsed, args.output,
    )
    return 0 if bad == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())