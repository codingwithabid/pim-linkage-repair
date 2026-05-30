#!/usr/bin/env python3
"""
export_duplicates_to_excel.py

Convert duplicate JSON files produced by repair_variants.py and
repair_products.py into a single, human-readable Excel workbook.

Duplicates are the cases that the repair toolchain refuses to auto-fix:

  - On the variant side (`duplicate_variants.json`): the spreadsheet
    claims the SAME variant belongs to two or more DIFFERENT products.
    A variant can only have one parent, so the toolchain can't pick a
    side — a human has to decide.

  - On the product side (`duplicate_variants_across_product.json`):
    the same variant appears in two or more DIFFERENT products'
    `variantIds` lists. Same human-decision problem from the other
    direction.

This script reads either or both files and produces ONE sheet that has
one row per (variant, product) pairing. So a variant claimed by 3
products becomes 3 rows, sharing the same variant_id but with
different product columns. That makes it easy to filter, sort by
variant, and assign owners in a spreadsheet rather than a JSON viewer.

Usage:
    # Read both files from one folder (the toolchain's output dir):
    python export_duplicates_to_excel.py --input ./out

    # Or point at the files directly:
    python export_duplicates_to_excel.py \\
        --variant-duplicates ./out/duplicate_variants.json \\
        --product-duplicates ./out/duplicate_variants_across_product.json \\
        --output ./duplicates_review.xlsx

    # Variant-side only:
    python export_duplicates_to_excel.py \\
        --variant-duplicates ./out/duplicate_variants.json

The output workbook has one sheet, "Duplicates", with these columns:

  source                  "variant-side" or "product-side" (which file
                          it came from)
  variant_id              the duplicated variant
  variant_name            (from API enrichment, may be empty if not run)
  variant_state           NORMAL / ARCHIVED / null
  variant_status          DRAFT / PUBLISHED / null
  variant_fetched_via_site site that served the variant data
  product_id              one of the claiming products
  product_name            from the sheet, or API enrichment
  product_template_id     ditto
  product_state           the claiming product's state (per file)
  product_status          the claiming product's status (per file)
  issues                  comma-joined issue codes for THIS (variant,
                          product) combination
  total_claimers          number of products claiming this variant
                          (constant across all rows for the variant —
                          a quick at-a-glance count)

Empty/null fields are written as blank cells rather than the JSON
string "null", so filters work cleanly in Excel.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Sheet definition — keep column order/headers in one place so it's easy to
# add/reorder columns later without hunting through the writer code.
# ---------------------------------------------------------------------------
COLUMNS: list[tuple[str, int]] = [
    # (header, column width)
    ("source",                    14),
    ("variant_id",                42),
    ("variant_name",              30),
    ("variant_state",             14),
    ("variant_status",            14),
    ("variant_fetched_via_site",  42),
    ("product_id",                42),
    ("product_name",              30),
    ("product_template_id",       42),
    ("product_state",             14),
    ("product_status",            14),
    ("issues",                    40),
    ("total_claimers",            16),
]


def _load_json(path: Path) -> dict[str, Any]:
    """Read a JSON file, defaulting to {} for missing/empty.

    A missing file isn't fatal — the toolchain may legitimately produce
    only one of the two duplicate files (e.g. you only ran the variant
    side). Empty `{}` files are also fine and mean "no duplicates found",
    which is the common case.
    """
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        return {}
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        sys.exit(f"ERROR: {path} is not valid JSON: {e}")
    if not isinstance(data, dict):
        sys.exit(f"ERROR: {path} should be a JSON object, got {type(data).__name__}")
    return data


def _cell(v: Any) -> Any:
    """Normalise a value for Excel output.

    None → empty string (so filters don't trip on "None").
    Lists → comma-joined string.
    Anything else → str()."""
    if v is None:
        return ""
    if isinstance(v, list):
        # Sort issue codes for stable output across runs. Other lists
        # (none expected at the moment) get joined as-is.
        try:
            return ", ".join(sorted(str(x) for x in v))
        except TypeError:
            return ", ".join(str(x) for x in v)
    return v if isinstance(v, (int, float)) else str(v)


def _expand_duplicates(
    duplicates: dict[str, Any],
    source_label: str,
) -> list[dict[str, Any]]:
    """Turn the nested duplicate-JSON into a flat list of rows.

    One row per (variant_id, product_id) pair. Variant-level fields
    (variant_name, state, status, siteId) repeat across all rows for
    the same variant — that's intentional, since a flat sheet is much
    easier to filter and sort than the nested original.

    Empty input gives an empty list — the caller decides what to do
    (still write the workbook with just headers, or skip).
    """
    rows: list[dict[str, Any]] = []
    for vid, vinfo in duplicates.items():
        # Variant-level fields, populated post-fetch by the enrichment
        # pass. May still be None if the user passed
        # --skip-duplicate-enrichment or the API call failed.
        variant_name        = vinfo.get("variantName", "")
        variant_state       = vinfo.get("state")
        variant_status      = vinfo.get("status")
        variant_via_site    = vinfo.get("siteId")
        products            = vinfo.get("products", []) or []
        total_claimers      = len(products)

        if not products:
            # Shouldn't normally happen — if a variant ended up in this
            # file it had ≥2 claiming products. But defensive: produce
            # one row showing the variant has no recorded claimers so
            # the operator notices the data inconsistency.
            rows.append({
                "source":                   source_label,
                "variant_id":               vid,
                "variant_name":             variant_name,
                "variant_state":            variant_state,
                "variant_status":           variant_status,
                "variant_fetched_via_site": variant_via_site,
                "product_id":               "",
                "product_name":             "",
                "product_template_id":      "",
                "product_state":            "",
                "product_status":           "",
                "issues":                   "",
                "total_claimers":           0,
            })
            continue

        for p in products:
            rows.append({
                "source":                   source_label,
                "variant_id":               vid,
                "variant_name":             variant_name,
                "variant_state":            variant_state,
                "variant_status":           variant_status,
                "variant_fetched_via_site": variant_via_site,
                "product_id":               p.get("productId", ""),
                "product_name":             p.get("productName", ""),
                "product_template_id":      p.get("productTemplateId", ""),
                "product_state":            p.get("state"),
                "product_status":           p.get("status"),
                "issues":                   p.get("issues", []) or [],
                "total_claimers":           total_claimers,
            })
    return rows


def write_workbook(rows: list[dict[str, Any]], output_path: Path) -> None:
    """Write rows to a single-sheet xlsx with header row + filter."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicates"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="305496")
    body_font    = Font(name="Arial", size=10)
    body_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center       = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = [name for name, _w in COLUMNS]
    ws.append(headers)
    for col_idx, (_name, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Body rows. Sort to group all rows of the same variant together,
    # then by source so variant-side and product-side don't interleave
    # within a variant. This makes the sheet readable straight from
    # File → Open without any operator filtering.
    rows_sorted = sorted(
        rows,
        key=lambda r: (
            str(r.get("variant_id") or ""),
            r.get("source", ""),
            str(r.get("product_id") or ""),
        ),
    )

    for r in rows_sorted:
        ws.append([_cell(r.get(name)) for name, _w in COLUMNS])
        row_idx = ws.max_row
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = body_align

    # Freeze the header and add an autofilter so users can immediately
    # filter by source/status/site.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, ws.max_row)}"

    wb.save(output_path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument(
        "--input", type=Path, default=None,
        help="Folder containing duplicate_variants.json and/or "
             "duplicate_variants_across_product.json. Either or both "
             "are picked up if they exist. Mutually exclusive with the "
             "explicit --variant-duplicates / --product-duplicates flags.",
    )
    ap.add_argument(
        "--variant-duplicates", type=Path, default=None,
        help="Explicit path to duplicate_variants.json (variant-side).",
    )
    ap.add_argument(
        "--product-duplicates", type=Path, default=None,
        help="Explicit path to duplicate_variants_across_product.json "
             "(product-side).",
    )
    ap.add_argument(
        "--output", "-o", type=Path, default=Path("duplicates_review.xlsx"),
        help="Output xlsx path (default: ./duplicates_review.xlsx).",
    )
    args = ap.parse_args()

    if args.input and (args.variant_duplicates or args.product_duplicates):
        ap.error(
            "Use EITHER --input OR --variant-duplicates/--product-duplicates, "
            "not both."
        )

    # Resolve paths. --input is a convenience; the explicit flags win
    # when --input isn't given.
    if args.input:
        variant_path = args.input / "duplicate_variants.json"
        product_path = args.input / "duplicate_variants_across_product.json"
    else:
        variant_path = args.variant_duplicates
        product_path = args.product_duplicates

    if not variant_path and not product_path:
        ap.error("Provide --input or at least one of --variant-duplicates / "
                 "--product-duplicates.")

    # Load whichever files were given. Missing files are silently skipped
    # (see _load_json) so the script can be wired into pipelines that
    # may or may not produce both sides.
    variant_data: dict[str, Any] = {}
    product_data: dict[str, Any] = {}
    if variant_path is not None:
        variant_data = _load_json(variant_path)
    if product_path is not None:
        product_data = _load_json(product_path)

    rows: list[dict[str, Any]] = []
    rows += _expand_duplicates(variant_data, "variant-side")
    rows += _expand_duplicates(product_data, "product-side")

    # Console summary so the operator knows what got produced. Even
    # zero-row outputs are written (with just the header) so downstream
    # automation can always assume a workbook exists.
    n_variants_v = len(variant_data)
    n_variants_p = len(product_data)
    print(
        f"variant-side duplicates:  {n_variants_v} variants ({sum(len(d.get('products', [])) for d in variant_data.values())} pairings)"
    )
    print(
        f"product-side duplicates:  {n_variants_p} variants ({sum(len(d.get('products', [])) for d in product_data.values())} pairings)"
    )
    print(f"total rows to write:      {len(rows)}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(rows, args.output)
    print(f"wrote {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())