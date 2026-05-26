#!/usr/bin/env python3
"""
repair_products.py
==================

Product-side counterpart of repair_variants.py.

This script repairs broken product→variant linkages in the Denver Mattress
PIM. The variant script repaired the variant side (product list inside a
variant payload); this one repairs the product side (variant list inside
a product payload).

The two sides are mirror images: variants point UP at their parent
products via `productIds`; products point DOWN at their child variants
via `variantIds`. When a variant says "I belong to product X" but product
X's `variantIds` array doesn't include that variant — that's the
PARENT_DOES_NOT_LINK_BACK issue this script fixes.

Pipeline (6 steps, mirroring repair_variants.py)
------------------------------------------------
  1. Read sheets — keep PARENT_DOES_NOT_LINK_BACK entries only.
  2. Build product→variants map from sheets, merging across all 3 sheets.
     Detect duplicates: a variant claimed by 2+ different products.
     Write duplicate_variants_across_product.json (placeholders).
  3. Enrich duplicates via the variant API (state, status, siteIds).
     No payload repairs — duplicates need human review.
  4. Snapshot variant states (state, status) by GETting each unique
     variant in the clean map. Write variant_states_pre_repair.json.
     (Mirrors step 4 of repair_variants.py, but for variants this time.)
  5. Build clean product_map.json (products with no duplicate variants),
     with placeholders to be filled in step 6.
  6. Fetch each product via the API. Append sheet variantIds to the
     payload's variantIds (dedupe). Write repaired payloads as
     <output>/<productTemplateId>/<siteIds>/<productId>.json. Rewrite
     product_map.json with authoritative API values.

CLI matches repair_variants.py: --input, --output, --api-host, --site-id,
--fallback-site-ids, --session-id, --cookie, --workers, --limit, --dry-run,
--log-dir, --verbose. Plus --map-only and --skip-duplicate-enrichment for
staged runs.

Usage
-----
    export API_COOKIE='...'
    python3 repair_products.py \\
        --input report-linkages.xlsx \\
        --output ./out-products \\
        --workers 4
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import threading
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any

import requests

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "openpyxl not installed. pip install openpyxl",
        file=sys.stderr,
    )
    raise


# ---------------------------------------------------------------------------
# Constants — Denver Mattress tenant, sites, sheet contract.
# ---------------------------------------------------------------------------
PRIMARY_SITE_ID = "201cb789-4198-488b-a5eb-4e7df0fb4bee"
FALLBACK_SITE_IDS = [
    "8d3ea3bc-f65b-4227-9fa6-6fae40e4575a",
    "fbfcd92b-d271-4002-9163-4f84986b41be",
]

# All 3 sheets share product_id and issue_label columns. The variant-list
# column has a different name in each sheet (a quirk of how the sheets
# were authored by different teams).
SHEETS: list[dict[str, str]] = [
    {
        "name":          "Variant Specific(+ve)",
        "variants_col":  "variants having issues",
    },
    {
        "name":          "Product Specific(+ve)",
        "variants_col":  "variants links to be added",
    },
    {
        "name":          "Product to verify(-ve)",
        "variants_col":  "links to be fixed",
    },
]

PRODUCT_ID_COL_CANDIDATES   = ("product_id", "productid", "product id")
ISSUE_LABEL_COL_CANDIDATES  = ("issue_label", "issuelabel", "issue label")
PRODUCT_NAME_COL_CANDIDATES = ("product_name", "productname", "product name")
TPL_ID_COL_CANDIDATES       = (
    "product_template_id", "producttemplateid", "product template id",
)

# We only ACT on PARENT_DOES_NOT_LINK_BACK rows. Other codes (STALE_LATEST_EMPTY,
# ORPHAN_BUT_REFERENCED, etc.) are handled by other scripts in the pipeline;
# rows with those codes are silently skipped here.
KEEP_ISSUE_CODES = {"PARENT_DOES_NOT_LINK_BACK"}

# Any code we recognize — used by the cell parser to detect a "CODE:" prefix.
# Recognizing OTHER codes here lets us cleanly strip the prefix on those
# rows too, even though we're going to filter them out by KEEP_ISSUE_CODES.
KNOWN_ISSUE_CODES = KEEP_ISSUE_CODES | {
    "STALE_LATEST_EMPTY",
    "ORPHAN_BUT_REFERENCED",
    "PARENT_CHANGED",
    "PRODUCT_DROPPED_VARIANT",
}

# Within the "variants having issues" cell, IDs can look like:
#   "abcd1234-5678-... (Some Name), efgh-9012 (Other Name)"
#   "BA-XYZ123, BA-ABC456"
#   "BA-XYZ123 (Some Name)"
# We capture the id token; the optional "(Name)" label is metadata only.
_VARIANT_ID_RE = re.compile(
    r"([A-Za-z0-9][A-Za-z0-9_\-]{2,})\s*(?:\([^)]*\))?"
)


# ---------------------------------------------------------------------------
# Payload strip list — fields the API returns but won't accept on PUT.
# Same logic as repair_variants.py._STRIP_FROM_PAYLOAD.
# ---------------------------------------------------------------------------
_STRIP_FROM_PAYLOAD = frozenset({
    # --- Identity & links ---
    "links",
    "id",  # server-owned, unusual but per direction in the variant script

    # --- Lifecycle flags the platform owns ---
    "hasDraft",
    "hasReady",
    "hasPublished",

    # --- Site routing (encoded in folder path) ---
    "siteIds",

    # --- Platform-derived metadata ---
    "syncedProductType",
    "productTemplateVersion",
    "productType",
    "sellable",
    "status",
    "state",

    # --- Timestamps (server-managed) ---
    "dateCreated",       "createdDate",
    "dateModified",      "modifiedDate",  "lastModifiedDate",
    "datePublished",     "publishDate",
    "dateArchived",      "archiveDate",
})


# ---------------------------------------------------------------------------
# Logging — three rotating handlers, mirroring repair_variants.py.
# ---------------------------------------------------------------------------
LOG_FMT = "%(asctime)s | %(levelname)-7s | %(name)-15s | %(message)s"
LOG_DATEFMT = "%Y-%m-%d %H:%M:%S"


class ProgressReporter:
    """
    Periodic progress logger for long-running parallel loops.

    Logs at every `step` items AND at every `interval_s` seconds of
    wall time, whichever comes first — so a slow loop still shows
    movement and a fast loop doesn't spam.

    Step is auto-computed from `total` if not given: ~10% of total but
    no less than 10 and no more than 200. Wall-clock interval defaults
    to 15s.

    Thread-safe; intended use is to call `tick()` after each completed
    item (typically inside the `as_completed` loop). The first and last
    items are always logged.
    """

    def __init__(
        self,
        log: logging.Logger,
        label: str,
        total: int,
        step: int | None = None,
        interval_s: float = 15.0,
    ):
        self.log = log
        self.label = label
        self.total = max(1, total)
        if step is None:
            step = max(10, min(200, self.total // 10 or 1))
        self.step = step
        self.interval_s = interval_s
        self._done = 0
        self._last_log_done = 0
        self._last_log_t = time.monotonic()
        self._lock = threading.Lock()

    def tick(self, extra: str = "") -> None:
        with self._lock:
            self._done += 1
            now = time.monotonic()
            elapsed_since_last = now - self._last_log_t
            steps_since_last = self._done - self._last_log_done
            is_last = self._done >= self.total
            # Log on step boundary OR after `interval_s` OR on the last item.
            if (
                steps_since_last >= self.step
                or elapsed_since_last >= self.interval_s
                or is_last
            ):
                rate = self._done / max(0.001, now - (self._last_log_t - elapsed_since_last))
                # Rate over the WHOLE run, not the last window — easier to read.
                pct = 100.0 * self._done / self.total
                suffix = f" {extra}" if extra else ""
                self.log.info(
                    "%s progress %d/%d (%.0f%% — ~%.1f/s)%s",
                    self.label, self._done, self.total, pct, rate, suffix,
                )
                self._last_log_done = self._done
                self._last_log_t = now


def setup_logging(log_dir: Path, verbose: bool = False) -> logging.Logger:
    """
    Three rotating file handlers + a console mirror at INFO+:

      products-repair.run.log      — everything (DEBUG if --verbose, else INFO)
      products-repair.success.log  — INFO-level per-product verdicts
      products-repair.errors.log   — WARNING+ (propagates to run.log too)

    Returns the main logger; submodule loggers (`products.ok`, `products.err`)
    are configured here too.
    """
    log_dir.mkdir(parents=True, exist_ok=True)
    formatter = logging.Formatter(LOG_FMT, datefmt=LOG_DATEFMT)

    def _rotating(path: Path, level: int) -> RotatingFileHandler:
        h = RotatingFileHandler(
            path, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8"
        )
        h.setLevel(level)
        h.setFormatter(formatter)
        return h

    main = logging.getLogger("products")
    main.setLevel(logging.DEBUG if verbose else logging.INFO)
    main.propagate = False
    main.handlers.clear()
    main.addHandler(
        _rotating(
            log_dir / "products-repair.run.log",
            logging.DEBUG if verbose else logging.INFO,
        )
    )
    console = logging.StreamHandler(sys.stderr)
    console.setLevel(logging.INFO)
    console.setFormatter(formatter)
    main.addHandler(console)

    ok = logging.getLogger("products.ok")
    ok.setLevel(logging.INFO)
    ok.propagate = False
    ok.handlers.clear()
    ok.addHandler(_rotating(log_dir / "products-repair.success.log", logging.INFO))

    err = logging.getLogger("products.err")
    err.setLevel(logging.WARNING)
    err.propagate = True  # bubble warnings up to run.log
    err.handlers.clear()
    err.addHandler(_rotating(log_dir / "products-repair.errors.log", logging.WARNING))

    main.info("repair_products logging initialised: dir=%s verbose=%s",
              log_dir, verbose)
    return main


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------

def _parse_issue_cell(
    cell_value: Any,
) -> tuple[dict[str, list[str]], list[str]]:
    """
    Parse a variant-list cell. The sheets use a few different cell shapes,
    and this function handles all of them.

    Format A — bare variant list (when issue_label is in its own column
    AND the row has only ONE issue code):

        "BA-XYZ123, BA-ABC456"
        "BA-XYZ123 (Foo), BA-ABC456 (Bar)"

    Format B — single-code prefix (the row has only one issue, but it's
    written as a prefix inside the cell):

        "PARENT_DOES_NOT_LINK_BACK: BA-XYZ123 (Foo), BA-ABC456 (Bar)"

    Format C — multi-code (compound) cell, ONE code per line. This is
    the shape compound rows use in the real spreadsheets:

        "PARENT_DOES_NOT_LINK_BACK: BA-FOO (cordial)
         STALE_LATEST_EMPTY: BA-BAR (cordisaal)"

    Returns (by_code, no_code_ids) where:
      - `by_code` maps each recognized CODE to the list of variant IDs
        directly under it. Empty if no CODE prefixes are found.
      - `no_code_ids` are the variant IDs found without any code prefix
        (for format A, where the caller knows the code from the
        issue_label column).

    Empty cell → ({}, []).

    Why this matters: when issue_label is COMPOUND (e.g.
    "PARENT_DOES_NOT_LINK_BACK | STALE_LATEST_EMPTY"), the cell groups
    variants under each code separately. Naively grabbing every ID
    would pull in variants meant for other scripts (e.g. variant-side
    STALE_LATEST_EMPTY repairs). The caller filters by KEEP_ISSUE_CODES
    to take only the relevant variants.
    """
    if cell_value is None:
        return {}, []
    text = str(cell_value).strip()
    if not text:
        return {}, []

    # Split into CODE-labeled SECTIONS using a regex that captures each
    # "CODE: ..." block. The lookahead stops at the start of the next
    # known code or end-of-string, so multi-line compound cells parse
    # correctly even with whitespace between sections.
    # The known codes list is a fixed alternation built from KNOWN_ISSUE_CODES.
    known_alt = "|".join(re.escape(c) for c in sorted(KNOWN_ISSUE_CODES, key=len, reverse=True))
    section_re = re.compile(
        rf"\b({known_alt})\s*:\s*([\s\S]*?)(?=(?:\b(?:{known_alt})\s*:)|\Z)",
        re.MULTILINE,
    )

    by_code: dict[str, list[str]] = {}
    matched_anything = False
    for m in section_re.finditer(text):
        matched_anything = True
        code = m.group(1)
        body = m.group(2) or ""
        ids: list[str] = []
        seen_in_section: set[str] = set()
        for id_m in _VARIANT_ID_RE.finditer(body):
            tok = id_m.group(1).strip()
            if tok and tok not in seen_in_section and tok not in KNOWN_ISSUE_CODES:
                seen_in_section.add(tok)
                ids.append(tok)
        if ids:
            # If the same code appears twice in one cell (defensive),
            # merge while preserving order and uniqueness.
            existing = by_code.setdefault(code, [])
            for vid in ids:
                if vid not in existing:
                    existing.append(vid)

    no_code_ids: list[str] = []
    if not matched_anything:
        # No CODE: prefixes anywhere → format A: bare variant list.
        seen: set[str] = set()
        for m in _VARIANT_ID_RE.finditer(text):
            tok = m.group(1).strip()
            if tok and tok not in seen and tok not in KNOWN_ISSUE_CODES:
                seen.add(tok)
                no_code_ids.append(tok)

    return by_code, no_code_ids


def read_sheets(
    xlsx_path: Path,
) -> list[dict[str, Any]]:
    """
    Read all 3 sheets, return a flat list of "product entries":

        [
          {
            "sheet":             "Variant Specific(+ve)",
            "productId":         "f1259472-...",
            "productName":       "Lounge 5 Pc. Theater Sectional",
            "productTemplateId": "71d5cab7-...",
            "variantIds":        ["BA-INTWPWK", "BA-XYZ"],
            "issues":            ["PARENT_DOES_NOT_LINK_BACK"],
          },
          ...
        ]

    Row shape (slightly different per sheet — the variant-list column
    name varies):
      - `product_id`              → productId
      - `issue_label`             → must equal one of KEEP_ISSUE_CODES
      - <per-sheet variant col>   → variant ID list (comma-separated in
                                    one cell, optional (Name) labels)
            Variant Specific(+ve):  variants having issues
            Product Specific(+ve):  variants links to be added
            Product to verify(-ve): links to be fixed
      - `product_name`            → productName (optional)
      - `product_template_id`     → productTemplateId (optional)

    Rows with a different issue_label (e.g. STALE_LATEST_EMPTY) are
    silently dropped — those are handled by other scripts in the pipeline.

    Why we return a flat list with duplicates (same productId across rows
    or sheets) instead of merging here: the merge step (#2 below) wants
    to track which sheet contributed what, and merging here would lose
    provenance.
    """
    log = logging.getLogger("products")
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)

    entries: list[dict[str, Any]] = []
    sheets_seen: list[str] = []
    counts = Counter()

    for sheet_spec in SHEETS:
        sheet_name = sheet_spec["name"]
        variants_col_name = sheet_spec["variants_col"]
        if sheet_name not in wb.sheetnames:
            log.warning("sheet missing: %r — skipping", sheet_name)
            counts["missing_sheets"] += 1
            continue
        sheets_seen.append(sheet_name)
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            log.warning("sheet %r is empty", sheet_name)
            continue

        # Normalize header names to lowercase + collapsed whitespace for
        # case- and spacing-tolerant lookup. Sheet authors use inconsistent
        # casing and sometimes double-spaces.
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            key = " ".join(str(cell).strip().lower().split())
            header_map[key] = idx

        def col(*names: str) -> int | None:
            # Don't use `a or b` — column 0 is falsy and would mis-trigger
            # the fallback. Explicit None checks instead.
            for n in names:
                v = header_map.get(" ".join(n.lower().split()))
                if v is not None:
                    return v
            return None

        product_id_col    = col(*PRODUCT_ID_COL_CANDIDATES)
        issue_label_col   = col(*ISSUE_LABEL_COL_CANDIDATES)
        # The variant-list column name is sheet-specific; only that one
        # name is acceptable per sheet (not a candidates list).
        variants_col      = col(variants_col_name)
        product_name_col  = col(*PRODUCT_NAME_COL_CANDIDATES)
        tpl_id_col        = col(*TPL_ID_COL_CANDIDATES)

        # product_id and the variant-list column are required. issue_label
        # is OPTIONAL because some sheets embed the code in the cell prefix
        # (e.g. "PARENT_DOES_NOT_LINK_BACK: BA-XYZ, ..."). _parse_issue_cell
        # extracts that prefix and we fall back to it when issue_label is
        # absent.
        missing = []
        if product_id_col  is None: missing.append("product_id")
        if variants_col    is None: missing.append(variants_col_name)
        if missing:
            log.warning(
                "sheet %r missing required column(s): %s — skipping. "
                "headers seen: %s",
                sheet_name, missing, sorted(header_map),
            )
            continue
        if issue_label_col is None:
            log.info(
                "sheet %r has no issue_label column — falling back to "
                "the CODE: prefix in the variants cell for row filtering.",
                sheet_name,
            )

        for row in rows:
            # product_id
            if product_id_col >= len(row):
                continue
            pid_cell = row[product_id_col]
            if pid_cell is None:
                continue
            product_id = str(pid_cell).strip()
            if not product_id:
                continue

            # Parse the variants cell into {code: [ids]} plus a no-code bucket
            # for sheets that use bare-list cells (format A).
            variants_cell = row[variants_col] if variants_col < len(row) else None
            by_code, no_code_ids = _parse_issue_cell(variants_cell)

            # Determine the row's labels. issue_label can be COMPOUND
            # (e.g. "PARENT_DOES_NOT_LINK_BACK | STALE_LATEST_EMPTY"),
            # so we split on '|' and treat each piece as an independent code.
            labels_from_col: set[str] = set()
            if issue_label_col is not None and issue_label_col < len(row):
                v = row[issue_label_col]
                if v is not None:
                    raw = str(v).strip()
                    if raw:
                        labels_from_col = {
                            piece.strip() for piece in raw.split("|") if piece.strip()
                        }

            # Combine: union of labels-from-column AND keys-from-cell.
            # When the cell uses format C (multi-code), `by_code` keys are
            # authoritative. When the cell is bare (format A), only the
            # issue_label column tells us what code applies.
            all_codes: set[str] = labels_from_col | set(by_code.keys())

            if not all_codes:
                counts["rows_no_issue_label"] += 1
                continue

            # Pick the codes we care about — intersection with KEEP_ISSUE_CODES.
            relevant_codes = all_codes & KEEP_ISSUE_CODES
            if not relevant_codes:
                counts["rows_filtered_out"] += 1
                continue

            # Build the variantIds list for THIS code(s) only:
            #   - If the cell has per-code sections, take ONLY the ones
            #     under our relevant codes.
            #   - Otherwise (bare cell, single label from issue_label), use
            #     the no_code_ids list.
            variant_ids: list[str] = []
            seen: set[str] = set()
            if by_code:
                for code in relevant_codes:
                    for vid in by_code.get(code, []):
                        if vid not in seen:
                            seen.add(vid)
                            variant_ids.append(vid)
            else:
                # Bare cell — all variants belong to the labels in
                # labels_from_col, and at least one of those is relevant.
                for vid in no_code_ids:
                    if vid not in seen:
                        seen.add(vid)
                        variant_ids.append(vid)

            if not variant_ids:
                # Compound row where our code lists no variants — defensive,
                # treat as if there's nothing to do here.
                counts["rows_no_variant_ids"] += 1
                continue

            # Optional metadata
            product_name = ""
            if product_name_col is not None and product_name_col < len(row):
                v = row[product_name_col]
                if v is not None:
                    product_name = str(v).strip()
            tpl_id = ""
            if tpl_id_col is not None and tpl_id_col < len(row):
                v = row[tpl_id_col]
                if v is not None:
                    tpl_id = str(v).strip()

            entries.append({
                "sheet":             sheet_name,
                "productId":         product_id,
                "productName":       product_name,
                "productTemplateId": tpl_id,
                "variantIds":        variant_ids,
                "issues":            sorted(relevant_codes),
            })
            counts["entries"] += 1

    log.info(
        "sheets read: sheets_seen=%d entries=%d "
        "rows_no_issue_label=%d rows_filtered_out=%d rows_no_variant_ids=%d",
        len(sheets_seen),
        counts.get("entries", 0),
        counts.get("rows_no_issue_label", 0),
        counts.get("rows_filtered_out", 0),
        counts.get("rows_no_variant_ids", 0),
    )
    return entries


def merge_entries_by_product(
    entries: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """
    Same productId from multiple rows/sheets → merge.

    The merged shape per product:

        {
          "productId":         <id>,
          "productName":       <first non-empty>,
          "productTemplateId": <first non-empty>,
          "variantIds":        [dedup'd, insertion-order],
          "issues":            sorted(set of all codes),
          "sources":           sorted(set of sheet names that contributed),
        }

    Why first-non-empty for name and templateId: sheets sometimes leave
    them blank in later rows after the first mention. We take the first
    populated value and ignore later blanks rather than overwriting with
    an empty.
    """
    log = logging.getLogger("products")
    merged: dict[str, dict[str, Any]] = {}

    for entry in entries:
        pid = entry["productId"]
        if pid not in merged:
            merged[pid] = {
                "productId":         pid,
                "productName":       entry["productName"],
                "productTemplateId": entry["productTemplateId"],
                "variantIds":        list(entry["variantIds"]),
                "issues":            set(entry["issues"]),
                "sources":           {entry["sheet"]},
            }
            continue
        m = merged[pid]
        if not m["productName"] and entry["productName"]:
            m["productName"] = entry["productName"]
        if not m["productTemplateId"] and entry["productTemplateId"]:
            m["productTemplateId"] = entry["productTemplateId"]
        # Append variantIds, preserving order, no duplicates.
        existing_ids = set(m["variantIds"])
        for vid in entry["variantIds"]:
            if vid not in existing_ids:
                existing_ids.add(vid)
                m["variantIds"].append(vid)
        m["issues"].update(entry["issues"])
        m["sources"].add(entry["sheet"])

    # Normalize sets → sorted lists for clean JSON output.
    for m in merged.values():
        m["issues"]  = sorted(m["issues"])
        m["sources"] = sorted(m["sources"])

    log.info(
        "merged entries: rows=%d unique_products=%d",
        len(entries), len(merged),
    )
    return merged


# ---------------------------------------------------------------------------
# Duplicate detection — a variant claimed by 2+ different products is a
# duplicate. The output groups by variant and lists every product claiming it.
# ---------------------------------------------------------------------------

def detect_duplicates(
    merged_products: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """
    Find every variantId claimed by 2+ DIFFERENT products in the merged
    map. Return a dict keyed by variantId:

        {
          "BA-INTWPWK": {
            "variantName": "",  # populated by API enrichment in step 3
            "products": [
              {
                "productId":         "...",
                "productName":       "...",
                "productTemplateId": "...",
                "state":             null,    # populated by API later
                "status":            null,    # populated by API later
                "issues":            ["PARENT_DOES_NOT_LINK_BACK"],
              },
              ...
            ],
            "state":   null,   # variant-level, API-enriched
            "status":  null,   # variant-level, API-enriched
            "siteId":  null,   # variant-level, API-enriched (single site per
                               # API contract for this lookup)
          },
          ...
        }

    The per-product status/state are placeholders here because the sheet
    knows about products but not their PIM state — we need the API for
    that. Step 3 (enrich_duplicates) fills them in.

    Why this matters: when a variant appears on 2+ products, we can't
    safely append it to ANY of those products' variantIds — we don't
    know which one is the "correct" owner. The operator has to decide.
    These products get excluded from the clean repair flow and surfaced
    in the duplicate file for manual review.
    """
    log = logging.getLogger("products")

    # variantId → list of (productId, product_dict) refs
    variant_to_products: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for pid, pdata in merged_products.items():
        for vid in pdata["variantIds"]:
            variant_to_products[vid].append(pdata)

    duplicates: dict[str, dict[str, Any]] = {}
    for vid, prods in variant_to_products.items():
        if len(prods) < 2:
            continue
        # Two different products mean this variant is duplicated. Group
        # by productId so the same product listed twice (shouldn't happen
        # after merge, but defensive) collapses into one entry.
        per_product: dict[str, dict[str, Any]] = {}
        for p in prods:
            pid = p["productId"]
            if pid in per_product:
                continue
            per_product[pid] = {
                "productId":         pid,
                "productName":       p.get("productName", ""),
                "productTemplateId": p.get("productTemplateId", ""),
                "state":             None,   # filled by enrichment
                "status":            None,   # filled by enrichment
                "issues":            list(p.get("issues", [])),
            }
        if len(per_product) < 2:
            # Same product listed twice — not actually a cross-product dup.
            continue
        duplicates[vid] = {
            "variantName": "",
            "products":    list(per_product.values()),
            "state":       None,
            "status":      None,
            "siteId":      None,
        }

    log.info(
        "duplicates detected: %d variants claimed by 2+ products",
        len(duplicates),
    )
    return duplicates


# ---------------------------------------------------------------------------
# HTTP session and fetch helpers — same pattern as repair_variants.py.
# ---------------------------------------------------------------------------

def make_session(
    session_id: str,
    cookie_header: str | None = None,
) -> requests.Session:
    """
    Single session shared across worker threads. The X-UpStart-Site
    header goes per-request (workers may target different sites), so we
    don't put it on the session itself — that would race.
    """
    s = requests.Session()
    headers = {
        "X-Upstart-Tenant":     "denvermattress",
        "Referer":               "https://denvermattress.nochannel-test-1.nochannel-test.upstart.team/",
        "X-Upstart-Session-Id": session_id,
        "Accept":                "application/json, text/plain, */*",
        "Content-Type":          "application/json",
    }
    if cookie_header:
        headers["Cookie"] = cookie_header
    s.headers.update(headers)
    return s


def _is_html_response(r: requests.Response) -> bool:
    """
    Detect the SPA-shell HTML page we get back when the cookie is missing
    or expired. The gateway serves the React app HTML at 200 OK instead
    of returning a 401 — without this check we'd think a fetch "worked"
    and try to parse HTML as JSON.
    """
    ctype = r.headers.get("Content-Type", "").lower()
    if "html" in ctype:
        return True
    body_head = r.text[:200].lstrip()
    return body_head.startswith("<!doctype") or body_head.startswith("<html")


def fetch_variant(
    session: requests.Session,
    api_host: str,
    variant_id: str,
    site_ids: list[str],
    timeout: float = 30.0,
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    """
    GET /pim/variants/<id>?view=...  with multi-site retry.

    Used by step 3 (duplicate enrichment) and step 4 (variant state
    snapshot). Tries LATEST_INCLUDE_DRAFT first, then LATEST, across the
    provided site list. The first successful site/view combo wins.

    Returns (payload, winning_site, error_summary). On total miss:
    (None, None, error_summary).
    """
    main_log = logging.getLogger("products")
    err_log = logging.getLogger("products.err")
    url = f"{api_host.rstrip('/')}/pim/variants/{variant_id}"

    last_err: str | None = None

    for view in ("LATEST_INCLUDE_DRAFT", "LATEST"):
        for site_id in site_ids:
            try:
                r = session.get(
                    url,
                    params={"view": view, "ignoreIfError": "1"},
                    headers={"X-UpStart-Site": site_id},
                    timeout=timeout,
                )
            except requests.RequestException as e:
                last_err = f"network: {e}"
                main_log.debug(
                    "fetch_variant network error variantId=%s site=%s view=%s detail=%r",
                    variant_id, site_id, view, str(e),
                )
                continue

            if r.status_code == 200:
                if _is_html_response(r):
                    err_log.error(
                        "kind=auth_html variantId=%s site=%s "
                        "HINT='cookie missing/expired'",
                        variant_id, site_id,
                    )
                    return None, None, "auth_html"
                try:
                    return r.json(), site_id, None
                except ValueError as e:
                    last_err = f"json_decode: {e}"
                    continue

            if r.status_code == 404:
                main_log.debug(
                    "fetch_variant 404 variantId=%s site=%s view=%s",
                    variant_id, site_id, view,
                )
                continue

            if r.status_code in (401, 403):
                err_log.error(
                    "fetch_variant auth error status=%d variantId=%s site=%s "
                    "HINT='Pass --cookie or set API_COOKIE'",
                    r.status_code, variant_id, site_id,
                )
                return None, None, f"auth http {r.status_code}"

            last_err = f"http {r.status_code}"
            main_log.debug(
                "fetch_variant unexpected status=%d variantId=%s site=%s",
                r.status_code, variant_id, site_id,
            )

    return None, None, last_err


def fetch_product(
    session: requests.Session,
    api_host: str,
    product_id: str,
    site_ids: list[str],
    timeout: float = 30.0,
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    """
    GET /pim/products/<id>?view=...  with multi-site retry.

    Mirrors fetch_variant exactly, just hitting /pim/products/ instead.
    Used in step 6 (repair products).
    """
    main_log = logging.getLogger("products")
    err_log = logging.getLogger("products.err")
    url = f"{api_host.rstrip('/')}/pim/products/{product_id}"

    last_err: str | None = None

    for view in ("LATEST_INCLUDE_DRAFT", "LATEST"):
        for site_id in site_ids:
            try:
                r = session.get(
                    url,
                    params={"view": view, "ignoreIfError": "1"},
                    headers={"X-UpStart-Site": site_id},
                    timeout=timeout,
                )
            except requests.RequestException as e:
                last_err = f"network: {e}"
                main_log.debug(
                    "fetch_product network error productId=%s site=%s view=%s detail=%r",
                    product_id, site_id, view, str(e),
                )
                continue

            if r.status_code == 200:
                if _is_html_response(r):
                    err_log.error(
                        "kind=auth_html productId=%s site=%s "
                        "HINT='cookie missing/expired'",
                        product_id, site_id,
                    )
                    return None, None, "auth_html"
                try:
                    return r.json(), site_id, None
                except ValueError as e:
                    last_err = f"json_decode: {e}"
                    continue

            if r.status_code == 404:
                main_log.debug(
                    "fetch_product 404 productId=%s site=%s view=%s",
                    product_id, site_id, view,
                )
                continue

            if r.status_code in (401, 403):
                err_log.error(
                    "fetch_product auth error status=%d productId=%s site=%s",
                    r.status_code, product_id, site_id,
                )
                return None, None, f"auth http {r.status_code}"

            last_err = f"http {r.status_code}"
            main_log.debug(
                "fetch_product unexpected status=%d productId=%s site=%s",
                r.status_code, product_id, site_id,
            )

    return None, None, last_err


# ---------------------------------------------------------------------------
# Payload repair — append sheet variantIds to payload.variantIds (dedupe).
# ---------------------------------------------------------------------------

def repair_product_payload(
    payload: dict[str, Any],
    product_id: str,
    sheet_variant_ids: list[str],
) -> dict[str, Any]:
    """
    Return a new payload with:
      - sheet_variant_ids appended to payload.variantIds (no duplicates)
      - _STRIP_FROM_PAYLOAD fields removed

    Order policy: existing variantIds keep their original order; new ones
    (from the sheet, not already present) get appended at the end. This
    preserves the platform's stored order and only adds the missing ids
    at the tail, which is the smallest possible change.

    A defensive note on the `id` field: the variant-side script strips
    `id` from variant payloads per direction. We do the same here for
    consistency; if PUT requires `id` for products specifically, this is
    the one knob to flip.
    """
    repaired = dict(payload)

    existing = repaired.get("variantIds") or []
    if not isinstance(existing, list):
        existing = []
    existing_strs = [str(v) for v in existing if v]
    existing_set  = set(existing_strs)

    appended = list(existing_strs)
    for vid in sheet_variant_ids:
        if vid not in existing_set:
            existing_set.add(vid)
            appended.append(vid)

    repaired["variantIds"] = appended

    for k in _STRIP_FROM_PAYLOAD:
        repaired.pop(k, None)

    return repaired


def write_product_payload(
    output_root: Path,
    repaired_payload: dict[str, Any],
    product_id: str,
    fetched_via_site: str,
    fallback_template_id: str | None,
    *,
    site_ids_source: dict[str, Any] | None = None,
) -> tuple[Path, list[str]]:
    """
    Write <output>/<productTemplateId>/<siteIds-joined>/<productId>.json
    and return (path, site_ids list).

    Folder naming is identical to the variant script:
      - comma-joined siteIds from the payload, in ORIGINAL ORDER (not sorted)
      - first site is the primary (X-UpStart-Site for any future push)
      - rest go in ?sites= query

    site_ids_source carries the un-stripped payload so we can read siteIds
    BEFORE repair_product_payload's strip removes it. If None, we look at
    the repaired payload itself (useful for tests).
    """
    tpl_id = (
        repaired_payload.get("productTemplateId")
        or repaired_payload.get("product_template_id")
        or fallback_template_id
        or "unknown"
    )

    source = site_ids_source if site_ids_source is not None else repaired_payload
    raw = source.get("siteIds") or []
    if not isinstance(raw, list):
        raw = []
    site_ids = [str(s) for s in raw if s]
    if not site_ids:
        # Defensive fallback: API should always return siteIds, but in case
        # it doesn't, file goes under the responding site so it's findable.
        site_ids = [f"fallback-{fetched_via_site}"] if fetched_via_site else ["unknown-site"]

    folder_name = ",".join(site_ids)
    folder = output_root / str(tpl_id) / folder_name
    folder.mkdir(parents=True, exist_ok=True)
    out = folder / f"{product_id}.json"
    out.write_text(json.dumps(repaired_payload, indent=2, ensure_ascii=False))
    return out, site_ids


# ---------------------------------------------------------------------------
# Step 3 — enrich duplicates via variant API
# ---------------------------------------------------------------------------

def enrich_duplicates(
    duplicates: dict[str, dict[str, Any]],
    session: requests.Session,
    api_host: str,
    site_ids: list[str],
    workers: int,
) -> None:
    """
    For each duplicate variant, fetch from the variant API and populate:
      - variant-level: variantName, state, status, siteId
      - per-product fields: state, status (the product the variant claims
        to belong to — its parent products in our sheet view aren't the
        same thing as the variant's `productIds`, so we ALSO fetch each
        claiming product to populate per-product state/status)

    Concurrent (per-variant), thread-safe via the shared session.
    Mutates `duplicates` in place.
    """
    log = logging.getLogger("products")
    ok_log = logging.getLogger("products.ok")
    err_log = logging.getLogger("products.err")

    if not duplicates:
        log.info("no duplicates to enrich")
        return

    # Collect every (productId) that's claimed by any duplicate. We'll
    # fetch each product once (a product may claim multiple duplicate
    # variants) and reuse the result.
    claiming_product_ids: set[str] = set()
    for vid, vdata in duplicates.items():
        for p in vdata["products"]:
            claiming_product_ids.add(p["productId"])

    log.info(
        "enriching %d duplicates and %d claiming products via API",
        len(duplicates), len(claiming_product_ids),
    )

    # --- Phase A: fetch all claiming products in parallel ---
    log.info("  phase A: fetching %d claiming products", len(claiming_product_ids))
    product_states: dict[str, dict[str, Any] | None] = {}
    state_lock = threading.Lock()
    progress_a = ProgressReporter(
        log, "  duplicate claiming-products", total=len(claiming_product_ids),
    )

    def _fetch_product_state(pid: str) -> tuple[str, dict[str, Any] | None, str | None]:
        payload, _, err = fetch_product(session, api_host, pid, site_ids)
        return pid, payload, err

    with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
        futs = {pool.submit(_fetch_product_state, pid): pid for pid in claiming_product_ids}
        for fut in as_completed(futs):
            pid, payload, err = fut.result()
            with state_lock:
                product_states[pid] = payload
            if err:
                err_log.warning(
                    "duplicate_product_fetch_failed productId=%s err=%s",
                    pid, err,
                )
            progress_a.tick()

    # --- Phase B: fetch each duplicate variant, populate the file ---
    def _enrich_variant(vid: str) -> None:
        vdata = duplicates[vid]
        payload, winning_site, err = fetch_variant(session, api_host, vid, site_ids)
        if payload is None:
            err_log.warning(
                "duplicate_variant_fetch_failed variantId=%s err=%s",
                vid, err,
            )
            return

        vdata["variantName"] = (
            payload.get("name")
            or payload.get("variantName")
            or vdata.get("variantName", "")
            or ""
        )
        vdata["state"]  = payload.get("state")
        vdata["status"] = payload.get("status")

        # The variant API returns siteIds (plural) — store the FIRST one
        # for the duplicate file (matches the format in the example you
        # shared; if you'd prefer the full list, swap to payload.get(...)).
        raw_sites = payload.get("siteIds") or []
        vdata["siteId"] = raw_sites[0] if raw_sites else winning_site

        # Per-product state/status: pull from the pre-fetched product states.
        for p in vdata["products"]:
            pstate = product_states.get(p["productId"])
            if pstate is None:
                continue
            p["state"]  = pstate.get("state")
            p["status"] = pstate.get("status")

        ok_log.info(
            "verdict=duplicate_enriched variantId=%s state=%s status=%s siteId=%s",
            vid, vdata["state"], vdata["status"], vdata["siteId"],
        )

    log.info("  phase B: fetching %d duplicate variants", len(duplicates))
    progress_b = ProgressReporter(
        log, "  duplicate variants", total=len(duplicates),
    )
    with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
        futs = [pool.submit(_enrich_variant, vid) for vid in duplicates]
        for fut in as_completed(futs):
            fut.result()  # surfaces exceptions; the enricher logs its own errors
            progress_b.tick()

    log.info("duplicate enrichment complete")


# ---------------------------------------------------------------------------
# Step 4 — snapshot variant states (state, status) by API GET
# ---------------------------------------------------------------------------

def snapshot_variant_states(
    variant_ids: list[str],
    session: requests.Session,
    api_host: str,
    site_ids: list[str],
    workers: int,
) -> dict[str, dict[str, Any]]:
    """
    For each unique variantId in the clean product map, GET from the API
    and record:

        {
          variantId: {
            "state":     <NORMAL | ARCHIVED | HIDDEN | null>,
            "status":    <PUBLISHED | DRAFT | null>,
            "fetchedViaSite": <site or null>,
          },
          ...
        }

    This is the inverse of repair_variants.py's product-states snapshot:
    that script knew about products from the sheet and recorded variant
    states; this script knows about products from the sheet but needs to
    record VARIANT states from the API to inform downstream decisions
    (e.g. whether a given variant is fit to be appended to a product).

    Concurrent. Variants that 404 across all sites get state=status=null.
    """
    log = logging.getLogger("products")
    err_log = logging.getLogger("products.err")

    snapshot: dict[str, dict[str, Any]] = {}
    snap_lock = threading.Lock()

    def _snap_one(vid: str) -> None:
        payload, winning_site, err = fetch_variant(session, api_host, vid, site_ids)
        with snap_lock:
            if payload is None:
                snapshot[vid] = {"state": None, "status": None, "fetchedViaSite": None}
                if err and err != "auth_html":
                    # auth_html already logged with HINT, no need to re-log here.
                    err_log.warning(
                        "variant_state_snapshot_failed variantId=%s err=%s",
                        vid, err,
                    )
                return
            snapshot[vid] = {
                "state":          payload.get("state"),
                "status":         payload.get("status"),
                "fetchedViaSite": winning_site,
            }

    log.info("snapshotting state+status for %d unique variants", len(variant_ids))
    progress = ProgressReporter(log, "variant snapshot", total=len(variant_ids))
    with ThreadPoolExecutor(max_workers=max(1, workers)) as pool:
        futs = [pool.submit(_snap_one, vid) for vid in variant_ids]
        for fut in as_completed(futs):
            fut.result()
            progress.tick()

    # Summarize.
    by_state = Counter((v.get("state") or "null") for v in snapshot.values())
    log.info("variant state snapshot complete: %s",
             dict(by_state.most_common()))
    return snapshot


# ---------------------------------------------------------------------------
# Step 5/6 — clean map + per-product repair
# ---------------------------------------------------------------------------

def build_clean_product_map(
    merged_products: dict[str, dict[str, Any]],
    duplicates: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """
    Build the clean product_map.json — products whose variantIds (from
    the sheets) contain NO duplicate variants.

    Why exclude products with any duplicate variant: we can't safely
    repair them either, because we don't know whether the duplicate
    really belongs to this product. Sending that variant in this product's
    variantIds list could entrench an incorrect link.

    Placeholders for status/state/siteIds/fetchedViaSite get filled by
    step 6 from the API response.
    """
    log = logging.getLogger("products")
    duplicate_variant_ids = set(duplicates.keys())

    clean: dict[str, dict[str, Any]] = {}
    skipped = 0
    for pid, pdata in merged_products.items():
        if any(vid in duplicate_variant_ids for vid in pdata["variantIds"]):
            skipped += 1
            continue
        clean[pid] = {
            "variantIds":        list(pdata["variantIds"]),
            "productName":       pdata.get("productName", ""),
            "productTemplateId": pdata.get("productTemplateId", ""),
            "status":            None,
            "state":             None,
            "siteIds":           None,
            "fetchedViaSite":    None,
            "issues":            list(pdata.get("issues", [])),
        }

    log.info(
        "clean product map built: %d products (skipped %d that contain duplicate variants)",
        len(clean), skipped,
    )
    return clean


def process_product(
    product_id: str,
    meta: dict[str, Any],
    session: requests.Session,
    api_host: str,
    site_ids: list[str],
    output_root: Path,
) -> dict[str, Any]:
    """
    Fetch one product, repair its payload, write the file. Returns a dict
    the main loop merges into product_map.json:

      {"productId": ..., "ok": bool, "state": ..., "status": ...,
       "siteIds": [...], "fetchedViaSite": ..., "path": ..., "reason": ...}
    """
    ok_log = logging.getLogger("products.ok")
    err_log = logging.getLogger("products.err")

    payload, winning_site, err = fetch_product(session, api_host, product_id, site_ids)
    if payload is None:
        err_log.error(
            "verdict=fetch_failed productId=%s reason=%s",
            product_id, err,
        )
        return {
            "productId":      product_id,
            "ok":             False,
            "state":          None,
            "status":         None,
            "siteIds":        None,
            "fetchedViaSite": None,
            "path":           None,
            "reason":         err or "unknown",
        }

    try:
        repaired = repair_product_payload(payload, product_id, meta["variantIds"])
        path, file_site_ids = write_product_payload(
            output_root,
            repaired,
            product_id,
            fetched_via_site=winning_site,
            fallback_template_id=meta.get("productTemplateId"),
            site_ids_source=payload,
        )
    except Exception as e:  # noqa: BLE001 - want to surface ANY write issue
        err_log.error(
            "verdict=write_failed productId=%s reason=%r",
            product_id, str(e),
        )
        return {
            "productId":      product_id,
            "ok":             False,
            "state":          payload.get("state"),
            "status":         payload.get("status"),
            "siteIds":        None,
            "fetchedViaSite": winning_site,
            "path":           None,
            "reason":         f"write: {e}",
        }

    final_state  = payload.get("state")
    final_status = payload.get("status")
    ok_log.info(
        "verdict=ok productId=%s siteIds=%s fetchedVia=%s state=%s status=%s path=%s",
        product_id, file_site_ids, winning_site, final_state, final_status, path,
    )
    return {
        "productId":      product_id,
        "ok":             True,
        "state":          final_state,
        "status":         final_status,
        "siteIds":        file_site_ids,
        "fetchedViaSite": winning_site,
        "path":           str(path),
        "reason":         None,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--input", required=True, type=Path,
                    help="Path to report-linkages.xlsx")
    ap.add_argument("--output", required=True, type=Path,
                    help="Output dir for product_map.json, duplicates, payloads")
    ap.add_argument(
        "--api-host",
        default="http://localhost:8080",
    )
    ap.add_argument("--site-id", default=PRIMARY_SITE_ID)
    ap.add_argument(
        "--fallback-site-ids",
        default=",".join(FALLBACK_SITE_IDS),
        help="Comma-separated fallback sites",
    )
    ap.add_argument(
        "--session-id",
        default=os.environ.get("API_SESSION_ID", "example"),
    )
    ap.add_argument("--cookie", default=os.environ.get("API_COOKIE", ""))
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--limit", type=int, default=0,
                    help="Process first N clean products (0=all). For testing.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Skip API calls; sheet+map work only.")
    ap.add_argument("--map-only", action="store_true",
                    help="Stop after step 5; don't fetch+repair products.")
    ap.add_argument("--skip-duplicate-enrichment", action="store_true",
                    help="Skip step 3 (duplicate API enrichment).")
    ap.add_argument("--log-dir", type=Path, default=None,
                    help="Where logs go. Default: <output>/logs")
    ap.add_argument("--verbose", "-v", action="store_true")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"Input not found: {args.input}", file=sys.stderr)
        return 1

    args.output.mkdir(parents=True, exist_ok=True)
    log_dir = args.log_dir or (args.output / "logs")
    log = setup_logging(log_dir, verbose=args.verbose)

    run_id = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    log.info(
        "run_start run_id=%s input=%s output=%s dry_run=%s map_only=%s",
        run_id, args.input, args.output, args.dry_run, args.map_only,
    )

    # --- Step 1: read sheets ---
    log.info("[1/6] read sheets")
    entries = read_sheets(args.input)
    if not entries:
        log.warning("no PARENT_DOES_NOT_LINK_BACK product entries found — nothing to do")
        return 0

    # --- Step 2: merge + detect duplicates + write placeholder file ---
    log.info("[2/6] merge entries and detect duplicates")
    merged = merge_entries_by_product(entries)
    duplicates = detect_duplicates(merged)
    dup_path = args.output / "duplicate_variants_across_product.json"
    dup_path.write_text(json.dumps(duplicates, indent=2, ensure_ascii=False))
    log.info("wrote %s (%d duplicate variants)", dup_path, len(duplicates))

    # --- Build the site retry list ---
    raw_fallbacks = [s.strip() for s in (args.fallback_site_ids or "").split(",")]
    site_ids: list[str] = []
    for s in [args.site_id, *raw_fallbacks]:
        if s and s not in site_ids:
            site_ids.append(s)
    if not site_ids:
        log.error("no site IDs configured")
        return 1
    log.info("siteIds (in retry order): %s", site_ids)

    if not args.cookie and not args.dry_run:
        log.warning(
            "no Cookie provided — if the gateway requires session cookies "
            "you'll get HTML responses. Pass --cookie or API_COOKIE env."
        )

    session = make_session(args.session_id, cookie_header=args.cookie or None)

    # --- Step 3: enrich duplicates ---
    if args.dry_run:
        log.info("[3/6] skip duplicate enrichment (dry-run)")
    elif args.skip_duplicate_enrichment:
        log.info("[3/6] skip duplicate enrichment (--skip-duplicate-enrichment)")
    else:
        log.info("[3/6] enrich duplicates via API")
        enrich_duplicates(duplicates, session, args.api_host, site_ids, args.workers)
        dup_path.write_text(json.dumps(duplicates, indent=2, ensure_ascii=False))
        log.info("rewrote %s with API-enriched duplicate data", dup_path)

    # --- Step 5 (early): build the clean map so step 4 knows what variants
    #     to snapshot. The clean map is the universe of "products we will
    #     actually repair"; the variants under those products are the
    #     universe of "variants we should record state for". ---
    log.info("[5/6] build clean product map (excluding products with duplicate variants)")
    product_map = build_clean_product_map(merged, duplicates)
    map_path = args.output / "product_map.json"
    map_path.write_text(json.dumps(product_map, indent=2, ensure_ascii=False))
    log.info("wrote %s (%d clean products)", map_path, len(product_map))

    # --- Step 4: snapshot variant states (skip on dry-run) ---
    unique_variants = sorted({
        vid for meta in product_map.values() for vid in meta["variantIds"]
    })
    snap_path = args.output / "variant_states_pre_repair.json"
    if args.dry_run:
        log.info("[4/6] skip variant state snapshot (dry-run)")
        snap_path.write_text(json.dumps({}, indent=2))
    else:
        log.info("[4/6] snapshot variant states (state, status) via API")
        snapshot = snapshot_variant_states(
            unique_variants, session, args.api_host, site_ids, args.workers,
        )
        snap_path.write_text(json.dumps(snapshot, indent=2, ensure_ascii=False))
        log.info("wrote %s (%d variants)", snap_path, len(snapshot))

    if args.map_only:
        log.info("--map-only: stopping before fetch+repair phase")
        return 0
    if args.dry_run:
        log.info("dry-run: stopping before fetch+repair phase")
        return 0

    # --- Step 6: fetch + repair + write payloads ---
    pids = list(product_map.keys())
    if args.limit > 0:
        pids = pids[: args.limit]
        log.info("--limit %d: processing first %d products", args.limit, len(pids))

    log.info("[6/6] fetch + repair + write payloads (%d products, %d workers)",
             len(pids), args.workers)

    t0 = time.time()
    n_ok = n_fail = 0
    results_lock = threading.Lock()

    def _do_one(pid: str) -> dict[str, Any]:
        return process_product(
            pid, product_map[pid], session, args.api_host, site_ids, args.output,
        )

    progress = ProgressReporter(log, "fetch+repair", total=len(pids))
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        futs = {pool.submit(_do_one, pid): pid for pid in pids}
        for fut in as_completed(futs):
            result = fut.result()
            pid = result["productId"]
            with results_lock:
                if result["ok"]:
                    n_ok += 1
                    product_map[pid]["status"]         = result["status"]
                    product_map[pid]["state"]          = result["state"]
                    product_map[pid]["siteIds"]        = result["siteIds"]
                    product_map[pid]["fetchedViaSite"] = result["fetchedViaSite"]
                else:
                    n_fail += 1
            progress.tick(extra=f"ok={n_ok} fail={n_fail}")

    elapsed = time.time() - t0
    log.info(
        "fetch+repair done: ok=%d fail=%d elapsed_s=%.1f",
        n_ok, n_fail, elapsed,
    )

    # --- Final rewrite of product_map.json with API truth ---
    map_path.write_text(json.dumps(product_map, indent=2, ensure_ascii=False))
    log.info("rewrote %s with API-enriched product data", map_path)

    log.info("run_end run_id=%s success=%s", run_id, n_fail == 0)
    return 0 if n_fail == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())